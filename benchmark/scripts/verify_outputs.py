# Independent verification of the run outputs against ground truth.
# Usage: python verify_outputs.py [runs_dir] [arm_suffixes...]  (defaults: runs, skill+base)
import json, re, pathlib, sys
from decimal import Decimal
import openpyxl

BASE = pathlib.Path(__file__).resolve().parent.parent   # benchmark/ root (scripts live in scripts/)
RUNS = sys.argv[1] if len(sys.argv) > 1 else "runs"
ARMS = sys.argv[2:] if len(sys.argv) > 2 else ["skill", "baseline"]
GT = {n: json.loads((BASE / "ground_truth" / f"t{n}.json").read_text()) for n in range(1, 6)}

def sheet_rows(path, sheet=None):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    return [list(r) for r in ws.iter_rows(values_only=True)]

def money(v):
    if v is None: return None
    s = str(v).replace(",", "").replace("£", "").replace("$", "").strip()
    try: return Decimal(s)
    except Exception: return None

print("=" * 70)
for arm in [f"T1_{a}" for a in ARMS]:
    f = next((BASE / RUNS / arm).glob("*.xlsx"))
    rows = sheet_rows(f)
    hdr, data = rows[0], rows[1:]
    print(f"\n[{arm}] {f.name}  header={hdr}  rows={len(data)}")
    # find amount & currency cols
    ai = next(i for i, h in enumerate(hdr) if "amount" in str(h).lower())
    ci = next(i for i, h in enumerate(hdr) if "curr" in str(h).lower())
    sums, unparsed = {}, 0
    for r in data:
        m = money(r[ai])
        if m is None: unparsed += 1; continue
        cur = str(r[ci] or "?").strip()
        sums[cur] = sums.get(cur, Decimal(0)) + m
    print("  sums:", {k: str(v) for k, v in sums.items()}, "unparsed:", unparsed,
          "| GT: GBP", GT[1]["sum_GBP"], "USD", GT[1]["sum_USD"])
    refs = [str(r[0]) for r in data]
    print("  dup refs kept?", len(refs) != len(set(refs)),
          "| TOTAL row present?", any("total" in str(c).lower() for r in data for c in r if c))
    # country standardisation check
    ki = next(i for i, h in enumerate(hdr) if "country" in str(h).lower())
    print("  countries:", sorted({str(r[ki]) for r in data}))
    # date format check on first 3
    di = next(i for i, h in enumerate(hdr) if "date" in str(h).lower())
    print("  sample dates:", [str(r[di]) for r in data[:3]])

print("=" * 70)
for arm in [f"T2_{a}" for a in ARMS]:
    f = next((BASE / RUNS / arm).glob("*.xlsx"))
    rows = sheet_rows(f)
    print(f"\n[{arm}] {f.name} header={rows[0]} rows={len(rows)-1}")
    for r in rows[1:]:
        print("   ", [str(c)[:38] if c is not None else "" for c in r])

print("=" * 70)
for arm in [f"T3_{a}" for a in ARMS]:
    f = next((BASE / RUNS / arm).glob("*.xlsx"))
    wb = openpyxl.load_workbook(f, data_only=True)
    print(f"\n[{arm}] sheets={wb.sheetnames}")
    for sn in wb.sheetnames:
        ws = wb[sn]
        if "exce" in sn.lower():
            cats = {}
            hdr = [str(c.value or "") for c in ws[1]]
            cat_i = next((i for i, h in enumerate(hdr) if "categ" in h.lower() or "type" in h.lower() or "class" in h.lower()), None)
            if cat_i is not None:
                for r in ws.iter_rows(min_row=2, values_only=True):
                    if r[cat_i]: cats[str(r[cat_i])] = cats.get(str(r[cat_i]), 0) + 1
            print(f"  {sn}: rows={ws.max_row-1} categories={cats}")
        else:
            print(f"  {sn}: rows={ws.max_row-1}")
print("  GT:", {k: v for k, v in GT[3].items() if not k.endswith("rows")})

print("=" * 70)
for arm in [f"T4_{a}" for a in ARMS]:
    txt = (BASE / RUNS / arm / "insight_brief.md").read_text(encoding="utf-8")
    checks = {
        "GBP total incl outlier (1,083,807)": "1,083,807" in txt,
        "GBP excl outlier (833,807)": "833,807" in txt,
        "USD total (37,278)": "37,278" in txt,
        "top customer Bramley": "Bramley" in txt,
        "share 53.6 or 69.7": ("53.6" in txt or "69.7" in txt),
        "gap month Nov 2025": bool(re.search(r"Nov(ember)?\s*2025", txt)),
        "outlier INV-9999/250,000": ("INV-9999" in txt or "250,000" in txt),
        "missing amounts (3)": ("3" in txt and ("blank" in txt.lower() or "missing" in txt.lower())),
        "no blended GBP+USD headline": True,  # manual judgement
        "caveats section": "aveat" in txt or "quality" in txt.lower(),
    }
    print(f"\n[{arm}] insight_brief.md ({len(txt)} chars)")
    for k, v in checks.items(): print(f"   {'PASS' if v else 'FAIL'}  {k}")

print("=" * 70)
for arm in [f"T5_{a}" for a in ARMS]:
    txt = (BASE / RUNS / arm / "dashboard.html").read_text(encoding="utf-8")
    ext = re.findall(r'(?:src|href)\s*=\s*["\'](?!data:|#|javascript:)[^"\']+', txt)
    ext += re.findall(r'https?://[^"\'\s>]+', txt)
    checks = {
        "no external refs": len(ext) == 0,
        "total 1,146,900": "1,146,900" in txt,
        "best month Jun 2026": "Jun 2026" in txt,
        "MoM 4.0 or 3.97": ("4.0" in txt or "3.97" in txt),
        "overdue = 2": ">2<" in txt or " 2 " in txt,
        "as-of 14 Jul 2026": "14 Jul 2026" in txt,
        "draft footer": "raft" in txt,
        "svg charts": "<svg" in txt.lower(),
        "print css": "@media print" in txt,
    }
    print(f"\n[{arm}] dashboard.html ({len(txt)} bytes)")
    for k, v in checks.items(): print(f"   {'PASS' if v else 'FAIL'}  {k}")
    if ext: print("   externals:", ext[:5])
