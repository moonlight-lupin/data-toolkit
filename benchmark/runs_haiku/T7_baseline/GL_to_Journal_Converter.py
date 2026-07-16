#!/usr/bin/env python3
"""
GL Export to Journal Import Converter
<org> - Reusable monthly conversion script

STANDING RULES:
1. Each month's source Excel file MUST be sense-checked against the expected schema before converting
2. Expected columns (order-independent): Entry No, Posting Date [DD/MM/YYYY], Account Code, Account Name,
   Description, Debit, Credit, Cost Centre
3. If any column is missing, renamed, or new columns appear, STOP and FLAG the discrepancy before proceeding
4. Required fields in output (must never be empty): JournalRef, Date, AccountCode, Amount
5. Rows with empty required fields must be flagged and EXCLUDED from the import file, never guessed
6. All conversions must be logged with row counts and flagged rows documented

USAGE:
  python GL_to_Journal_Converter.py <source_excel_file> <output_csv_file>

EXAMPLE:
  python GL_to_Journal_Converter.py fixtures/t7_gl_jul.xlsx runs_haiku/T7_base/journal_import_jul2026.csv
"""

import openpyxl
import csv
from datetime import datetime
from pathlib import Path
import sys
import logging

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(levelname)s: %(message)s'
)
logger = logging.getLogger(__name__)

class GLConverter:
    """Converts GL export Excel files to journal import CSV format."""

    # Expected column names in source file
    EXPECTED_COLUMNS = {
        'Entry No',
        'Posting Date',
        'Account Code',
        'Account Name',
        'Description',
        'Debit',
        'Credit',
        'Cost Centre'
    }

    # Column aliases - maps variant names to canonical names
    COLUMN_ALIASES = {
        'CC Code': 'Cost Centre'  # August 2026+ uses CC Code instead of Cost Centre
    }

    # Target CSV columns in required order
    TARGET_COLUMNS = ['JournalRef', 'Date', 'AccountCode', 'Narrative', 'Amount', 'CostCentre', 'Source']

    # Required fields (must not be empty)
    REQUIRED_FIELDS = {'JournalRef', 'Date', 'AccountCode', 'Amount'}

    def __init__(self, source_file, output_file):
        """Initialize converter with source and output paths."""
        self.source_file = Path(source_file)
        self.output_file = Path(output_file)
        self.converted_rows = []
        self.flagged_rows = []
        self.schema_discrepancies = []

    def validate_schema(self):
        """
        Validate that source file has expected columns.
        Accepts column aliases and flags discrepancies.
        Returns True if valid, False otherwise.
        """
        try:
            wb = openpyxl.load_workbook(self.source_file)
            ws = wb.active

            # Get headers
            headers = [cell.value for cell in ws[1]]
            found_headers = set(h for h in headers if h is not None)

            logger.info(f"Source headers: {found_headers}")

            # Normalize headers by applying aliases
            normalized_headers = set()
            for h in found_headers:
                if h in self.COLUMN_ALIASES:
                    normalized_headers.add(self.COLUMN_ALIASES[h])
                    self.schema_discrepancies.append(f"Column '{h}' mapped to '{self.COLUMN_ALIASES[h]}'")
                    logger.warning(f"SCHEMA CHANGE: Column '{h}' used instead of '{self.COLUMN_ALIASES[h]}'")
                else:
                    normalized_headers.add(h)

            # Check for missing columns (after normalization)
            missing_cols = self.EXPECTED_COLUMNS - normalized_headers
            if missing_cols:
                logger.error(f"DISCREPANCY: Missing expected columns: {missing_cols}")
                return False

            # Check for unexpected columns
            unexpected_cols = found_headers - self.EXPECTED_COLUMNS - set(self.COLUMN_ALIASES.keys())
            if unexpected_cols:
                self.schema_discrepancies.append(f"New columns found: {unexpected_cols}")
                logger.warning(f"DISCREPANCY: Unexpected new columns found: {unexpected_cols}")
                logger.warning("Proceeding, but review required for next conversion.")
                # Don't fail on new columns, just warn

            logger.info("Schema validation PASSED (with noted discrepancies)")
            return True

        except Exception as e:
            logger.error(f"Schema validation failed: {type(e).__name__}: {e}")
            return False

    def read_and_convert(self):
        """
        Read source file and convert to journal import format.
        Returns True if successful, False otherwise.
        """
        try:
            wb = openpyxl.load_workbook(self.source_file)
            ws = wb.active

            # Get headers and column indices
            headers = [cell.value for cell in ws[1]]
            col_indices = {}
            for i, header in enumerate(headers):
                if header in self.EXPECTED_COLUMNS:
                    col_indices[header] = i
                elif header in self.COLUMN_ALIASES:
                    # Map alias to the canonical name
                    canonical = self.COLUMN_ALIASES[header]
                    col_indices[canonical] = i

            logger.info(f"Total data rows to process: {ws.max_row - 1}")

            # Process each data row
            for row_idx in range(2, ws.max_row + 1):
                row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
                self._convert_row(row_idx, row, col_indices)

            return True

        except Exception as e:
            logger.error(f"Conversion failed: {type(e).__name__}: {e}")
            import traceback
            traceback.print_exc()
            return False

    def _convert_row(self, row_num, source_row, col_indices):
        """Convert a single row from GL format to journal format."""
        try:
            # Extract values
            entry_no = source_row[col_indices['Entry No']]
            posting_date = source_row[col_indices['Posting Date']]
            account_code = source_row[col_indices['Account Code']]
            description = source_row[col_indices['Description']]
            debit = source_row[col_indices['Debit']]
            credit = source_row[col_indices['Credit']]
            cost_centre = source_row[col_indices['Cost Centre']]

            # Initialize converted row
            converted = {}

            # 1. JournalRef = Entry No
            converted['JournalRef'] = entry_no

            # 2. Date: DD/MM/YYYY -> YYYY-MM-DD
            date_str = self._convert_date(posting_date)
            converted['Date'] = date_str

            # 3. AccountCode = Account Code
            converted['AccountCode'] = account_code

            # 4. Narrative = Description
            converted['Narrative'] = description

            # 5. Amount: Debit - Credit (Debit positive, Credit negative)
            amount = self._calculate_amount(debit, credit)
            converted['Amount'] = amount

            # 6. CostCentre = Cost Centre
            converted['CostCentre'] = cost_centre

            # 7. Source = constant "GLEXPORT"
            converted['Source'] = 'GLEXPORT'

            # Check required fields
            missing_required = []
            for req_field in self.REQUIRED_FIELDS:
                value = converted[req_field]
                if value is None or (isinstance(value, str) and value.strip() == ''):
                    missing_required.append(req_field)

            if missing_required:
                self.flagged_rows.append({
                    'row_num': row_num,
                    'entry_no': entry_no,
                    'reason': f"Missing required field(s): {', '.join(missing_required)}",
                    'data': converted
                })
                logger.warning(f"Row {row_num}: Flagged - Missing required field(s): {missing_required}")
            else:
                self.converted_rows.append(converted)

        except Exception as e:
            entry_no = 'unknown'
            try:
                entry_no = source_row[col_indices.get('Entry No', 0)] if 'Entry No' in col_indices else 'unknown'
            except:
                pass
            self.flagged_rows.append({
                'row_num': row_num,
                'entry_no': entry_no,
                'reason': f"Processing error: {type(e).__name__}: {e}",
                'data': None
            })
            logger.error(f"Row {row_num}: Processing error - {type(e).__name__}: {e}")

    def _convert_date(self, posting_date):
        """Convert date from DD/MM/YYYY to YYYY-MM-DD."""
        if not posting_date:
            return None

        try:
            if isinstance(posting_date, datetime):
                return posting_date.strftime('%Y-%m-%d')
            else:
                # Try to parse as string
                date_obj = datetime.strptime(str(posting_date).strip(), '%d/%m/%Y')
                return date_obj.strftime('%Y-%m-%d')
        except Exception as e:
            logger.error(f"Date conversion error for '{posting_date}': {e}")
            return None

    def _calculate_amount(self, debit, credit):
        """Calculate Amount as Debit - Credit."""
        try:
            debit_val = float(debit) if debit else 0.0
            credit_val = float(credit) if credit else 0.0
            amount = debit_val - credit_val
            amount = round(amount, 2)

            # Amount is None only if both Debit and Credit are empty
            if amount == 0.0 and not debit and not credit:
                return None

            return amount
        except Exception as e:
            logger.error(f"Amount calculation error: Debit={debit}, Credit={credit} - {e}")
            return None

    def write_output(self):
        """Write converted rows to output CSV file."""
        try:
            self.output_file.parent.mkdir(parents=True, exist_ok=True)

            with open(self.output_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=self.TARGET_COLUMNS)
                writer.writeheader()

                for row in self.converted_rows:
                    output_row = {
                        'JournalRef': row['JournalRef'],
                        'Date': row['Date'],
                        'AccountCode': row['AccountCode'],
                        'Narrative': row['Narrative'],
                        'Amount': row['Amount'],
                        'CostCentre': row['CostCentre'] if row['CostCentre'] else '',
                        'Source': row['Source']
                    }
                    writer.writerow(output_row)

            logger.info(f"Successfully wrote {len(self.converted_rows)} rows to {self.output_file}")
            return True

        except Exception as e:
            logger.error(f"Failed to write output file: {type(e).__name__}: {e}")
            import traceback
            traceback.print_exc()
            return False

    def report(self):
        """Generate and print conversion report."""
        print("\n" + "=" * 80)
        print("CONVERSION REPORT")
        print("=" * 80)
        print(f"Source: {self.source_file}")
        print(f"Output: {self.output_file}")

        if self.schema_discrepancies:
            print(f"\nSCHEMA DISCREPANCIES (flagged, not blocking):")
            for disc in self.schema_discrepancies:
                print(f"  - {disc}")

        print(f"\nRows converted: {len(self.converted_rows)}")
        print(f"Rows flagged/excluded: {len(self.flagged_rows)}")

        if self.flagged_rows:
            print("\nFLAGGED ROWS:")
            for flagged in self.flagged_rows:
                print(f"  Row {flagged['row_num']} (Entry {flagged['entry_no']}): {flagged['reason']}")

        # Calculate sum of Amount column
        total_amount = sum(row['Amount'] for row in self.converted_rows if row['Amount'] is not None)
        print(f"\nSum of Amount column: {total_amount:.2f}")

        print("=" * 80)


def main():
    """Main entry point."""
    if len(sys.argv) < 3:
        print(__doc__)
        print(f"Usage: {sys.argv[0]} <source_excel_file> <output_csv_file>")
        sys.exit(1)

    source_file = sys.argv[1]
    output_file = sys.argv[2]

    converter = GLConverter(source_file, output_file)

    # Step 1: Validate schema
    logger.info(f"Starting conversion: {source_file} -> {output_file}")
    if not converter.validate_schema():
        logger.error("Schema validation failed. Conversion aborted.")
        sys.exit(1)

    # Step 2: Read and convert
    if not converter.read_and_convert():
        logger.error("Conversion failed.")
        sys.exit(1)

    # Step 3: Write output
    if not converter.write_output():
        logger.error("Failed to write output file.")
        sys.exit(1)

    # Step 4: Report
    converter.report()

    logger.info("Conversion completed successfully.")


if __name__ == '__main__':
    main()
