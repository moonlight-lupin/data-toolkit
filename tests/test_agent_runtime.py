#!/usr/bin/env python3
from __future__ import annotations

import copy
import json
import subprocess
import sys
import tempfile
from decimal import Decimal
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

import agent_runtime as ar


APPROVAL_KEY = "operator-test-key"


def signed_receipt(result, decision):
    request = result["approvals_required"][0]
    return ar.sign_approval_receipt(
        request, decision, approved_by="Test Reviewer", key=APPROVAL_KEY,
        issued_at="2026-07-15T08:00:00+00:00",
    )


def check(name, fn):
    try:
        fn()
        print(f"PASS {name}")
        return True
    except Exception as exc:
        print(f"FAIL {name}: {exc}")
        return False


def test_json_ingest():
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "rows.json"
        p.write_text(json.dumps([{"id": 1, "party": {"name": "A"}}, {"id": 2, "party": {"name": "B"}}]), encoding="utf-8")
        table = ar.read_table(str(p))
        assert table.header == ["id", "party"]
        assert table.rows[0]["party"]["name"] == "A"
        assert "2 record" in table.note


def test_json_path_and_auto_list():
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "wrapped.json"
        p.write_text(json.dumps({"meta": {"as_of": "2026-07-15"}, "records": [{"x": 1}, {"x": 2}]}), encoding="utf-8")
        explicit = ar.read_table({"path": str(p), "json_path": "records"})
        auto = ar.read_table(str(p))
        assert len(explicit.rows) == len(auto.rows) == 2
        assert "auto-selected" in auto.note


def test_plan_validation_approval():
    plan = {
        "version": 1,
        "skill": "data-convert",
        "inputs": ["source.csv"],
        "spec": {"target": {}},
        "output": "out.csv",
    }
    result = ar.validate_plan(plan, check_sources=False)
    assert result["status"] == "needs_approval"
    assert result["approvals_required"][0]["kind"] == "plan_confirmation"
    plan["approval"] = {"confirmed": True}
    result = ar.validate_plan(plan, check_sources=False)
    assert result["status"] == "success"


def test_all_skill_plan_shapes_validate():
    common = {"version": 1, "input": "source.csv", "output": "out.bin", "approval": {"confirmed": True}}
    plans = [
        {**common, "skill": "data-tidy", "recipe": {"columns": []}},
        {**common, "skill": "data-extract", "mode": "fields", "fields": [{"name": "X", "labels": ["x"]}]},
        {"version": 1, "skill": "data-reconcile", "inputs": ["a.csv", "b.csv"], "output": "r.xlsx", "approval": {"confirmed": True}},
        {**common, "skill": "data-analyse", "operations": []},
        {"version": 1, "skill": "data-visualise", "inputs": [], "output": "d.html", "dashboard": {"title": "D", "blocks": []}},
        {**common, "skill": "data-convert", "spec": {"target": {}}},
    ]
    for plan in plans:
        result = ar.validate_plan(plan, check_sources=False)
        assert not result["errors"], (plan["skill"], result)


def test_convert_json_flatten_union_and_nest():
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        a = td / "a.json"
        b = td / "b.json"
        a.write_text(json.dumps([{"id": 1, "party": {"name": "A"}, "amount": 10}]), encoding="utf-8")
        b.write_text(json.dumps([{"id": 2, "party": {"name": "B"}, "amount": 20, "extra": "x"}]), encoding="utf-8")
        out = td / "flat.csv"
        plan = {
            "version": 1,
            "skill": "data-convert",
            "inputs": [str(a), str(b)],
            "spec": {
                "reshape": [{"op": "union", "how": "outer"}, {"op": "flatten"}],
                "target": {"format": "csv"},
            },
            "output": str(out),
            "approval": {"confirmed": True},
        }
        result = ar.run_plan(plan, base_dir=td)
        assert result["status"] == "success", result
        text = out.read_text(encoding="utf-8")
        assert "party.name" in text and "extra" in text and "A" in text and "B" in text
        assert result["metrics"]["sources"] == 2

        source = td / "lines.csv"
        source.write_text("id,sku,amount\n1,A,10\n1,B,20\n", encoding="utf-8")
        nested_out = td / "nested.csv"
        nest_plan = {
            "version": 1,
            "skill": "data-convert",
            "inputs": [str(source)],
            "spec": {
                "reshape": [{"op": "nest", "key_cols": ["id"], "into": "lines", "child_cols": ["sku", "amount"]}],
                "target": {"format": "json"},
            },
            "output": str(nested_out),
            "approval": {"confirmed": True},
        }
        nested_result = ar.run_plan(nest_plan, base_dir=td)
        written = Path(nested_result["artifacts"][0]["path"])
        assert written.suffix == ".json" and written.exists(), nested_result
        payload = json.loads(written.read_text(encoding="utf-8"))
        assert len(payload) == 1 and len(payload[0]["lines"]) == 2


def test_multiple_conversion_inputs_require_explicit_union():
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        a = td / "a.json"
        b = td / "b.json"
        a.write_text('[{"x":1}]', encoding="utf-8")
        b.write_text('[{"x":2}]', encoding="utf-8")
        out = td / "new-output-dir" / "out.csv"
        plan = {
            "version": 1, "skill": "data-convert",
            "inputs": [str(a), str(b)],
            "spec": {"target": {"format": "csv"}},
            "output": str(out), "approval": {"confirmed": True},
        }
        result = ar.run_plan(plan, base_dir=td)
        assert result["status"] == "error", result
        assert "explicit union" in result["errors"][0]
        assert not out.exists()


def test_extract_fields_validation_does_not_parse_as_table():
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        document = td / "form.pdf"
        document.write_bytes(b"not a real PDF; validation should only check existence")
        plan = {
            "version": 1, "skill": "data-extract", "input": str(document),
            "mode": "fields", "fields": [{"name": "Name", "labels": ["name"]}],
            "output": str(td / "out.xlsx"), "approval": {"confirmed": True},
        }
        result = ar.validate_plan(plan, base_dir=td)
        assert not result["errors"], result


def test_analysis_zero_total_is_preserved():
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        src = td / "zero.json"
        src.write_text(json.dumps([{"group": "A", "amount": 10}, {"group": "B", "amount": -10}]), encoding="utf-8")
        out = td / "analysis.json"
        plan = {
            "version": 1,
            "skill": "data-analyse",
            "input": str(src),
            "operations": [{"op": "breakdown", "by": "group", "value": "amount"}],
            "output": str(out),
            "approval": {"confirmed": True},
        }
        result = ar.run_plan(plan, base_dir=td)
        assert result["status"] == "success", result
        breakdown = result["details"]["results"][0]["result"]
        assert breakdown["grand_total"] == "0", breakdown
        assert breakdown["top1_share"] is None and breakdown["top3_share"] is None
        assert all(g["share"] is None for g in breakdown["groups"])


def test_tidy_json_runtime():
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        src = td / "messy.json"
        src.write_text(json.dumps([{"Name": " Alice ", "Amount": "1,200.00"}]), encoding="utf-8")
        out = td / "clean.xlsx"
        plan = {
            "version": 1,
            "skill": "data-tidy",
            "input": str(src),
            "recipe": {"columns": [
                {"source": "Name", "target": "Name", "type": "text"},
                {"source": "Amount", "target": "Amount", "type": "number"}
            ]},
            "output": str(out),
            "approval": {"confirmed": True},
        }
        result = ar.run_plan(plan, base_dir=td)
        assert result["status"] in ("success", "success_with_warnings"), result
        assert out.exists() and out.with_suffix(".report.md").exists()


def test_reconcile_json_runtime():
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        a = td / "a.json"
        b = td / "b.json"
        a.write_text(json.dumps([{"invoice_no": "INV-1", "amount": "100", "date": "01 Jul 2026"}]), encoding="utf-8")
        b.write_text(json.dumps([{"invoice_no": "INV-1", "amount": "100", "date": "01 Jul 2026"}]), encoding="utf-8")
        out = td / "recon.xlsx"
        plan = {
            "version": 1,
            "skill": "data-reconcile",
            "inputs": [str(a), str(b)],
            "options": {"preset": "invoice_tracker_vs_ledger"},
            "output": str(out),
            "approval": {"confirmed": True},
        }
        result = ar.run_plan(plan, base_dir=td)
        assert result["status"] == "success", result
        assert result["metrics"]["matched"] == 1 and out.exists()


def test_visualise_declarative_runtime():
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        src = td / "rows.json"
        src.write_text(json.dumps([{"Owner": "A", "Status": "Open"}]), encoding="utf-8")
        out = td / "dashboard.html"
        plan = {
            "version": 1,
            "skill": "data-visualise",
            "input": str(src),
            "dashboard": {
                "title": "Agent runtime dashboard",
                "as_of": "15 Jul 2026",
                "blocks": [
                    {"type": "kpi_row", "items": [{"label": "Open", "value": 1, "status": "amber"}]},
                    {"type": "table", "title": "Rows", "rows": "$source",
                     "rag": {"Status": {"Open": "amber"}}},
                    {"type": "sparkline", "title": "Tiny trend",
                     "data": [["W1", 1], ["W2", 3], ["W3", 2]]},
                    {"type": "waterfall", "title": "Bridge", "steps": [
                        {"label": "Start", "value": 10, "kind": "start"},
                        {"label": "Up", "value": 5, "kind": "delta"},
                        {"label": "End", "value": 15, "kind": "total"},
                    ]},
                ]
            },
            "output": str(out),
        }
        result = ar.run_plan(plan, base_dir=td)
        assert result["status"] == "success", result
        html = out.read_text(encoding="utf-8")
        assert "Agent runtime dashboard" in html and "Owner" in html and "Open" in html
        assert "border-left:3px solid" in html
        assert "spark" in html and "Bridge" in html


def test_visualise_from_analysis_json():
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        analysis = td / "analysis.json"
        analysis.write_text(json.dumps({
            "source": "sales.json",
            "results": [
                {"op": "numeric_summary", "name": "Amount", "result": {
                    "n": 3, "total": "300", "mean": "100", "median": "100",
                }},
                {"op": "period_series", "name": "Monthly", "result": {
                    "grain": "month",
                    "periods": [
                        {"period": "2026-01", "count": 1, "total": "100", "delta": None},
                        {"period": "2026-02", "count": 1, "total": "200", "delta": "100"},
                    ],
                }},
            ],
        }), encoding="utf-8")
        out = td / "insight.html"
        plan = {
            "version": 1,
            "skill": "data-visualise",
            "input": str(analysis),
            "dashboard": {
                "title": "Insight board",
                "as_of": "18 Jul 2026",
                "blocks": "$analysis",
            },
            "output": str(out),
        }
        result = ar.run_plan(plan, base_dir=td)
        assert result["status"] == "success", result
        assert result["metrics"]["from_analysis"] is True
        assert result["metrics"].get("format", "html") == "html"
        html = out.read_text(encoding="utf-8")
        assert "Insight board" in html and "Monthly" in html
        assert "period bridge" in html and "spark" in html

        filtered_out = td / "filtered.html"
        filtered = {
            "version": 1,
            "skill": "data-visualise",
            "input": str(analysis),
            "dashboard": {
                "title": "Amount only",
                "blocks": [{"type": "from_analysis", "ops": ["numeric_summary"]}],
            },
            "output": str(filtered_out),
        }
        result2 = ar.run_plan(filtered, base_dir=td)
        assert result2["status"] == "success", result2
        html2 = filtered_out.read_text(encoding="utf-8")
        assert "Amount" in html2 and "period bridge" not in html2


def test_visualise_xlsx_charts_from_analysis():
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        analysis = td / "analysis.json"
        analysis.write_text(json.dumps({
            "results": [
                {"op": "breakdown", "name": "By customer", "result": {
                    "by": "Customer",
                    "groups": [
                        {"key": "Acme", "total": "100", "count": 2},
                        {"key": "Beta", "total": "60", "count": 1},
                    ],
                }},
                {"op": "period_series", "name": "Monthly", "result": {
                    "grain": "month",
                    "periods": [
                        {"period": "2026-01", "total": "100", "delta": None},
                        {"period": "2026-02", "total": "160", "delta": "60"},
                    ],
                }},
            ],
        }), encoding="utf-8")
        out = td / "insight-charts.xlsx"
        plan = {
            "version": 1,
            "skill": "data-visualise",
            "format": "xlsx",
            "input": str(analysis),
            "dashboard": {"title": "Insight charts", "blocks": "$analysis"},
            "output": str(out),
        }
        result = ar.run_plan(plan, base_dir=td)
        assert result["status"] == "success", result
        assert result["metrics"]["format"] == "xlsx"
        assert result["metrics"]["charts"] >= 2
        assert out.is_file()
        from openpyxl import load_workbook
        wb = load_workbook(out)
        assert wb.sheetnames
        assert any(ws._charts for ws in wb.worksheets)

        explicit = td / "explicit.xlsx"
        plan2 = {
            "version": 1,
            "skill": "data-visualise",
            "dashboard": {
                "title": "Manual",
                "blocks": [{
                    "type": "chart",
                    "chart_type": "pie",
                    "title": "Mix",
                    "categories": ["A", "B"],
                    "series": [{"name": "Share", "values": [60, 40]}],
                }],
            },
            "output": str(explicit),
        }
        result2 = ar.run_plan(plan2, base_dir=td)
        assert result2["status"] == "success", result2
        assert result2["metrics"]["format"] == "xlsx"
        assert explicit.is_file()


def test_dry_run_writes_nothing():
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        src = td / "x.json"
        src.write_text(json.dumps([{"x": 1}]), encoding="utf-8")
        out = td / "new-output-dir" / "out.csv"
        plan = {
            "version": 1,
            "skill": "data-convert",
            "input": str(src),
            "spec": {"target": {"format": "csv"}},
            "output": str(out),
        }
        result = ar.run_plan(plan, base_dir=td, dry_run=True)
        assert result["status"] == "success", result
        assert not out.exists()
        assert not out.parent.exists()


def test_drift_requires_bound_receipt_and_blocks_writes():
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        src = td / "source.csv"
        src.write_text("id,new_column\n1,x\n", encoding="utf-8")
        out = td / "new-output-dir" / "out.csv"
        plan = {
            "version": 1, "skill": "data-convert", "input": str(src),
            "spec": {
                "source": {"expected_columns": ["id"]},
                "map": {"id": {"from": "id"}},
                "target": {"format": "csv", "columns": [{"name": "id"}]},
            },
            "output": str(out), "options": {"allow_drift": True},
            "approval": {"confirmed": True},
        }
        blocked = ar.run_plan(plan, base_dir=td, approval_key=APPROVAL_KEY)
        assert blocked["status"] == "needs_approval", blocked
        assert blocked["approvals_required"][0]["kind"] == "source_drift"
        assert not out.exists() and not out.parent.exists()
        assert any("allow_drift is ignored" in str(w) for w in blocked["warnings"])

        bad = signed_receipt(blocked, {"allow_drift": True})
        bad["signature"] = "0" * 64
        bad_plan = copy.deepcopy(plan)
        bad_plan["approval_receipts"] = [bad]
        still_blocked = ar.run_plan(bad_plan, base_dir=td, approval_key=APPROVAL_KEY)
        assert still_blocked["status"] == "needs_approval", still_blocked
        assert not out.exists() and not out.parent.exists()

        good_plan = copy.deepcopy(plan)
        good_plan["approval_receipts"] = [signed_receipt(blocked, {"allow_drift": True})]
        allowed = ar.run_plan(good_plan, base_dir=td, approval_key=APPROVAL_KEY)
        assert allowed["status"] == "success_with_warnings", allowed
        assert out.exists()


def _aggregation_plan(td: Path, output_name: str = "recon.xlsx"):
    a = td / "agg-a.json"
    b = td / "agg-b.json"
    a.write_text(json.dumps([
        {"party": "Acme", "batch": "BR1", "amount": "8000", "date": "05 Jun 2026"}
    ]), encoding="utf-8")
    b.write_text(json.dumps([
        {"party": "Acme", "batch": "BR1", "amount": "3000", "date": "03 Jun 2026"},
        {"party": "Acme", "batch": "BR1", "amount": "5000", "date": "04 Jun 2026"},
    ]), encoding="utf-8")
    return {
        "version": 1, "skill": "data-reconcile", "inputs": [str(a), str(b)],
        "options": {
            "mode": "amount_date", "amount": "amount", "date": "date",
            "aggregate": True, "group_col": "batch", "party_col": "party",
            "date_window": 5,
        },
        "output": str(td / "new-output-dir" / output_name),
        "approval": {"confirmed": True},
    }


def test_aggregation_receipts_bind_selection_and_dry_run_action():
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        plan = _aggregation_plan(td)
        dry = ar.run_plan(plan, base_dir=td, dry_run=True, approval_key=APPROVAL_KEY)
        assert dry["status"] == "needs_approval" and dry["action"] == "dry-run", dry
        assert dry["approvals_required"][0]["kind"] == "aggregation_proposals"
        assert not Path(plan["output"]).parent.exists()

        asserted_only = copy.deepcopy(plan)
        asserted_only["accepted_aggregations"] = []
        blocked = ar.run_plan(asserted_only, base_dir=td, approval_key=APPROVAL_KEY)
        assert blocked["status"] == "needs_approval", blocked

        reject_plan = copy.deepcopy(plan)
        reject_plan["approval_receipts"] = [signed_receipt(dry, {"accepted_aggregations": []})]
        rejected = ar.run_plan(reject_plan, base_dir=td, approval_key=APPROVAL_KEY)
        assert rejected["status"] == "success_with_warnings", rejected
        assert rejected["details"]["summary"].get("aggregated", 0) == 0
        assert Path(reject_plan["output"]).exists()

        accept_plan = _aggregation_plan(td, "accepted.xlsx")
        request_result = ar.run_plan(accept_plan, base_dir=td, approval_key=APPROVAL_KEY)
        receipt = signed_receipt(request_result, {"accepted_aggregations": [0]})
        accept_plan["approval_receipts"] = [receipt]
        accepted = ar.run_plan(accept_plan, base_dir=td, approval_key=APPROVAL_KEY)
        assert accepted["status"] == "success", accepted
        assert accepted["details"]["summary"]["aggregated"] == 1
        assert accepted["metrics"]["exceptions"] == 0

        mismatch = copy.deepcopy(accept_plan)
        mismatch["accepted_aggregations"] = []
        mismatch_result = ar.run_plan(mismatch, base_dir=td, approval_key=APPROVAL_KEY)
        assert mismatch_result["status"] == "error", mismatch_result
        assert "does not match" in mismatch_result["errors"][0]


def test_accepted_aggregations_without_proposals_is_warning_only():
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        a = td / "same-a.json"
        b = td / "same-b.json"
        row = [{"invoice_no": "INV-1", "amount": "100", "date": "01 Jul 2026"}]
        a.write_text(json.dumps(row), encoding="utf-8")
        b.write_text(json.dumps(row), encoding="utf-8")
        plan = {
            "version": 1, "skill": "data-reconcile", "inputs": [str(a), str(b)],
            "options": {"preset": "invoice_tracker_vs_ledger", "aggregate": True},
            "accepted_aggregations": [],
            "output": str(td / "same.xlsx"), "approval": {"confirmed": True},
        }
        result = ar.run_plan(plan, base_dir=td, approval_key=APPROVAL_KEY)
        assert result["status"] == "success_with_warnings", result
        assert any("no aggregation proposals" in str(w) for w in result["warnings"])
        assert Path(plan["output"]).exists()


def test_direct_breakdown_preserves_zero_total():
    import analyse
    result = analyse.breakdown(["Group", "Amount"], [["A", "10"], ["B", "-10"]],
                               "Group", value="Amount")
    assert result["grand_total"] == Decimal(0), result
    assert result["top1_share"] is None and result["top3_share"] is None


def test_cli_schema_and_inspect():
    cli = ROOT / "bin" / "data-toolkit"
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "x.json"
        p.write_text('[{"a":1}]', encoding="utf-8")
        raw = subprocess.check_output([sys.executable, str(cli), "inspect", str(p)], text=True)
        result = json.loads(raw)
        assert result["status"] == "success" and result["metrics"]["rows"] == 1
        raw = subprocess.check_output([sys.executable, str(cli), "schema"], text=True)
        assert json.loads(raw)["details"]["example"]["version"] == 1


def main():
    tests = [
        ("JSON ingest", test_json_ingest),
        ("JSON path + auto list", test_json_path_and_auto_list),
        ("approval gate", test_plan_validation_approval),
        ("six skill plan shapes", test_all_skill_plan_shapes_validate),
        ("convert JSON/flatten/union/nest", test_convert_json_flatten_union_and_nest),
        ("explicit multi-source union", test_multiple_conversion_inputs_require_explicit_union),
        ("extract fields validation", test_extract_fields_validation_does_not_parse_as_table),
        ("analysis zero total", test_analysis_zero_total_is_preserved),
        ("tidy JSON runtime", test_tidy_json_runtime),
        ("reconcile JSON runtime", test_reconcile_json_runtime),
        ("visualise declarative runtime", test_visualise_declarative_runtime),
        ("visualise from analysis.json", test_visualise_from_analysis_json),
        ("visualise xlsx charts", test_visualise_xlsx_charts_from_analysis),
        ("dry-run", test_dry_run_writes_nothing),
        ("drift signed receipt", test_drift_requires_bound_receipt_and_blocks_writes),
        ("aggregation signed receipts", test_aggregation_receipts_bind_selection_and_dry_run_action),
        ("aggregation no proposals", test_accepted_aggregations_without_proposals_is_warning_only),
        ("direct zero-total breakdown", test_direct_breakdown_preserves_zero_total),
        ("CLI schema + inspect", test_cli_schema_and_inspect),
    ]
    passed = sum(1 for name, fn in tests if check(name, fn))
    print(f"agent runtime: {passed}/{len(tests)} passed")
    if passed != len(tests):
        raise SystemExit(1)


if __name__ == "__main__":
    main()
