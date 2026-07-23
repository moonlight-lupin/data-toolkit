"""Reconciliation — match A <-> B, then triage the discrepancies.

Reconciles any two record sets (A = "ours" / source, B = "theirs" / target): matches
line-by-line on a shared key OR heuristically on amount + date, then runs **Discrepancy
Triage** — every unreconciled item is classified (category, materiality, probable cause,
suggested action), not just listed. Produces a reconciliation working paper (.xlsx) for
review and sign off.

Design notes:
- Pure `match()` / `triage()` work on already-parsed rows (list[dict]) — no I/O, so the
  self-test runs offline with no dependencies.
- `reconcile_files()` is the higher-level wrapper that reads A and B in any format via the
  shared engine (`ingest`) and normalises (`dataclean`) before matching.
- Deterministic. NEVER force-fits a match — an unmatched item stays flagged. It is a
  WORKING PAPER for a qualified person; it does NOT post adjustments or write to any system.

Run `python reconcile.py --self-test` (offline) for a worked check.
"""
from __future__ import annotations

import argparse
import datetime as _dt
import re
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path

# --------------------------------------------------------------------------- #
# Discrepancy Triage taxonomy — the heart of the skill.
# Each category maps to a probable cause and a suggested action. {A}/{B} are filled
# with the two side labels at render time.
# --------------------------------------------------------------------------- #
TRIAGE = {
    "missing_in_B":     ("In {A} but not {B} — unrecorded/omitted in {B}, or a timing/cut-off item",
                         "Investigate; record in {B} or chase {B}"),
    "missing_in_A":     ("In {B} but not {A} — unrecorded/omitted in {A}, or a timing/cut-off item",
                         "Investigate; record in {A} or query {B}"),
    "amount_mismatch":  ("Same item matched, but the amounts differ",
                         "Investigate the difference; correct the wrong side"),
    "rounding":         ("Amounts differ within the rounding / FX tolerance",
                         "Accept within tolerance (note only)"),
    "sign_flip":        ("Equal magnitude, opposite sign — a debit/credit or direction error",
                         "Correct the sign / side of the entry"),
    "duplicate":        ("Appears more than once on the {side} side",
                         "Confirm and remove the duplicate"),
    "timing_difference":("Matches on amount but the dates differ — likely in-transit / cut-off",
                         "Monitor — expected to clear next period"),
    "ambiguous_match":  ("Equal amount but the dates differ beyond the matching window — "
                         "possibly the same item, possibly coincidental",
                         "Confirm it is the same item before reconciling; else treat the two as separate"),
    "currency_mismatch":("Same item on both sides, but the currencies differ — the amounts are "
                         "not comparable as-is",
                         "Investigate; convert at the correct rate, or fix the mis-booked currency"),
    "currency_unknown": ("Amounts tie but at least one side's currency is unknown — can't confirm "
                         "it's the same money (strict-currency mode)",
                         "Establish the currency on both sides, then re-run"),
    "parse_error":      ("A matched item has a missing or unparseable amount/date, so it cannot be "
                         "reconciled safely",
                         "Fix the source value, then re-run"),
}

# --------------------------------------------------------------------------- #
# Presets — column-mapping + match defaults for common recurring reconciliations.
# The engine is generic; a preset just pre-fills sensible defaults (override per run).
# --------------------------------------------------------------------------- #
PRESETS = {
    "invoice_tracker_vs_ledger": {
        "label": "Internal invoice tracker vs accounting records",
        "a_label": "Tracker", "b_label": "Ledger",
        "mode": "key", "key": "invoice_no", "amount": "amount", "date": "date",
        "note": "Completeness: every tracked invoice is booked, and every booked entry is tracked. "
                "missing_in_B = invoiced-not-booked; missing_in_A = booked-not-tracked.",
    },
    "bank_vs_ledger": {
        "label": "Bank statement vs cashbook / ledger",
        "a_label": "Bank", "b_label": "Cashbook",
        "mode": "amount_date", "key": None, "amount": "amount", "date": "date",
        "note": "Banks rarely share a clean key — match on amount + date; un-cleared items are "
                "usually timing_difference (in transit) until they clear.",
    },
    "fa_vs_internal": {
        "label": "Fund administrator vs internal records",
        "a_label": "FA", "b_label": "Internal",
        "mode": "key", "key": "ref", "amount": "amount", "date": "date",
        "note": "The outsourced-admin check (NAV / cash / capital accounts / positions). "
                "Sensitive or confidential data — keep strictly local.",
    },
    "payments_vs_bank": {
        "label": "Payments (AP / PRF) vs bank",
        "a_label": "Approved", "b_label": "Bank",
        "mode": "key", "key": "payment_ref", "amount": "amount", "date": "date",
        "note": "Payments requested/approved vs what actually left the bank. missing_in_B = "
                "approved-not-paid; missing_in_A = paid-without-approval (flag).",
    },
}


# --------------------------------------------------------------------------- #
# Parsing helpers
# --------------------------------------------------------------------------- #
def to_amount(x):
    """Best-effort numeric parse -> Decimal (exact, for finance): strips currency symbols and
    thousands separators, treats (brackets) as negative. Decimal (not float) so reconciliation
    sums and tolerances are exact — no binary-float drift making a tie look like a 0.01 break."""
    if x is None or isinstance(x, bool):
        return None
    if isinstance(x, Decimal):
        return x
    if isinstance(x, int):
        return Decimal(x)
    if isinstance(x, float):
        return Decimal(str(x))                    # via str() so 0.1 stays 0.1
    s = str(x).strip()
    if not s:
        return None
    neg = s.startswith("(") and s.endswith(")")
    s = re.sub(r"[^0-9.\-]", "", s.replace(",", ""))
    if s in ("", "-", ".", "-."):
        return None
    try:
        v = Decimal(s)
    except InvalidOperation:
        return None
    return -v if neg else v


def to_date(x):
    """Parse common date forms to a date; None if unparseable (kept raw, flagged elsewhere)."""
    if x is None or str(x).strip() == "":
        return None
    if isinstance(x, _dt.datetime):
        return x.date()
    if isinstance(x, _dt.date):
        return x
    s = str(x).strip()
    for fmt in ("%Y-%m-%d", "%Y%m%d", "%d/%m/%Y", "%d-%m-%Y", "%d-%b-%Y",
                "%d %b %Y", "%d %B %Y", "%m/%d/%Y", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return _dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def norm_key(x):
    return re.sub(r"\s+", "", str(x).strip().lower()) if x not in (None, "") else None


def _resolve_col(rows, name):
    """Resolve a configured column name against a side's actual headers, case- and
    whitespace-insensitively, so `--amount amount` still finds a bank CSV's 'Amount '.
    Exact match wins; a unique normalised match is used; otherwise the name is returned
    as-is (missing columns then behave as before — values come back None and get flagged)."""
    if not name or not rows:
        return name
    keys = list(rows[0].keys())
    if name in keys:
        return name
    want = str(name).strip().lower()
    hits = [k for k in keys if str(k).strip().lower() == want]
    return hits[0] if len(hits) == 1 else name


def _row_amount(r, amount_col, debit_col=None, credit_col=None):
    """Signed amount for a row. When debit/credit columns are configured and either cell
    parses, the signed amount is debit - credit (accounting convention: debits positive);
    else fall back to the single amount column. Handles the common bank/GL export layout
    with separate Debit and Credit columns — including a mixed pair of files where only
    one side splits them (the other side just uses its amount column)."""
    if debit_col or credit_col:
        d = to_amount(r.get(debit_col)) if debit_col else None
        c = to_amount(r.get(credit_col)) if credit_col else None
        if d is not None or c is not None:
            return (d or Decimal(0)) - (c or Decimal(0))
    return to_amount(r.get(amount_col))


# Currency detection (mirrors dataclean's, kept local so the pure matcher runs offline).
# A bare "$" is deliberately ambiguous -> None (it is NOT assumed USD).
_CCY_SYMBOLS = {"US$": "USD", "S$": "SGD", "A$": "AUD", "HK$": "HKD", "NZ$": "NZD",
                "C$": "CAD", "R$": "BRL", "£": "GBP", "€": "EUR", "¥": "JPY"}
_CCY_CODES = {"USD", "SGD", "GBP", "EUR", "JPY", "AUD", "HKD", "NZD", "CAD", "CHF",
              "CNY", "RMB", "INR", "BRL"}


def to_currency(x):
    """Best-effort ISO currency code from a cell -> code or None. Reads an explicit code
    (USD/SGD/…) or a disambiguated symbol (US$/S$/A$/HK$/NZ$/C$/£/€/¥). A bare '$' stays
    ambiguous (None) — never assumed USD, the SG/AU/HK foot-gun this guards against."""
    if x is None:
        return None
    s = str(x).strip().upper()
    if not s:
        return None
    for code in _CCY_CODES:
        if re.search(rf"\b{code}\b", s):
            return "CNY" if code == "RMB" else code
    for sym, code in _CCY_SYMBOLS.items():           # US$ before S$ (substring) — dict order holds
        if sym.upper() in s:
            return code
    return None


def _rec_ccy(row, currency_col, amount_col, debit_col=None, credit_col=None):
    """The record's currency: an explicit currency column wins (normalised), else detect from
    the amount (or debit/credit) cell's symbol/code; None when genuinely unknown."""
    if currency_col and row.get(currency_col) not in (None, ""):
        raw = row.get(currency_col)
        return to_currency(raw) or str(raw).strip().upper() or None
    for col in (amount_col, debit_col, credit_col):
        if col:
            ccy = to_currency(row.get(col))
            if ccy:
                return ccy
    return None


def _ccy_relation(a_ccy, b_ccy, strict=False):
    """Classify a pair's currency relation -> 'ok' | 'mismatch' | 'unknown'.
    - both known & equal           -> 'ok'
    - both known & different        -> 'mismatch'
    - at least one unknown          -> permissive: 'ok' (matchable); strict: 'unknown'
    Strict mode is for audit/finance work — it refuses to reconcile what it can't currency-check."""
    if a_ccy is not None and b_ccy is not None:
        return "ok" if a_ccy == b_ccy else "mismatch"
    return "unknown" if strict else "ok"


# --------------------------------------------------------------------------- #
# Stage 1 — match A <-> B
# --------------------------------------------------------------------------- #
# --------------------------------------------------------------------------- #
# Absurd-result guard — a 0%/near-0% match against two sides that share thousands of equal
# amounts is almost always a MISCONFIGURATION (sign convention, debit/credit mapping, date
# wiring), not a real break. The engine can catch mechanically what a careful operator catches
# by eye; at 5,000 rows nobody spots it by eye. We record the amount-only matching POTENTIAL
# here (where both sides are in scope) and let `summarise` raise the warning, so aggregation
# confirmed after matching still counts toward the achieved rate.
# --------------------------------------------------------------------------- #
ABSURD_MIN_POTENTIAL = 20          # below this, a low match rate is unremarkable — stay quiet
ABSURD_RATE = 0.20                 # matched < 20% of what equal amounts alone could pair


def _amount_pairs(a_amts, b_amts):
    """1:1 pairs formable on exactly-equal amounts alone (multiset intersection) — an upper
    bound on what any amount-based matching should achieve."""
    from collections import Counter
    cb = Counter(b_amts)
    return sum(min(n, cb[amt]) for amt, n in Counter(a_amts).items() if amt in cb)


def _amount_potential(A, B):
    """Amount-only pairing potential both as-is and with B negated. A much larger `flipped`
    count is the fingerprint of a wrong sign convention (--flip-b / debit-credit mapping)."""
    a = [x["amt"] for x in A if x["amt"] is not None]
    b = [x["amt"] for x in B if x["amt"] is not None]
    return {"equal": _amount_pairs(a, b), "flipped": _amount_pairs(a, [-v for v in b])}


def match(rows_a, rows_b, *, key=None, amount="amount", date=None, currency=None,
          mode="key", tol=0.01, date_window_days=5, strict_currency=False,
          debit=None, credit=None, flip_b=False):
    """Match two row sets. Returns a dict of buckets, each a list of records.

    Buckets: matched, value_diffs, currency_diffs, currency_unknown, a_only, b_only, dup_a,
    dup_b, ambiguous. Every record carries its source row(s) and the parsed
    amount/date/key/currency.

    `currency` names a currency column (else the code is detected from the amount cell's
    symbol). Currencies are COMPARED, so 100 USD never silently matches 100 SGD: in key mode a
    key match whose currencies differ goes to `currency_diffs` (not amount-compared); in
    amount_date mode an equal-amount pair only matches when the currencies are compatible.

    `strict_currency=True` (audit/finance mode) additionally refuses to match when a side's
    currency is UNKNOWN — those pairs go to `currency_unknown` rather than being assumed
    compatible. Default is permissive (unknown treated as compatible).

    In `amount_date` mode `date_window_days` is a HARD constraint: an equal-amount pair only
    counts as matched when the dates are within the window (a small in-window gap is a genuine
    timing difference). An equal-amount pair OUTSIDE the window is NOT reconciled — it goes to
    `ambiguous` (a possible-but-uncertain match for the reviewer to confirm), so transactions
    weeks apart are never silently passed off as timing differences. `date_window_days=None`
    disables the window (any nearest-date equal-amount pair matches — the old behaviour).

    `debit`/`credit` name separate Debit / Credit columns (the common bank/GL export layout):
    the signed amount is then debit - credit; a row where neither parses falls back to the
    `amount` column, so a debit/credit file can reconcile against a signed-amount file.
    `flip_b=True` negates B's amounts before comparing — for sides that book the same money
    with opposite sign conventions (e.g. bank statement vs the GL cash account).
    All column names are resolved case- and whitespace-insensitively per side."""
    tol = Decimal(str(tol))

    def _mk(rows, flip=False):
        acol = _resolve_col(rows, amount)
        dcol = _resolve_col(rows, date) if date else None
        kcol = _resolve_col(rows, key) if key else None
        ccol = _resolve_col(rows, currency) if currency else None
        drcol = _resolve_col(rows, debit) if debit else None
        crcol = _resolve_col(rows, credit) if credit else None
        out = []
        for i, r in enumerate(rows):
            amt = _row_amount(r, acol, drcol, crcol)
            if flip and amt is not None:
                amt = -amt
            out.append({"row": r, "amt": amt,
                        "dt": to_date(r.get(dcol)) if dcol else None,
                        "key": norm_key(r.get(kcol)) if kcol else None,
                        "ccy": _rec_ccy(r, ccol, acol, drcol, crcol), "i": i})
        return out
    A, B = _mk(rows_a), _mk(rows_b, flip=flip_b)

    res = {"matched": [], "value_diffs": [], "currency_diffs": [], "currency_unknown": [],
           "parse_errors": [], "warnings": [],
           "a_only": [], "b_only": [], "dup_a": [], "dup_b": [], "ambiguous": []}

    if mode == "key":
        from collections import defaultdict
        ib, ia = defaultdict(list), defaultdict(list)
        for b in B:
            ib[b["key"]].append(b)
        for a in A:
            ia[a["key"]].append(a)
        for k, items in ia.items():
            if k is not None and len(items) > 1:
                res["dup_a"].extend(items[1:])
        for k, items in ib.items():
            if k is not None and len(items) > 1:
                res["dup_b"].extend(items[1:])
        dupb_ids = {id(x) for x in res["dup_b"]}
        usedb, seen_a = set(), set()
        for a in A:
            k = a["key"]
            if k is not None and k in seen_a:
                continue                          # duplicate occurrence -> already in dup_a
            if k is not None:
                seen_a.add(k)
            cand = ib.get(k) if k is not None else None
            if cand:
                b = cand[0]
                usedb.add(id(b))
                if a["amt"] is None or b["amt"] is None:
                    res["parse_errors"].append({"a": a, "b": b, "issue": "missing/unparseable amount"})
                    continue
                da = a["amt"] - b["amt"]
                rel = _ccy_relation(a["ccy"], b["ccy"], strict_currency)
                if rel == "mismatch":               # same key, different currency — not comparable
                    res["currency_diffs"].append({"a": a, "b": b, "diff": da})
                elif rel == "unknown":              # strict: can't currency-check this key match
                    res["currency_unknown"].append({"a": a, "b": b, "diff": da})
                else:
                    res["matched" if abs(da) <= tol else "value_diffs"].append({"a": a, "b": b, "diff": da})
            else:
                res["a_only"].append({"a": a})
        for b in B:
            if id(b) not in usedb and id(b) not in dupb_ids:
                res["b_only"].append({"b": b})

    else:  # amount_date heuristic
        usedb = set()
        bidx = list(B)
        win = date_window_days
        # amount_date's safety rail IS the date window. With NO date column named, the window
        # can't be applied — so equal-amount pairs are NOT silently matched on amount alone: they
        # are held as `ambiguous` for confirmation and a warning is raised. (A named-but-
        # unparseable date routes each affected row to a_only via the guards below instead.)
        no_date = not date
        for a in A:
            if a["amt"] is None:
                res["a_only"].append({"a": a})
                continue
            if date and a["dt"] is None:
                res["a_only"].append({"a": a})
                continue
            best_in = best_out = best_unknown = None   # in-window / out-of-window(or no-date) / strict-unknown
            for b in bidx:
                if id(b) in usedb or b["amt"] is None:
                    continue
                if date and b["dt"] is None:
                    continue
                if abs(a["amt"] - b["amt"]) > tol:
                    continue
                rel = _ccy_relation(a["ccy"], b["ccy"], strict_currency)
                if rel == "mismatch":              # equal amount, different currency — not the same money
                    continue
                gap = abs((a["dt"] - b["dt"]).days) if (a["dt"] and b["dt"]) else None
                if rel == "unknown":               # strict: equal amount but currency unverifiable
                    if best_unknown is None or gap is None or best_unknown[1] is None \
                            or gap < best_unknown[1]:
                        best_unknown = (b, gap)
                elif no_date:                      # window unenforceable -> hold for confirmation
                    if best_out is None:
                        best_out = (b, None)
                elif win is None or gap <= win:
                    if best_in is None or gap < best_in[1]:
                        best_in = (b, gap)
                elif best_out is None or best_out[1] is None or gap < best_out[1]:
                    best_out = (b, gap)
            if best_in is not None:
                b, gap = best_in
                usedb.add(id(b))
                rec = {"a": a, "b": b, "diff": (a["amt"] - b["amt"])}
                if gap:
                    rec["date_gap"] = gap          # in-window gap -> genuine timing difference
                res["matched"].append(rec)
            elif best_out is not None:             # equal amount, but window not satisfied/enforceable
                b, gap = best_out
                usedb.add(id(b))                   # reserve b so it isn't also counted as b_only
                rec = {"a": a, "b": b}
                if gap is None:                    # no date column -> matched on amount ALONE
                    rec["no_date"] = True
                    if not res["warnings"]:
                        res["warnings"].append(
                            "amount_date mode ran without a resolvable date column: equal-amount "
                            "pairs are held as ambiguous (matched on amount alone), not reconciled. "
                            "Provide a date column (--date) to enforce the matching window.")
                else:
                    rec["date_gap"] = gap
                res["ambiguous"].append(rec)
            elif best_unknown is not None:         # strict: equal amount, currency unknown
                b, gap = best_unknown
                usedb.add(id(b))
                res["currency_unknown"].append({"a": a, "b": b, "diff": (a["amt"] - b["amt"])})
            else:
                res["a_only"].append({"a": a})
        for b in bidx:
            if id(b) not in usedb:
                res["b_only"].append({"b": b})
        # Second pass: classify the one-sided residue (duplicate / sign_flip / amount_mismatch),
        # which 1:1 amount+date matching alone leaves as bare missing_in_A / missing_in_B.
        _refine_residues(res, tol=tol, win=win)
    res["amount_potential"] = _amount_potential(A, B)
    return res


# --------------------------------------------------------------------------- #
# Stage 2 — Discrepancy Triage
# --------------------------------------------------------------------------- #
def _materiality(v, material, escalate):
    v = abs(v or 0)
    if v >= escalate:
        return "escalate"
    if v >= material:
        return "material"
    return "immaterial"


# Common GST/VAT/WHT rates checked when an amount mismatch looks like net-vs-gross.
# Advisory only — a matching ratio adds a hint to the probable cause, never a category change.
_TAX_RATES = (Decimal("0.05"), Decimal("0.07"), Decimal("0.08"), Decimal("0.09"),
              Decimal("0.10"), Decimal("0.15"), Decimal("0.20"))


def _tax_hint(a_amt, b_amt):
    """If the gap between two amounts is a common GST/VAT/WHT rate applied to the smaller
    side (i.e. one side looks net, the other gross), return an advisory note; else None."""
    if not a_amt or not b_amt:
        return None
    lo, hi = sorted((abs(a_amt), abs(b_amt)))
    if lo == 0:
        return None
    ratio = (hi - lo) / lo
    for r in _TAX_RATES:
        if abs(ratio - r) <= Decimal("0.001"):
            pct = (r * 100).quantize(Decimal(1))
            return f"difference is {pct}% of the smaller side — possible GST/VAT/WHT (net vs gross)"
    return None


# --------------------------------------------------------------------------- #
# Stage 1.5 — refine the amount_date residue (duplicate / sign_flip / amount_mismatch)
# In `key` mode these come free from the shared key; amount_date matching alone leaves them
# as bare one-sided items, so a second pass over a_only / b_only reclassifies them into the
# SAME buckets key-mode uses (value_diffs / dup_a / dup_b) — triage then labels them.
# --------------------------------------------------------------------------- #
def _sig(x):
    """Duplicate signature for a residue item (no shared key in amount_date): amount + date."""
    return (x["amt"], x["dt"])


def _refine_residues(res, *, tol, win, rounding=Decimal("0.05")):
    """Reclassify the one-sided residue. Conservative — with no shared key, it pairs only on
    defensible signals: opposite-equal amounts (sign_flip), a net-vs-gross tax ratio
    (amount_mismatch), or an exact amount+date twin of an already-matched / prior residue item
    (duplicate). Every result is a CONFIRM item for the reviewer, never an auto-adjustment."""
    a_items = [d["a"] for d in res["a_only"]]
    b_items = [d["b"] for d in res["b_only"]]
    used_a, used_b = set(), set()

    def _win_ok(x, y):
        if win is None or not x["dt"] or not y["dt"]:
            return True
        return abs((x["dt"] - y["dt"]).days) <= win

    # (1) sign_flip — equal magnitude, opposite sign (a + b ~ 0), currency-compatible, in window
    for a in a_items:
        if id(a) in used_a or a["amt"] is None or abs(a["amt"]) <= rounding:
            continue
        for b in b_items:
            if id(b) in used_b or b["amt"] is None:
                continue
            if _ccy_relation(a["ccy"], b["ccy"]) == "mismatch":
                continue
            if abs(a["amt"] + b["amt"]) <= rounding and _win_ok(a, b):
                res["value_diffs"].append({"a": a, "b": b, "diff": a["amt"] - b["amt"]})
                used_a.add(id(a)); used_b.add(id(b))
                break

    # (2) amount_mismatch — one item mis-stated. Without a key, pair only when same sign, in
    # window, currency-compatible AND the gap is a common net-vs-gross tax ratio; nearest date wins.
    for a in a_items:
        if id(a) in used_a or a["amt"] is None:
            continue
        best = None
        for b in b_items:
            if id(b) in used_b or b["amt"] is None:
                continue
            if (a["amt"] > 0) != (b["amt"] > 0):        # same sign only
                continue
            if abs(a["amt"] - b["amt"]) <= tol:         # equal -> would already have matched
                continue
            if _ccy_relation(a["ccy"], b["ccy"]) == "mismatch" or not _win_ok(a, b):
                continue
            if _tax_hint(a["amt"], b["amt"]):
                gap = abs((a["dt"] - b["dt"]).days) if (a["dt"] and b["dt"]) else 0
                if best is None or gap < best[1]:
                    best = (b, gap)
        if best is not None:
            res["value_diffs"].append({"a": a, "b": best[0], "diff": a["amt"] - best[0]["amt"]})
            used_a.add(id(a)); used_b.add(id(best[0]))

    # (3) duplicate — an unmatched item that is an exact amount+date twin of an already-matched
    # item (its real pair was reconciled) or of another still-unmatched item on the same side.
    for items, side, used, bucket in (
            (a_items, "a", used_a, "dup_a"), (b_items, "b", used_b, "dup_b")):
        matched_sigs = {}
        for m in res["matched"]:
            it = m[side]
            if it["amt"] is not None:
                matched_sigs[_sig(it)] = matched_sigs.get(_sig(it), 0) + 1
        seen = set()
        for x in items:
            if id(x) in used or x["amt"] is None:
                continue
            k = _sig(x)
            if matched_sigs.get(k) or k in seen:        # twin already matched, or a prior residue twin
                res[bucket].append(x); used.add(id(x))
            else:
                seen.add(k)

    if used_a or used_b:
        res["a_only"] = [d for d in res["a_only"] if id(d["a"]) not in used_a]
        res["b_only"] = [d for d in res["b_only"] if id(d["b"]) not in used_b]
    return res


def triage(res, *, a_label="A", b_label="B", material=1000.0, escalate=10000.0,
           rounding=0.05, date_window_days=5, as_of=None):
    """Classify every unreconciled item. Returns a list of exception dicts.

    `as_of` (a date, or a parseable date string) ages the one-sided items: each
    missing_in_A/missing_in_B exception with a parsed date gains `age_days`, so stale
    open items (unbanked receipts, uncleared payments) rank visibly in the working paper."""
    out = []
    material, escalate, rounding = Decimal(str(material)), Decimal(str(escalate)), Decimal(str(rounding))
    as_of = to_date(as_of)

    def cause_action(cat, side=None):
        c, act = TRIAGE[cat]
        f = {"A": a_label, "B": b_label, "side": side or ""}
        return c.format(**f), act.format(**f)

    for m in res["matched"]:
        if m.get("date_gap"):
            c, act = cause_action("timing_difference")
            out.append(_exc("timing_difference", m["a"], m["b"], Decimal("0"), "monitor", c, act,
                            material, escalate))
    for d in res.get("ambiguous", []):
        a, b = d["a"], d["b"]
        c, act = cause_action("ambiguous_match")
        if d.get("no_date"):
            c = f"{c} (no date column — matched on amount alone; confirm it is the same item)"
        else:
            c = f"{c} (dates {d['date_gap']} days apart)"
        out.append(_exc("ambiguous_match", a, b, a["amt"], "confirm", c, act, material, escalate))
    for d in res.get("currency_diffs", []):
        a, b = d["a"], d["b"]
        c, act = cause_action("currency_mismatch")
        c = f"{c} ({a['ccy'] or '?'} vs {b['ccy'] or '?'})"
        out.append(_exc("currency_mismatch", a, b, a["amt"], "investigate", c, act, material, escalate))
    for d in res.get("currency_unknown", []):
        a, b = d["a"], d["b"]
        c, act = cause_action("currency_unknown")
        c = f"{c} ({a['ccy'] or '?'} vs {b['ccy'] or '?'})"
        out.append(_exc("currency_unknown", a, b, a["amt"], "confirm", c, act, material, escalate))
    for d in res.get("parse_errors", []):
        a, b = d.get("a"), d.get("b")
        c, act = cause_action("parse_error")
        if d.get("issue"):
            c = f"{c} ({d['issue']})"
        mag = (a or {}).get("amt") or (b or {}).get("amt")
        out.append(_exc("parse_error", a, b, mag, "correct", c, act, material, escalate))
    for d in res["value_diffs"]:
        diff = d["diff"]
        a, b = d["a"], d["b"]
        adiff = round(abs(diff), 2)            # avoid float dust at the band edge
        if adiff <= rounding:
            cat = "rounding"
        elif a["amt"] is not None and b["amt"] is not None \
                and round(abs(a["amt"] + b["amt"]), 2) <= rounding and round(abs(a["amt"]), 2) > rounding:
            cat = "sign_flip"
        else:
            cat = "amount_mismatch"
        c, act = cause_action(cat)
        if cat == "amount_mismatch":
            hint = _tax_hint(a["amt"], b["amt"])
            if hint:
                c = f"{c} ({hint})"
        out.append(_exc(cat, a, b, diff, _verb(cat), c, act, material, escalate))

    def _aged(e, item):
        if as_of and item.get("dt"):
            e["age_days"] = (as_of - item["dt"]).days
            e["probable_cause"] += f" (aged {e['age_days']}d at {as_of.strftime('%d %b %Y')})"
        return e

    for d in res["a_only"]:
        c, act = cause_action("missing_in_B")
        a = d["a"]
        out.append(_aged(_exc("missing_in_B", a, None, a["amt"], "investigate", c, act,
                              material, escalate), a))
    for d in res["b_only"]:
        c, act = cause_action("missing_in_A")
        b = d["b"]
        out.append(_aged(_exc("missing_in_A", None, b, b["amt"], "investigate", c, act,
                              material, escalate), b))
    for d in res["dup_a"]:
        c, act = cause_action("duplicate", side=a_label)
        out.append(_exc("duplicate", d, None, d["amt"], "remove", c, act, material, escalate))
    for d in res["dup_b"]:
        c, act = cause_action("duplicate", side=b_label)
        out.append(_exc("duplicate", None, d, d["amt"], "remove", c, act, material, escalate))

    order = {"escalate": 0, "material": 1, "immaterial": 2}
    out.sort(key=lambda e: (order.get(e["materiality"], 3), -abs(e["magnitude"] or 0)))
    return out


def _verb(cat):
    return {"rounding": "accept", "sign_flip": "correct", "amount_mismatch": "investigate"}.get(cat, "investigate")


def _exc(cat, a, b, magnitude, action, cause, action_text, material, escalate):
    mat = "within tolerance" if cat == "rounding" else _materiality(magnitude, material, escalate)
    return {
        "category": cat, "magnitude": magnitude, "materiality": mat,
        "probable_cause": cause, "suggested_action": action_text, "action": action,
        "a": (a or {}).get("row") if a else None,
        "b": (b or {}).get("row") if b else None,
        "currency": (a or {}).get("ccy") or (b or {}).get("ccy"),
        "age_days": None,
        "status": "open",
    }


# --------------------------------------------------------------------------- #
# Aggregation matching (sum-to-one / sum-to-sum) — a 2nd pass, CONFIRM-FIRST.
# Heuristic, so proposals are NEVER auto-accepted: the caller presents them and the
# user confirms; apply_aggregations() only moves the CONFIRMED ones to matched.
# Bounded by a shared key + a date window + counterparty (and a subset-size cap) to
# stay tractable and avoid coincidental matches.
# --------------------------------------------------------------------------- #
from itertools import combinations as _combos  # noqa: E402


def _same(a_row, b_row, col):
    return norm_key(a_row.get(col)) == norm_key(b_row.get(col))


def _within(a_item, b_item, days):
    if days is None or not a_item["dt"] or not b_item["dt"]:
        return True
    return abs((a_item["dt"] - b_item["dt"]).days) <= days


def _compat(t, c, group_col, party_col, date_window):
    if group_col and not _same(t["row"], c["row"], group_col):
        return False
    if party_col and not _same(t["row"], c["row"], party_col):
        return False
    return _within(t, c, date_window)


def _currency_compatible(items):
    """True when known currencies in the proposed aggregation do not conflict."""
    known = [x.get("ccy") for x in items if x.get("ccy") is not None]
    return all(_ccy_relation(a, b) != "mismatch" for i, a in enumerate(known) for b in known[i + 1:])


def _subset_sum(target_amt, cands, tol, max_subset):
    """Smallest subset (size >= 2) of cands whose amounts tie to target within tol; else None."""
    items = [c for c in cands if c["amt"] is not None]
    for size in range(2, min(max_subset, len(items)) + 1):
        for combo in _combos(items, size):
            if abs(sum(c["amt"] for c in combo) - target_amt) <= tol:
                return list(combo)
    return None


def _basis(group_col, party_col, date_window, extra=""):
    bits = []
    if group_col:
        bits.append(f"same {group_col}")
    if party_col:
        bits.append(f"same {party_col}")
    if date_window is not None:
        bits.append(f"within {date_window}d")
    bits.append("amounts tie")
    return (" + ".join(bits) + extra).strip()


def _proposal(kind, A, B, basis):
    sa = round(sum(x["amt"] or 0 for x in A), 2)
    sb = round(sum(x["amt"] or 0 for x in B), 2)
    return {"kind": kind, "basis": basis,
            "a_items": [x["row"] for x in A], "b_items": [x["row"] for x in B],
            "a_sum": sa, "b_sum": sb, "residual": round(sa - sb, 2),
            "_A": A, "_B": B}            # internal handles for apply (in-process only)


def propose_aggregations(res, *, group_col=None, party_col=None, date_window=None,
                         tol=0.01, max_subset=5):
    """Read-only: propose sum-to-one and sum-to-sum matches over the still-unmatched items.
    Returns a list of proposal dicts. NEVER mutates res and NEVER auto-accepts — the caller
    confirms, then calls apply_aggregations()."""
    tol = Decimal(str(tol))
    a_rem = [d["a"] for d in res["a_only"]]
    b_rem = [d["b"] for d in res["b_only"]]
    used, proposals = set(), []
    basis = _basis(group_col, party_col, date_window)

    # (1) sum-to-sum control totals — tie group sums, grouping by the available key(s)
    # combined (e.g. batch + counterparty), so a shared run still splits per counterparty.
    group_cols = [c for c in (group_col, party_col) if c]
    if group_cols:
        from collections import defaultdict
        def gk(x):
            return tuple(norm_key(x["row"].get(c)) for c in group_cols)
        ga, gb = defaultdict(list), defaultdict(list)
        for x in a_rem:
            ga[gk(x)].append(x)
        for x in b_rem:
            gb[gk(x)].append(x)
        for k in set(ga) & set(gb):
            if any(v is None for v in k):
                continue
            A, B = ga[k], gb[k]
            if len(A) + len(B) < 3:            # a 1:1 within a group isn't an aggregation
                continue
            if not _currency_compatible(A + B):
                continue
            if abs(sum(x["amt"] or 0 for x in A) - sum(x["amt"] or 0 for x in B)) <= tol:
                label = "+".join(str(v) for v in k)
                proposals.append(_proposal("many_to_many", A, B,
                                           _basis(group_col, party_col, date_window,
                                                  f" (group {'+'.join(group_cols)}={label})")))
                used |= {id(x) for x in A + B}

    # (2) sum-to-one subset-sum on what's left, constrained by the keys.
    for side, targets, pool, kind in [("A", a_rem, b_rem, "one_to_many"),
                                      ("B", b_rem, a_rem, "many_to_one")]:
        for t in targets:
            if id(t) in used or t["amt"] is None:
                continue
            cands = [c for c in pool if id(c) not in used and _compat(t, c, group_col, party_col, date_window)]
            cands = [c for c in cands if _currency_compatible([t, c])]
            hit = _subset_sum(t["amt"], cands, tol, max_subset)
            if hit and not _currency_compatible([t] + hit):
                hit = None
            if hit:
                A, B = ([t], hit) if side == "A" else (hit, [t])
                proposals.append(_proposal(kind, A, B, basis))
                used |= {id(x) for x in [t] + hit}
    return proposals


def apply_aggregations(res, proposals, accepted):
    """Move CONFIRMED proposals' items out of the exception pool into 'matched_agg'.
    `accepted` = indices (into `proposals`) the user confirmed. In-process only."""
    res.setdefault("matched_agg", [])
    keep = set()
    for i in set(accepted):
        p = proposals[i]
        res["matched_agg"].append(p)
        keep |= {id(x) for x in p["_A"]} | {id(x) for x in p["_B"]}
    res["a_only"] = [d for d in res["a_only"] if id(d["a"]) not in keep]
    res["b_only"] = [d for d in res["b_only"] if id(d["b"]) not in keep]
    return res


def render_proposals(proposals, *, a_label="A", b_label="B"):
    """A short, human-readable list of the proposed aggregations to confirm."""
    if not proposals:
        return "No aggregation proposals."
    L = ["## Proposed aggregations — CONFIRM before they count as reconciled", ""]
    for i, p in enumerate(proposals):
        tie = "ties" if abs(p["residual"]) < 0.005 else f"residual {p['residual']:,}"
        L.append(f"**[{i}] {p['kind']}** — {a_label} {p['a_sum']:,} vs {b_label} {p['b_sum']:,} ({tie}); {p['basis']}")
        L.append(f"    {a_label}: " + " | ".join(str(r) for r in p["a_items"]))
        L.append(f"    {b_label}: " + " | ".join(str(r) for r in p["b_items"]))
    L.append("")
    L.append("> Confirm which to accept; only confirmed proposals are applied. The rest stay as exceptions.")
    return "\n".join(L)


def finalize(res, *, a_label="A", b_label="B", material=1000.0, escalate=10000.0, as_of=None):
    """Triage what remains (after any confirmed aggregations) and summarise."""
    exceptions = triage(res, a_label=a_label, b_label=b_label, material=material,
                        escalate=escalate, as_of=as_of)
    summary = summarise(res, exceptions)
    return exceptions, summary


# --------------------------------------------------------------------------- #
# Statement completeness — opening + net movement = closing
# --------------------------------------------------------------------------- #
def check_balance(rows, *, amount="amount", debit=None, credit=None,
                  opening=None, closing=None):
    """Completeness check for a statement-style extract: does opening balance + net movement
    tie to the stated closing balance? Catches a truncated or filtered export BEFORE it
    silently 'reconciles'. Returns a dict; `ties` is None when a balance is missing/unparseable.
    Movement is summed in the file's own sign convention (debit - credit when those columns
    are configured)."""
    acol = _resolve_col(rows, amount)
    drcol = _resolve_col(rows, debit) if debit else None
    crcol = _resolve_col(rows, credit) if credit else None
    movement, unparsed = Decimal("0"), 0
    for r in rows:
        amt = _row_amount(r, acol, drcol, crcol)
        if amt is None:
            unparsed += 1
        else:
            movement += amt
    op, cl = to_amount(opening), to_amount(closing)
    computed = (op + movement) if op is not None else None
    ties = None
    if computed is not None and cl is not None:
        ties = abs(computed - cl) <= Decimal("0.01")
    return {"opening": op, "movement": movement, "computed_closing": computed,
            "stated_closing": cl, "ties": ties, "rows": len(rows), "unparsed": unparsed}


def render_balance_check(chk, label="A"):
    """One-line human summary of a check_balance() result."""
    if chk["ties"] is None:
        return (f"Balance check ({label}): incomplete — opening/closing missing or unparseable "
                f"(net movement {_money(chk['movement'])} over {chk['rows']} row(s))")
    verdict = "TIES" if chk["ties"] else \
        f"DOES NOT TIE (gap {_money(chk['computed_closing'] - chk['stated_closing'])}) — extract may be truncated/filtered"
    line = (f"Balance check ({label}): opening {_money(chk['opening'])} + movement "
            f"{_money(chk['movement'])} = {_money(chk['computed_closing'])} vs stated closing "
            f"{_money(chk['stated_closing'])} — {verdict}")
    if chk["unparsed"]:
        line += f"; {chk['unparsed']} row(s) unparseable and excluded"
    return line


# --------------------------------------------------------------------------- #
# Summary + report
# --------------------------------------------------------------------------- #
def _absurd_match_warning(res, reconciled):
    """Warn when almost nothing reconciled although equal amounts alone could have paired
    far more — the mechanical signature of a misconfigured run. Returns the message or None.

    Deliberately blunt: this says the RESULT is not trustworthy, which is a different claim
    from the individual exceptions being real. Silent on small files (< ABSURD_MIN_POTENTIAL)
    and whenever a sane share of the potential was achieved."""
    pot = res.get("amount_potential") or {}
    equal, flipped = pot.get("equal", 0), pot.get("flipped", 0)
    potential = max(equal, flipped)
    if potential < ABSURD_MIN_POTENTIAL or reconciled >= potential * ABSURD_RATE:
        return None
    msg = (f"UNRELIABLE RESULT — only {reconciled} item(s) reconciled, but {potential} pair(s) "
           f"of exactly-equal amounts exist across the two sides. A near-zero match rate at "
           f"this scale is almost always a configuration error, not a real break. ")
    if flipped > equal * 2:
        msg += ("The two sides appear to be SIGN-INVERTED relative to each other "
                f"({flipped} pairs match when one side is negated, vs {equal} as-is) — check "
                "--flip-b and the debit/credit mapping. ")
    else:
        msg += "Check the sign convention (--flip-b), the debit/credit mapping "
    msg += ("and that the date column is wired (--date) before accepting this working paper.")
    return msg


def summarise(res, exceptions, *, amount="amount"):
    agg = res.get("matched_agg", [])
    matched_n = len(res["matched"])
    agg_a_items = sum(len(p["a_items"]) for p in agg)
    reconciled = matched_n + agg_a_items
    total_n = reconciled + len(exceptions)
    mat_val = sum(abs((m["a"]["amt"] or 0)) for m in res["matched"]) + sum(abs(p["a_sum"]) for p in agg)
    exc_val = sum(abs(e["magnitude"] or 0) for e in exceptions)
    by_cat = {}
    for e in exceptions:
        by_cat.setdefault(e["category"], {"n": 0, "val": Decimal("0")})
        by_cat[e["category"]]["n"] += 1
        by_cat[e["category"]]["val"] += abs(e["magnitude"] or 0)
    # Per-currency split of the value figures — the headline totals sum across currencies
    # (indicative only when mixed), so a mixed recon also reports each currency separately.
    by_ccy = {}
    def _ccy_add(ccy, field, v):
        d = by_ccy.setdefault(ccy or "?", {"matched_val": Decimal("0"),
                                           "exception_val": Decimal("0")})
        d[field] += abs(v or 0)
    for m in res["matched"]:
        _ccy_add(m["a"].get("ccy") or m["b"].get("ccy"), "matched_val", m["a"]["amt"])
    for p in agg:
        items = p.get("_A") or []
        ccy = next((x.get("ccy") for x in items if x.get("ccy")), None)
        _ccy_add(ccy, "matched_val", p["a_sum"])
    for e in exceptions:
        _ccy_add(e.get("currency"), "exception_val", e["magnitude"])
    reds = sum(1 for e in exceptions if e["materiality"] in ("material", "escalate"))
    rag = "GREEN" if reds == 0 and exceptions == [] else ("RED" if reds else "AMBER")
    warnings = list(res.get("warnings", []))
    absurd = _absurd_match_warning(res, reconciled)
    if absurd:
        warnings.insert(0, absurd)          # lead with it — it questions the whole result
    return {
        "matched": matched_n, "aggregated": len(agg), "exceptions": len(exceptions), "total": total_n,
        "pct_reconciled": round(100 * reconciled / total_n, 1) if total_n else 100.0,
        "value_matched": round(mat_val, 2), "value_in_exception": round(exc_val, 2),
        "by_category": by_cat, "by_currency": by_ccy,
        "material_or_escalate": reds, "rag": rag,
        "warnings": warnings, "unreliable": bool(absurd),
    }


CURRENCY_DP = {"JPY": 0}


def _quantize_money(v, ccy=None):
    q = Decimal(1).scaleb(-CURRENCY_DP.get(ccy, 2))
    return Decimal(str(v or 0)).quantize(q)


def _money(v, ccy=None):
    return f"{_quantize_money(v, ccy):,}"


def _md_escape(v):
    return str("" if v is None else v).replace("|", "\\|").replace("\r\n", "<br>").replace("\n", "<br>")


def render_report(summary, exceptions, *, a_label="A", b_label="B", title="Reconciliation",
                  balances=None):
    L = [f"# {_md_escape(title)}: {_md_escape(a_label)} vs {_md_escape(b_label)}", ""]
    s = summary
    agg = f", {s['aggregated']} aggregation(s)" if s.get("aggregated") else ""
    L.append(f"**{s['rag']}{' · ⛔ UNRELIABLE' if s.get('unreliable') else ''}** — "
             f"{s['pct_reconciled']}% of items reconciled "
             f"({s['matched']} matched 1:1{agg}); {s['exceptions']} exception(s), "
             f"{s['material_or_escalate']} material/escalate.")
    L.append(f"Value matched: {_money(s['value_matched'])} · value in exception: {_money(s['value_in_exception'])}")
    for w in s.get("warnings", []):
        L.append(f"> ⚠️ **Warning** — {_md_escape(w)}")
    for label, chk in (balances or []):
        L.append(_md_escape(render_balance_check(chk, label)))
    L.append("")
    by_ccy = s.get("by_currency") or {}
    if len(by_ccy) > 1:
        L.append("Mixed currencies — the value totals above sum across currencies (indicative "
                 "only); per currency:")
        L.append("| Currency | Value matched | Value in exception |")
        L.append("|---|---|---|")
        for c in sorted(by_ccy):
            d = by_ccy[c]
            L.append(f"| {_md_escape(c)} | {_money(d['matched_val'], c)} | {_money(d['exception_val'], c)} |")
        L.append("")
    if s["by_category"]:
        L.append("| Category | # | Value |")
        L.append("|---|---|---|")
        for c, d in sorted(s["by_category"].items(), key=lambda kv: -kv[1]["val"]):
            L.append(f"| {_md_escape(c)} | {d['n']} | {_money(d['val'])} |")
        L.append("")
    if exceptions:
        L.append("## Exceptions (triaged — highest first)")
        L.append("| Category | Magnitude | Materiality | Probable cause | Action |")
        L.append("|---|---|---|---|---|")
        for e in exceptions[:50]:
            L.append(f"| {_md_escape(e['category'])} | {_money(e['magnitude'], e.get('currency'))} "
                     f"| {_md_escape(e['materiality'])} | {_md_escape(e['probable_cause'])} "
                     f"| {_md_escape(e['suggested_action'])} |")
    L.append("")
    L.append("> Working paper for review by a qualified person — not a posting; "
             "no adjustment is made. Unmatched items are flagged, never force-fitted.")
    return "\n".join(L)


def write_workpaper(res, exceptions, summary, out_path, *, a_label="A", b_label="B"):
    """Write the .xlsx working paper (Summary + Matched + Exceptions). Needs openpyxl."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    s = summary
    for r in [["Reconciliation", f"{a_label} vs {b_label}"], ["RAG", s["rag"]],
              ["% reconciled", s["pct_reconciled"]], ["Matched", s["matched"]],
              ["Exceptions", s["exceptions"]], ["Material/escalate", s["material_or_escalate"]],
              ["Value matched", s["value_matched"]], ["Value in exception", s["value_in_exception"]],
              [], ["Category", "#", "Value"]]:
        ws.append(r)
    for c, d in sorted(s["by_category"].items(), key=lambda kv: -kv[1]["val"]):
        ws.append([c, d["n"], _quantize_money(d["val"])])

    we = wb.create_sheet("Exceptions")
    we.append(["Category", "Magnitude", "Currency", "Materiality", "Age (days)",
               "Probable cause", "Suggested action",
               f"{a_label} row", f"{b_label} row", "Status"])
    for e in exceptions:
        we.append([e["category"], _quantize_money(e["magnitude"], e.get("currency")),
                   e.get("currency") or "", e["materiality"],
                   e.get("age_days") if e.get("age_days") is not None else "",
                   e["probable_cause"],
                   e["suggested_action"], str(e["a"]) if e["a"] else "", str(e["b"]) if e["b"] else "",
                   e["status"]])

    wm = wb.create_sheet("Matched")
    wm.append([f"{a_label} row", f"{b_label} row", "Diff"])
    for m in res["matched"]:
        wm.append([str(m["a"]["row"]), str(m["b"]["row"]), m.get("diff", 0.0)])

    if res.get("matched_agg"):
        wa = wb.create_sheet("Aggregations")
        wa.append(["Kind", "Basis", f"{a_label} sum", f"{b_label} sum", "Residual",
                   f"{a_label} items", f"{b_label} items"])
        for p in res["matched_agg"]:
            wa.append([p["kind"], p["basis"], p["a_sum"], p["b_sum"], p["residual"],
                       " | ".join(str(r) for r in p["a_items"]),
                       " | ".join(str(r) for r in p["b_items"])])
    wb.save(out_path)
    return out_path


# --------------------------------------------------------------------------- #
# File wrapper (uses the shared engine) + redaction
# --------------------------------------------------------------------------- #
def _load_engine():
    """Add the shared scripts/ (the toolkit-root engine, or a vendored sibling scripts/)."""
    cands = [
        Path(__file__).resolve().parents[3] / "scripts",            # toolkit-root engine
        Path(__file__).resolve().parent / "scripts",                # vendored sibling
    ]
    for p in cands:
        if (p / "ingest.py").is_file() and str(p) not in sys.path:
            sys.path.insert(0, str(p))
            break
    import ingest, dataclean  # noqa: F401
    return ingest, dataclean


def _records(raw):
    """ingest.read_any returns (list-of-lists, note). Turn it into list[dict] using the
    detected header where possible. (Messy inputs should be tidied with data-tidy first.)"""
    rows = raw[0] if isinstance(raw, tuple) else raw
    rows = [r for r in rows if any(str(c).strip() for c in r)]
    if not rows:
        return []
    try:
        import dataclean  # type: ignore
    except ImportError:
        try:
            _, dataclean = _load_engine()
        except Exception:  # noqa: BLE001
            dataclean = None
    hr = dataclean.detect_header(rows) if dataclean and hasattr(dataclean, "detect_header") else 0
    header = [str(h).strip() for h in rows[hr]]
    out = []
    for r in rows[hr + 1:]:
        r = list(r) + [None] * (len(header) - len(r))
        out.append({header[i]: r[i] for i in range(len(header))})
    return out


def reconcile_files(path_a, path_b, *, preset=None, aggregate=False, group_col=None,
                    party_col=None, date_window=None, auto_confirm=False,
                    sheet_a=None, sheet_b=None, strict_currency=False, **opts):
    """Read A and B in any format (shared `ingest`), match, optionally propose aggregations,
    then triage. Returns (res, exceptions, summary, proposals).

    `sheet_a`/`sheet_b` pick the worksheet for multi-tab `.xlsx` inputs (else ingest selects
    the single data sheet, or raises if several are non-empty). `date_window` (days) is the
    hard matching window in amount_date mode — equal-amount pairs outside it are flagged
    ambiguous, not reconciled (defaults to 5 days when not given).

    Confirm-first: with `aggregate=True` the proposals are returned UN-applied unless
    `auto_confirm=True`. The agent should present `proposals`, get the user's confirmation,
    call `apply_aggregations(res, proposals, accepted)`, then `finalize(res, ...)`.

    Extra keyword options (also settable via a preset): `debit`/`credit` (separate Debit /
    Credit columns -> signed amount), `flip_b` (negate B's amounts — opposite sign
    conventions), `as_of` (age one-sided items in the triage), and `opening_a`/`closing_a` /
    `opening_b`/`closing_b` (statement completeness checks, reported in
    `summary['balance_checks']`)."""
    ingest, _ = _load_engine()
    cfg = dict(PRESETS.get(preset, {})) if preset else {}
    cfg.update({k: v for k, v in opts.items() if v is not None})
    rows_a = _records(ingest.read_any(path_a, sheet=sheet_a))
    rows_b = _records(ingest.read_any(path_b, sheet=sheet_b))
    res = match(rows_a, rows_b, key=cfg.get("key"), amount=cfg.get("amount", "amount"),
                date=cfg.get("date"), currency=cfg.get("currency"), mode=cfg.get("mode", "key"),
                tol=cfg.get("tol", 0.01), strict_currency=strict_currency,
                debit=cfg.get("debit"), credit=cfg.get("credit"),
                flip_b=cfg.get("flip_b", False),
                date_window_days=(date_window if date_window is not None else 5))
    proposals = []
    if aggregate:
        proposals = propose_aggregations(res, group_col=group_col, party_col=party_col,
                                         date_window=date_window, tol=cfg.get("tol", 0.01))
        if auto_confirm and proposals:
            apply_aggregations(res, proposals, range(len(proposals)))
    exc, summary = finalize(res, a_label=cfg.get("a_label", "A"), b_label=cfg.get("b_label", "B"),
                            material=cfg.get("material", 1000.0), escalate=cfg.get("escalate", 10000.0),
                            as_of=cfg.get("as_of"))
    balances = []
    for label, rows, op, cl in ((cfg.get("a_label", "A"), rows_a, cfg.get("opening_a"), cfg.get("closing_a")),
                                (cfg.get("b_label", "B"), rows_b, cfg.get("opening_b"), cfg.get("closing_b"))):
        if op is not None or cl is not None:
            balances.append((label, check_balance(rows, amount=cfg.get("amount", "amount"),
                                                  debit=cfg.get("debit"), credit=cfg.get("credit"),
                                                  opening=op, closing=cl)))
    if balances:
        summary["balance_checks"] = balances
    return res, exc, summary, proposals


def redact(text, terms, amounts=False):
    """Mask party/asset terms (and optionally amounts) before a recon artefact leaves entitled use."""
    out = text
    for t in sorted([t for t in terms if t], key=len, reverse=True):
        out = re.sub(re.escape(t), "[redacted]", out, flags=re.I)
    if amounts:
        out = re.sub(r"-?\d[\d,]*\.?\d*", "[amount]", out)
    return out


def catalogue_md():
    L = ["# Reconciliation presets", ""]
    for k, p in PRESETS.items():
        L.append(f"## `{k}` — {p['label']}")
        L.append(f"- {p['a_label']} (A) vs {p['b_label']} (B); default match: **{p['mode']}**"
                 + (f" on `{p['key']}`" if p.get("key") else " (amount + date)"))
        L.append(f"- {p['note']}")
        L.append("")
    return "\n".join(L)


# --------------------------------------------------------------------------- #
# Self-test (offline; no engine/openpyxl needed)
# --------------------------------------------------------------------------- #
def _self_test():
    A = [
        {"invoice_no": "INV-001", "amount": "1,000.00", "date": "01 Jun 2026"},
        {"invoice_no": "INV-002", "amount": "2,500.00", "date": "02 Jun 2026"},
        {"invoice_no": "INV-003", "amount": "750.00", "date": "03 Jun 2026"},   # not booked
        {"invoice_no": "INV-004", "amount": "12,000.00", "date": "04 Jun 2026"},
        {"invoice_no": "INV-004", "amount": "12,000.00", "date": "04 Jun 2026"},  # duplicate in A
    ]
    B = [
        {"invoice_no": "INV-001", "amount": "1000.00"},
        {"invoice_no": "INV-002", "amount": "2500.05"},   # rounding
        {"invoice_no": "INV-004", "amount": "11500.00"},  # material mismatch 500
        {"invoice_no": "INV-009", "amount": "300.00"},    # booked, not tracked
    ]
    res = match(A, B, key="invoice_no", amount="amount", mode="key")
    exc = triage(res, a_label="Tracker", b_label="Ledger", material=400, escalate=10000)
    cats = sorted({e["category"] for e in exc})
    s = summarise(res, exc)
    print("matched:", len(res["matched"]), "| exceptions:", len(exc), "| categories:", cats)
    print("RAG:", s["rag"], "| % reconciled:", s["pct_reconciled"])
    assert len(res["matched"]) == 1, "INV-001 should match exactly"
    assert "rounding" in cats and "amount_mismatch" in cats, "rounding + mismatch expected"
    assert "missing_in_B" in cats and "missing_in_A" in cats, "both missings expected"
    assert "duplicate" in cats, "duplicate INV-004 in A expected"
    assert any(e["materiality"] == "material" for e in exc), "the 500 diff should be material"

    # aggregation (sum-to-one) + confirm-first
    A2 = [{"party": "Acme", "batch": "BR1", "amount": "8000", "date": "05 Jun 2026"},
          {"party": "Beta", "batch": "BR2", "amount": "2000", "date": "06 Jun 2026"}]
    B2 = [{"party": "Acme", "batch": "BR1", "amount": "3000", "date": "03 Jun 2026"},
          {"party": "Acme", "batch": "BR1", "amount": "5000", "date": "04 Jun 2026"}]
    r2 = match(A2, B2, amount="amount", date="date", mode="amount_date")
    props = propose_aggregations(r2, group_col="batch", party_col="party", date_window=5)
    assert props and abs(props[0]["residual"]) < 0.005, "8000 should tie to 3000+5000"
    e_before, _ = finalize(r2)
    assert len(e_before) == 4, "confirm-first: nothing applied until confirmed"
    apply_aggregations(r2, props, [0])
    e_after, s_after = finalize(r2)
    assert len(r2["matched_agg"]) == 1 and len(e_after) == 1, "after confirm: 1 agg, Beta remains"
    print("aggregation + confirm-first: PASS")

    # amount_date date window is a HARD constraint (gap 2): an equal amount OUTSIDE the window
    # is flagged ambiguous, never silently passed off as a timing difference.
    A3 = [{"amount": "100.00", "date": "01 Jun 2026"}]
    B3 = [{"amount": "100.00", "date": "20 Jun 2026"}]
    r3 = match(A3, B3, amount="amount", date="date", mode="amount_date", date_window_days=5)
    assert not r3["matched"] and len(r3["ambiguous"]) == 1, ("out-of-window must not match", r3)
    assert any(e["category"] == "ambiguous_match" for e in triage(r3)), "expect ambiguous_match"
    r4 = match(A3, B3, amount="amount", date="date", mode="amount_date", date_window_days=30)
    assert len(r4["matched"]) == 1 and not r4["ambiguous"], "within window should match"
    assert r4["matched"][0].get("date_gap") == 19, r4["matched"][0]
    print("amount_date window: PASS")

    # amount_date with NO date column: the window can't be enforced, so equal amounts are held as
    # ambiguous (matched on amount ALONE), never silently reconciled — and the run WARNS.
    And = [{"amount": "3,236.53", "date": "13/06/2026"}]       # dates present but not passed as a column
    Bnd = [{"amount": "3,236.53", "date": "25/06/2026"}]       # 12 days apart — would be a false match
    rnd = match(And, Bnd, amount="amount", mode="amount_date")  # NB: no date= argument
    assert not rnd["matched"] and len(rnd["ambiguous"]) == 1, ("no-date must not match", rnd)
    assert rnd["ambiguous"][0].get("no_date") and rnd["warnings"], rnd
    exc_nd = triage(rnd)
    assert exc_nd and exc_nd[0]["category"] == "ambiguous_match", exc_nd
    assert "amount alone" in exc_nd[0]["probable_cause"], exc_nd[0]
    assert summarise(rnd, exc_nd)["warnings"], "warning must surface in the summary"
    print("amount_date no-date safety: PASS")

    # Absurd-result guard: a near-zero match rate against two sides sharing many equal amounts
    # is a misconfiguration signature, not a real break — the summary must say so out loud.
    n = ABSURD_MIN_POTENTIAL + 10
    Aab = [{"amount": str(100 + i), "date": "05/01/2026"} for i in range(n)]
    Bab = [{"amount": str(100 + i), "date": "05/01/2026"} for i in range(n)]
    # (a) healthy run — same data wired correctly — must stay SILENT
    ok = summarise(match(Aab, Bab, amount="amount", date="date", mode="amount_date"), [])
    assert not ok["unreliable"] and not ok["warnings"], ("healthy run must not warn", ok)
    # (b) date wiring lost -> everything held ambiguous, nothing reconciled -> must FIRE
    bad_res = match(Aab, Bab, amount="amount", mode="amount_date")      # no date=
    bad = summarise(bad_res, triage(bad_res))
    assert bad["unreliable"], ("near-zero match vs equal amounts must warn", bad)
    assert "UNRELIABLE RESULT" in bad["warnings"][0], bad["warnings"]
    assert bad_res["amount_potential"]["equal"] == n, bad_res["amount_potential"]
    # (c) sign-inverted sides -> the message must name the sign convention specifically
    Binv = [{"amount": str(-(100 + i)), "date": "05/01/2026"} for i in range(n)]
    inv_res = match(Aab, Binv, amount="amount", date="date", mode="amount_date")
    inv = summarise(inv_res, triage(inv_res))
    assert inv["unreliable"] and "SIGN-INVERTED" in inv["warnings"][0], inv["warnings"]
    # (d) small files stay quiet — a low match rate there is unremarkable
    Asm = [{"amount": "10", "date": "05/01/2026"}, {"amount": "20", "date": "05/01/2026"}]
    sm_res = match(Asm, [{"amount": "10", "date": "30/06/2026"}], amount="amount",
                   date="date", mode="amount_date")
    assert not summarise(sm_res, triage(sm_res))["unreliable"], "must not warn on tiny inputs"
    print("absurd-result guard: PASS")

    # amount_date residue triage (Stage 1.5): duplicate / sign_flip / amount_mismatch are
    # classified, not left as bare missing_in_A / missing_in_B (a bank-vs-cashbook staple).
    Ar = [
        {"amount": "1000.00", "date": "01 Jun 2026"},   # matches Br[0]
        {"amount": "1000.00", "date": "01 Jun 2026"},   # DUPLICATE of the above (cashbook double-entry)
        {"amount": "3292.08", "date": "05 Jun 2026"},   # sign flip vs Br[1] (+ vs -)
        {"amount": "-2400.00", "date": "07 Jun 2026"},  # net; Br[2] is gross (9% GST) -> amount_mismatch
    ]
    Br = [
        {"amount": "1000.00", "date": "01 Jun 2026"},
        {"amount": "-3292.08", "date": "05 Jun 2026"},
        {"amount": "-2616.00", "date": "07 Jun 2026"},  # -2400 * 1.09
    ]
    rr = match(Ar, Br, amount="amount", date="date", mode="amount_date")
    cats_r = sorted({e["category"] for e in triage(rr)})
    assert len(rr["matched"]) == 1, rr["matched"]
    assert "duplicate" in cats_r, ("expect the cashbook duplicate", cats_r)
    assert "sign_flip" in cats_r, ("expect the sign flip", cats_r)
    assert "amount_mismatch" in cats_r, ("expect the GST net-vs-gross mismatch", cats_r)
    assert not any(e["category"] in ("missing_in_A", "missing_in_B") for e in triage(rr)), \
        ("residue should be classified, not left one-sided", triage(rr))
    print("amount_date residue triage: PASS")

    # Decimal exactness: amounts tie without binary-float drift (a float recon can break 0.1+0.2)
    assert to_amount("0.1") + to_amount("0.2") == to_amount("0.3"), "Decimal must be exact"
    print("Decimal amounts: PASS")

    # currency-aware matching: 100 USD must NOT reconcile against 100 SGD
    Ac = [{"ref": "R1", "amount": "100.00", "ccy": "USD"},
          {"ref": "R2", "amount": "50.00", "ccy": "SGD"}]
    Bc = [{"ref": "R1", "amount": "100.00", "ccy": "SGD"},   # same key, wrong currency
          {"ref": "R2", "amount": "50.00", "ccy": "SGD"}]
    rc = match(Ac, Bc, key="ref", amount="amount", currency="ccy", mode="key")
    assert len(rc["matched"]) == 1 and len(rc["currency_diffs"]) == 1, rc
    assert any(e["category"] == "currency_mismatch" for e in triage(rc)), "expect currency_mismatch"
    # code detected from the amount cell's symbol when there's no currency column
    rs = match([{"amount": "US$ 100"}], [{"amount": "S$ 100"}], amount="amount", mode="amount_date")
    assert not rs["matched"] and not rs["ambiguous"], "US$100 must not match S$100"
    assert to_currency("$100") is None, "bare $ stays ambiguous (not USD)"
    print("currency-aware matching: PASS")

    # strict currency: unknown currency must NOT match in strict mode (audit/finance)
    Au = [{"ref": "R1", "amount": "100.00"}]          # no currency, none detectable
    Bu = [{"ref": "R1", "amount": "100.00"}]
    assert len(match(Au, Bu, key="ref", amount="amount")["matched"]) == 1, "permissive: unknown matches"
    rstrict = match(Au, Bu, key="ref", amount="amount", strict_currency=True)
    assert not rstrict["matched"] and len(rstrict["currency_unknown"]) == 1, rstrict
    assert any(e["category"] == "currency_unknown" for e in triage(rstrict)), "expect currency_unknown"
    print("strict currency: PASS")
    print("self-test: PASS")
    return 0


def main(argv=None):
    ap = argparse.ArgumentParser(description="Reconcile A <-> B and triage the discrepancies.")
    ap.add_argument("a", nargs="?", help="source A (ours)")
    ap.add_argument("b", nargs="?", help="source B (theirs)")
    ap.add_argument("--preset", choices=list(PRESETS), help="recurring-reconciliation preset")
    ap.add_argument("--key", help="match key column (override)")
    ap.add_argument("--amount", default="amount")
    ap.add_argument("--debit", help="debit column (bank/GL exports with separate Debit/Credit "
                                    "columns; signed amount = debit - credit)")
    ap.add_argument("--credit", help="credit column (see --debit)")
    ap.add_argument("--flip-b", action="store_true",
                    help="negate B's amounts before comparing (sides book the same money with "
                         "opposite signs, e.g. bank statement vs GL cash account)")
    ap.add_argument("--date")
    ap.add_argument("--as-of", help="ageing date (e.g. today) — one-sided exceptions with a "
                                    "parsed date gain age_days in the working paper")
    ap.add_argument("--opening-a", help="opening balance of side A — enables the "
                                        "opening + movement = closing completeness check")
    ap.add_argument("--closing-a", help="stated closing balance of side A")
    ap.add_argument("--opening-b", help="opening balance of side B")
    ap.add_argument("--closing-b", help="stated closing balance of side B")
    ap.add_argument("--currency", help="currency column (else the code is read from the amount cell); "
                                       "currencies are compared so 100 USD != 100 SGD")
    ap.add_argument("--strict-currency", action="store_true",
                    help="audit/finance mode: don't match when a side's currency is unknown "
                         "(routes to currency_unknown instead of assuming compatibility)")
    ap.add_argument("--mode", choices=["key", "amount_date"])
    ap.add_argument("--material", type=str, default="1000.0")
    ap.add_argument("--escalate", type=str, default="10000.0")
    ap.add_argument("--aggregate", action="store_true", help="propose sum-to-one / sum-to-sum matches")
    ap.add_argument("--group-col", help="shared reference/batch column bounding aggregation")
    ap.add_argument("--party-col", help="counterparty/account column bounding aggregation")
    ap.add_argument("--date-window", type=int,
                    help="amount_date matching + aggregation window, +/- N days (default 5 for "
                         "matching); equal-amount pairs outside it are flagged ambiguous, not matched")
    ap.add_argument("--sheet-a", help="worksheet name for a multi-tab .xlsx source A")
    ap.add_argument("--sheet-b", help="worksheet name for a multi-tab .xlsx source B")
    ap.add_argument("--auto-confirm", action="store_true",
                    help="accept all aggregation proposals without review (headless/testing only)")
    ap.add_argument("--out", help="write the .xlsx working paper here")
    ap.add_argument("--catalogue", action="store_true", help="print the presets catalogue")
    ap.add_argument("--self-test", action="store_true")
    a = ap.parse_args(argv)
    if a.self_test:
        return _self_test()
    if a.catalogue:
        print(catalogue_md())
        return 0
    if not (a.a and a.b):
        ap.error("need source A and B (or --self-test / --catalogue)")
    res, exc, summary, proposals = reconcile_files(
        a.a, a.b, preset=a.preset, key=a.key, amount=a.amount, date=a.date,
        currency=a.currency, mode=a.mode, material=a.material, escalate=a.escalate,
        aggregate=a.aggregate, group_col=a.group_col, party_col=a.party_col,
        date_window=a.date_window, auto_confirm=a.auto_confirm,
        sheet_a=a.sheet_a, sheet_b=a.sheet_b, strict_currency=a.strict_currency,
        debit=a.debit, credit=a.credit, flip_b=(a.flip_b or None), as_of=a.as_of,
        opening_a=a.opening_a, closing_a=a.closing_a,
        opening_b=a.opening_b, closing_b=a.closing_b)
    p = PRESETS.get(a.preset, {})
    al, bl = p.get("a_label", "A"), p.get("b_label", "B")
    if proposals and not a.auto_confirm:
        print(render_proposals(proposals, a_label=al, b_label=bl))
        print()
    print(render_report(summary, exc, a_label=al, b_label=bl, title=p.get("label", "Reconciliation"),
                        balances=summary.get("balance_checks")))
    if a.out:
        write_workpaper(res, exc, summary, a.out, a_label=al, b_label=bl)
        print(f"\nworking paper -> {a.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
