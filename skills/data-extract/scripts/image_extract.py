"""Vision-model image → structured table extraction for data-extract.

Accepts chart / table / UI / diagram / general images (``.png`` ``.jpg`` ``.jpeg``
``.gif`` ``.webp`` ``.bmp``), classifies them, calls an OpenAI-compatible vision
endpoint, parses any Markdown table into a DataFrame, and writes a styled
``.xlsx``. Batch mode processes a directory into one workbook (one sheet per
image, plus a combined sheet with a ``source`` column).

Does **not** fall back to Tesseract for chart data — Tesseract cannot read
charts. If no vision endpoint/key is configured the script exits clearly.

    python skills/data-extract/scripts/image_extract.py chart.png -o out.xlsx
    python skills/data-extract/scripts/image_extract.py ./shots/ -o batch.xlsx

Env / flags: ``--model`` / ``VISION_MODEL``, ``--api-key`` / ``VISION_API_KEY``
or ``OPENAI_API_KEY``, ``--base-url`` / ``VISION_BASE_URL`` (default
``https://api.openai.com/v1``).
"""

from __future__ import annotations

import argparse
import base64
import hashlib
import io
import json
import os
import re
import sys
import time
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}
MAX_BYTES = 5 * 1024 * 1024
MAX_PX = 2048

PROMPTS = {
    "chart": (
        "Extract the chart title, axis labels, legend, and every data point value. "
        "Output as a Markdown table. Do not round numbers."
    ),
    "table": (
        "Extract all table content as a Markdown table. Preserve row/column structure. "
        "Do not round numbers."
    ),
    "ui": (
        "Describe from a frontend developer's perspective: layout, components, text, colours."
    ),
    "diagram": (
        "Describe all nodes and connections (A→B), including branch conditions."
    ),
    "general": (
        "Describe the image clearly. If any tabular or numeric data is visible, "
        "also output it as a Markdown table. Do not round numbers."
    ),
}

_TYPE_HINTS = (
    ("chart", ("chart", "graph", "plot", "bar", "line", "pie", "scatter", "histogram")),
    ("table", ("table", "grid", "spreadsheet", "ledger", "screenshot-table")),
    ("ui", ("ui", "screenshot", "dashboard", "mockup", "wireframe", "app")),
    ("diagram", ("diagram", "flowchart", "flow", "uml", "architecture", "node")),
)


# --------------------------------------------------------------------------- #
# Classification + prompts
# --------------------------------------------------------------------------- #
def classify_image(path: str, caption_hint: str | None = None) -> str:
    """Classify by filename hints (and optional light caption). Default: general."""
    name = Path(path).stem.lower().replace("-", " ").replace("_", " ")
    blob = f"{name} {caption_hint or ''}".lower()
    for kind, hints in _TYPE_HINTS:
        if any(h in blob for h in hints):
            return kind
    return "general"


def prompt_for(kind: str) -> str:
    return PROMPTS.get(kind, PROMPTS["general"])


# --------------------------------------------------------------------------- #
# Image prep + cache
# --------------------------------------------------------------------------- #
def _file_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for block in iter(lambda: f.read(1 << 20), b""):
            h.update(block)
    return h.hexdigest()


def compress_image(path: str | Path, max_bytes: int = MAX_BYTES, max_px: int = MAX_PX):
    """Resize / recompress when over size or dimension limits. Returns (bytes, mime, compressed)."""
    path = Path(path)
    raw = path.read_bytes()
    suffix = path.suffix.lower()
    mime = {
        ".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
        ".gif": "image/gif", ".webp": "image/webp", ".bmp": "image/bmp",
    }.get(suffix, "application/octet-stream")

    need = len(raw) > max_bytes
    try:
        from PIL import Image
    except ImportError:
        if need:
            raise ImportError(
                "Pillow is required to compress images >5MB. Install with: pip install Pillow"
            )
        return raw, mime, False

    img = Image.open(io.BytesIO(raw))
    w, h = img.size
    if max(w, h) > max_px:
        need = True
        scale = max_px / float(max(w, h))
        img = img.resize((max(1, int(w * scale)), max(1, int(h * scale))), Image.Resampling.LANCZOS)

    if not need:
        return raw, mime, False

    # Prefer JPEG for photos/screenshots to hit the byte budget; keep PNG for graphics with alpha.
    out = io.BytesIO()
    has_alpha = img.mode in ("RGBA", "LA") or (img.mode == "P" and "transparency" in img.info)
    if has_alpha and len(raw) <= max_bytes * 1.5:
        img = img.convert("RGBA")
        img.save(out, format="PNG", optimize=True)
        return out.getvalue(), "image/png", True

    rgb = img.convert("RGB")
    quality = 85
    while quality >= 40:
        out = io.BytesIO()
        rgb.save(out, format="JPEG", quality=quality, optimize=True)
        if out.tell() <= max_bytes:
            break
        quality -= 10
    return out.getvalue(), "image/jpeg", True


def _cache_dir(base: Path | None = None) -> Path:
    d = base or (Path.home() / ".cache" / "data-toolkit" / "image_extract")
    d.mkdir(parents=True, exist_ok=True)
    return d


def _cache_key(file_hash: str, prompt: str, model: str) -> str:
    ph = hashlib.sha256(f"{prompt}|{model}".encode()).hexdigest()
    return f"{file_hash[:16]}_{ph[:16]}.json"


def cache_get(file_hash: str, prompt: str, model: str, cache_dir: Path | None = None):
    p = _cache_dir(cache_dir) / _cache_key(file_hash, prompt, model)
    if p.is_file():
        return json.loads(p.read_text(encoding="utf-8"))
    return None


def cache_put(file_hash: str, prompt: str, model: str, payload: dict, cache_dir: Path | None = None):
    p = _cache_dir(cache_dir) / _cache_key(file_hash, prompt, model)
    p.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return p


# --------------------------------------------------------------------------- #
# Vision API (OpenAI-compatible)
# --------------------------------------------------------------------------- #
def vision_config(api_key=None, base_url=None, model=None):
    key = api_key or os.environ.get("VISION_API_KEY") or os.environ.get("OPENAI_API_KEY")
    url = (base_url or os.environ.get("VISION_BASE_URL") or "https://api.openai.com/v1").rstrip("/")
    mdl = model or os.environ.get("VISION_MODEL") or "gpt-4o"
    return key, url, mdl


def call_vision(
    image_bytes: bytes,
    mime: str,
    prompt: str,
    *,
    api_key=None,
    base_url=None,
    model=None,
    timeout: int = 120,
    _request=None,
) -> dict:
    """Call an OpenAI-compatible chat-completions vision endpoint.

    Returns ``{"description": str, "usage": dict, "model": str}``.
    Retries once on transient failure; raises on permanent failure.
    """
    import requests as _requests

    req = _request or _requests.post
    key, url, mdl = vision_config(api_key, base_url, model)
    if not key:
        raise RuntimeError(
            "No vision API key configured. Set VISION_API_KEY or OPENAI_API_KEY "
            "(or pass --api-key). Image/chart extraction requires a vision-capable "
            "OpenAI-compatible endpoint — it will not fall back to Tesseract."
        )

    b64 = base64.b64encode(image_bytes).decode("ascii")
    payload = {
        "model": mdl,
        "messages": [
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {
                        "type": "image_url",
                        "image_url": {"url": f"data:{mime};base64,{b64}"},
                    },
                ],
            }
        ],
        "temperature": 0,
    }
    headers = {"Authorization": f"Bearer {key}", "Content-Type": "application/json"}
    endpoint = f"{url}/chat/completions"

    last_err = None
    for attempt in range(2):
        try:
            resp = req(endpoint, headers=headers, json=payload, timeout=timeout)
            if resp.status_code >= 500 or resp.status_code == 429:
                last_err = RuntimeError(f"transient vision API error HTTP {resp.status_code}: {resp.text[:300]}")
                if attempt == 0:
                    time.sleep(1.5)
                    continue
                raise last_err
            if resp.status_code >= 400:
                raise RuntimeError(
                    f"vision API error HTTP {resp.status_code}: {resp.text[:500]}"
                )
            data = resp.json()
            text = data["choices"][0]["message"]["content"]
            usage = data.get("usage") or {}
            return {"description": text, "usage": usage, "model": data.get("model", mdl)}
        except (RuntimeError, KeyError, ValueError, TypeError):
            raise
        except Exception as e:  # noqa: BLE001 — network / timeout
            last_err = e
            if attempt == 0:
                time.sleep(1.5)
                continue
            raise RuntimeError(f"vision API call failed: {e}") from e
    raise RuntimeError(f"vision API call failed: {last_err}")


# --------------------------------------------------------------------------- #
# Markdown table parser
# --------------------------------------------------------------------------- #
_SEP_RE = re.compile(r"^\s*\|?\s*:?-{3,}:?\s*(\|\s*:?-{3,}:?\s*)+\|?\s*$")
_CURRENCY_RE = re.compile(r"^[\s$£€¥₹₽₩]*(US\$|S\$|A\$|HK\$|NZ\$|C\$|R\$)?")


def _split_row(line: str) -> list[str]:
    """Split a Markdown table row on unescaped pipes."""
    s = line.strip()
    if s.startswith("|"):
        s = s[1:]
    if s.endswith("|"):
        s = s[:-1]
    cells, buf, i = [], [], 0
    while i < len(s):
        if s[i] == "\\" and i + 1 < len(s):
            buf.append(s[i + 1])
            i += 2
            continue
        if s[i] == "|":
            cells.append("".join(buf).strip())
            buf = []
            i += 1
            continue
        buf.append(s[i])
        i += 1
    cells.append("".join(buf).strip())
    return cells


def _auto_number(v: str):
    """Convert currency / percent / comma-thousands strings; else return original."""
    if v is None:
        return None
    s = str(v).strip()
    if s == "" or s.lower() in {"none", "null", "nan", "—", "-", "–"}:
        return None
    pct = s.endswith("%")
    raw = s[:-1].strip() if pct else s
    # Strip currency symbols / codes at ends
    raw = re.sub(r"^(US\$|S\$|A\$|HK\$|NZ\$|C\$|R\$|\$|£|€|¥|₹)", "", raw).strip()
    raw = re.sub(r"(USD|SGD|GBP|EUR|AUD|HKD|JPY)$", "", raw, flags=re.I).strip()
    raw = raw.replace(",", "")
    neg = False
    if raw.startswith("(") and raw.endswith(")"):
        neg, raw = True, raw[1:-1]
    try:
        num = float(raw)
    except ValueError:
        return v
    if neg:
        num = -num
    if pct:
        num = num / 100.0
    if num == int(num) and not pct and "." not in str(v).replace(",", ""):
        return int(num)
    return num


def parse_markdown_table(text: str):
    """Extract the first Markdown table from ``text`` → DataFrame, or ``None``.

    Handles pipe escaping, strips the separator row, and auto-converts numeric
    columns (comma thousands, ``%`` suffixes, currency symbols).
    """
    if not text or "|" not in text:
        return None

    lines = text.splitlines()
    # Find a header + separator block
    start = None
    for i in range(len(lines) - 1):
        if "|" in lines[i] and _SEP_RE.match(lines[i + 1] or ""):
            start = i
            break
    if start is None:
        # Some models omit the separator — accept ≥2 pipe-rows as a table.
        pipe_rows = [ln for ln in lines if "|" in ln and not _SEP_RE.match(ln)]
        if len(pipe_rows) < 2:
            return None
        header = _split_row(pipe_rows[0])
        body = [_split_row(r) for r in pipe_rows[1:]]
    else:
        header = _split_row(lines[start])
        body = []
        for ln in lines[start + 2:]:
            if "|" not in ln:
                if body:
                    break
                continue
            if _SEP_RE.match(ln):
                continue
            body.append(_split_row(ln))

    if not header or not body:
        return None

    width = len(header)
    rows = []
    for r in body:
        if len(r) < width:
            r = r + [""] * (width - len(r))
        elif len(r) > width:
            r = r[:width]
        rows.append([_auto_number(c) for c in r])

    try:
        import pandas as pd
    except ImportError:
        # Minimal stand-in so callers without pandas still get structure.
        class _Mini:
            def __init__(self, columns, data):
                self.columns = list(columns)
                self._data = data

            def __len__(self):
                return len(self._data)

            def to_dict(self, orient="records"):
                return [dict(zip(self.columns, row)) for row in self._data]

        return _Mini(header, rows)

    import pandas as pd
    df = pd.DataFrame(rows, columns=header)
    # Second pass: if a column is mostly numeric, keep it numeric (None for blanks).
    for col in df.columns:
        converted = df[col].map(
            lambda x: x if isinstance(x, (int, float)) or x is None else _auto_number(x)
        )
        nums = sum(isinstance(v, (int, float)) for v in converted)
        if nums >= max(1, int(0.6 * len(converted))):
            df[col] = converted
    return df


# --------------------------------------------------------------------------- #
# XLSX export
# --------------------------------------------------------------------------- #
def write_styled_xlsx(sheets: dict, out_path: str) -> str:
    """Write ``{sheet_name: DataFrame}`` to a styled workbook (bold white-on-blue headers)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")

    for name, df in sheets.items():
        title = re.sub(r"[\[\]\*\?/\\:]", "_", str(name))[:31] or "Sheet"
        ws = wb.create_sheet(title)
        cols = list(df.columns)
        ws.append([str(c) for c in cols])
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for _, row in df.iterrows():
            ws.append([("" if v is None or (isinstance(v, float) and v != v) else v) for v in row.tolist()])
        for i, col in enumerate(cols, 1):
            width = max(len(str(col)), *(len(str(v)) if v is not None else 0 for v in df[col].tolist()[:200]))
            ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 10), 48)
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


# --------------------------------------------------------------------------- #
# Orchestration
# --------------------------------------------------------------------------- #
def extract_image(
    path: str,
    *,
    image_type: str | None = None,
    api_key=None,
    base_url=None,
    model=None,
    cache_dir: Path | None = None,
    force: bool = False,
    _request=None,
) -> dict:
    """Extract structured data from one image.

    Returns JSON-shaped dict: ``{file, type, description, usage, cached, dataframe?, error?}``.
    """
    path = str(path)
    p = Path(path)
    if not p.is_file():
        return {"file": path, "type": None, "description": None, "usage": {},
                "cached": False, "error": f"not a file: {path}"}

    kind = image_type or classify_image(path)
    prompt = prompt_for(kind)
    key, url, mdl = vision_config(api_key, base_url, model)
    file_hash = _file_sha256(p)

    if not force:
        hit = cache_get(file_hash, prompt, mdl, cache_dir=cache_dir)
        if hit is not None:
            hit = dict(hit)
            hit["cached"] = True
            hit["file"] = path
            if hit.get("description"):
                hit["dataframe"] = parse_markdown_table(hit["description"])
            return hit

    if not key:
        return {
            "file": path, "type": kind, "description": None, "usage": {},
            "cached": False,
            "error": (
                "No vision API key configured. Set VISION_API_KEY or OPENAI_API_KEY. "
                "Will not fall back to Tesseract for chart/image data."
            ),
        }

    try:
        img_bytes, mime, compressed = compress_image(p)
        result = call_vision(
            img_bytes, mime, prompt,
            api_key=key, base_url=url, model=mdl, _request=_request,
        )
    except Exception as e:  # noqa: BLE001
        return {"file": path, "type": kind, "description": None, "usage": {},
                "cached": False, "error": str(e), "compressed": False}

    payload = {
        "file": path,
        "type": kind,
        "description": result["description"],
        "usage": result.get("usage") or {},
        "cached": False,
        "model": result.get("model", mdl),
        "compressed": compressed,
    }
    cache_put(file_hash, prompt, mdl, {k: v for k, v in payload.items() if k != "dataframe"},
              cache_dir=cache_dir)
    payload["dataframe"] = parse_markdown_table(result["description"])
    return payload


def extract_batch(
    source: str,
    out_path: str,
    *,
    api_key=None,
    base_url=None,
    model=None,
    cache_dir: Path | None = None,
    force: bool = False,
    combine: bool = True,
    _request=None,
) -> dict:
    """Process one image or a directory → styled ``.xlsx``. Returns a summary dict."""
    src = Path(source)
    if src.is_dir():
        files = sorted(p for p in src.iterdir() if p.suffix.lower() in IMAGE_EXTS and p.is_file())
    elif src.is_file() and src.suffix.lower() in IMAGE_EXTS:
        files = [src]
    else:
        raise ValueError(f"Not an image or image directory: {source}")

    if not files:
        raise ValueError(f"No images found in {source}")

    try:
        import pandas as pd
    except ImportError as e:
        raise ImportError("pandas is required for image_extract DataFrame export") from e

    sheets = {}
    results = []
    combined_frames = []
    for f in files:
        r = extract_image(
            str(f), api_key=api_key, base_url=base_url, model=model,
            cache_dir=cache_dir, force=force, _request=_request,
        )
        results.append({k: v for k, v in r.items() if k != "dataframe"})
        df = r.get("dataframe")
        sheet_name = f.stem[:28]
        if df is not None and len(df) > 0:
            sheets[sheet_name] = df
            if combine:
                c = df.copy()
                c.insert(0, "source", f.name)
                combined_frames.append(c)
        else:
            # Keep a description sheet so the workbook still records the result.
            desc = r.get("description") or r.get("error") or "(no table extracted)"
            sheets[sheet_name] = pd.DataFrame({"description": [desc]})

    if combine and combined_frames:
        sheets["combined"] = pd.concat(combined_frames, ignore_index=True)

    write_styled_xlsx(sheets, out_path)
    return {
        "out": out_path,
        "count": len(files),
        "results": results,
        "sheets": list(sheets.keys()),
    }


def main(argv=None):
    ap = argparse.ArgumentParser(description="Extract structured data from images via a vision model")
    ap.add_argument("source", help="Image file or directory of images")
    ap.add_argument("-o", "--output", default="extracted_images.xlsx", help="Output .xlsx path")
    ap.add_argument("--model", default=None, help="Vision model id (or VISION_MODEL)")
    ap.add_argument("--api-key", default=None, help="API key (or VISION_API_KEY / OPENAI_API_KEY)")
    ap.add_argument("--base-url", default=None, help="OpenAI-compatible base URL (or VISION_BASE_URL)")
    ap.add_argument("--cache-dir", default=None, help="Override result cache directory")
    ap.add_argument("--force", action="store_true", help="Ignore cache")
    ap.add_argument("--type", dest="image_type", default=None,
                    choices=list(PROMPTS), help="Force image type / prompt")
    ap.add_argument("--json", action="store_true", help="Print JSON summary to stdout")
    args = ap.parse_args(argv)

    cache = Path(args.cache_dir) if args.cache_dir else None
    key, url, mdl = vision_config(args.api_key, args.base_url, args.model)
    if not key:
        print(
            "ERROR: No vision API key. Set VISION_API_KEY or OPENAI_API_KEY "
            "(or pass --api-key). Will not fall back to Tesseract for chart data.",
            file=sys.stderr,
        )
        return 2

    try:
        summary = extract_batch(
            args.source, args.output,
            api_key=key, base_url=url, model=mdl,
            cache_dir=cache, force=args.force,
        )
    except Exception as e:  # noqa: BLE001
        print(f"ERROR: {e}", file=sys.stderr)
        return 1

    if args.json:
        print(json.dumps(summary, ensure_ascii=False, indent=2, default=str))
    else:
        print(f"Wrote {summary['out']} ({summary['count']} image(s), sheets: {', '.join(summary['sheets'])})")
        for r in summary["results"]:
            flag = "cached" if r.get("cached") else ("error" if r.get("error") else "ok")
            print(f"  [{flag}] {r.get('file')} type={r.get('type')}")
            if r.get("error"):
                print(f"           {r['error']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
