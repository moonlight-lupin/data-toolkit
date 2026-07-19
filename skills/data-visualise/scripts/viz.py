"""viz.py — a brandable, self-contained HTML dashboard / visualisation engine.

Turns any tabular data — a list of dicts, or an existing .xlsx — into a
self-contained HTML artefact that opens in a browser and prints cleanly to PDF.
Ships **unbranded** — a neutral "Data Toolkit" wordmark (no logo image), a teal /
cool-paper palette and a clean type pairing — and lets any firm drop in its own
brand (name, colours, fonts, logo) without touching the code.

Design principles
-----------------
* **Self-contained & offline.** Pure inline HTML + CSS + **inline SVG** charts.
  NO JavaScript chart libraries, NO CDN, NO remote images. One file, no network.
  This keeps sensitive/confidential data off any cloud (the toolkit's
  data-handling rule) and makes the file print reliably and travel as a single
  attachment.
* **Brandable.** A neutral default is built in; pass a `theme` dict (brand name,
  primary/accent colours, fonts, optional logo path) to re-skin every artefact.
  See DEFAULT_THEME below.
* **Composable.** Small building blocks (kpi_card, bar_chart, line_chart,
  donut_chart, heatmap, sparkline, waterfall, table, status_pill, section, grid)
  each return an HTML string; `dashboard(...)` assembles them into a full page
  with a header, an "as-of" stamp, a print stylesheet and a footer disclaimer.
  `suggest_blocks_from_analysis` turns a data-analyse `analysis.json` into
  editable declarative blocks (compute stays in analyse; this skill only draws).
* **Drafts, not advice.** Output is a draft artefact for a qualified person
  to review — never auto-distributed.

Quick start
-----------
    from viz import kpi_row, bar_chart, table, dashboard
    blocks = [
        kpi_row([{"label": "Open tasks", "value": 12, "status": "amber"},
                 {"label": "Overdue", "value": 3, "status": "red"}]),
        bar_chart([("Mon", 4), ("Tue", 7), ("Wed", 5)], title="Tasks by day"),
        table(rows, title="Detail"),
    ]
    dashboard("Operations dashboard", blocks, subtitle="Weekly", as_of="14 Jun 2026",
              out_path="dashboard.html")

To apply your own brand, pass a theme (any subset overrides the neutral default):
    my_theme = {"brand_name": "Acme Co",
                "colours": {"burgundy": "#0B3D91", "rose": "#1565C0"},
                "logo_path": "assets/acme-logo.png"}
    dashboard("Operations dashboard", blocks, theme=my_theme, out_path="dashboard.html")

Run `python viz.py` for an offline self-test that builds a demo dashboard.
"""

from __future__ import annotations

import base64 as _b64
import datetime as dt
from decimal import Decimal
import html as _html
import os
import sys
import webbrowser
from pathlib import Path

# Windows cp1252 consoles can't print non-ASCII (·, —, ⚠) — guard it.
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass


# --------------------------------------------------------------------------- #
# Theme — a NEUTRAL, white-label default. The toolkit ships unbranded: a generic
# "Data Toolkit" wordmark (no logo image), a teal/paper palette and a clean type
# pairing. A firm re-skins every artefact by passing a `theme` dict to dashboard()
# (or calling apply_theme before building blocks) — name, colours, fonts, logo.
# Colour-token *names* (burgundy / rose / pink …) are historical and stay stable
# across the rendering code. See references/brand.md.
# --------------------------------------------------------------------------- #
DEFAULT_THEME = {
    # Shown in the header text wordmark and footer. Neutral by default.
    "brand_name": "Data Toolkit",
    # No default logo image — the header shows a text wordmark of brand_name.
    # A firm supplies its own via theme["logo_path"] (a transparent PNG or SVG).
    "logo_path": None,
    # Font stacks — a clean sans body with a geometric-display heading. The dashboard
    # is self-contained (no CDN), so these load only if the reader has them installed
    # and otherwise fall back gracefully down the stack. When labels contain CJK or
    # other non-Latin scripts, dashboard() augments the stack (see _augment_font_stack).
    "font": "'Inter','Segoe UI','Helvetica Neue',Arial,sans-serif",
    "font_heading": "'Space Grotesk','Inter','Segoe UI',system-ui,sans-serif",
    # Palette — the toolkit's neutral default scheme (a teal accent on cool paper).
    # Token names are historical (kept stable so existing themes keep working).
    "colours": {
        "burgundy": "#163F3A",   # primary — deep teal
        "rose":     "#4FB3A0",   # accent 1 — bright teal
        "pink":     "#20574F",   # accent 2 — soft teal
        "pink_lt":  "#A7D9CF",   # light teal tint
        "pink_vlt": "#E7EBE9",   # zebra striping (very light)
        "ink":      "#14171A",   # body text
        "grey":     "#565C63",   # muted text
        "grey_faint": "#8C9298", # faintest text
        "grey_lt":  "#D9DEDB",   # hairlines
        "bg":       "#F1F3F2",   # page background (paper)
        "white":    "#FFFFFF",   # cards
        # status (RAG)
        "green":    "#2E7D57",
        "amber":    "#B26B00",
        "red":      "#9B2226",
    },
}

# Active module-level brand state, initialised from DEFAULT_THEME. Building
# blocks read these; apply_theme() rebinds them so block colours follow a brand.
BRAND = dict(DEFAULT_THEME["colours"])
FONT = DEFAULT_THEME["font"]
FONT_HEADING = DEFAULT_THEME["font_heading"]
BRAND_NAME = DEFAULT_THEME["brand_name"]
# Path to the header logo (none by default — text wordmark; a firm supplies its own via theme).
LOGO_PATH = Path(DEFAULT_THEME["logo_path"]) if DEFAULT_THEME["logo_path"] else None


def _series_palette() -> list:
    """Sequence used to colour multi-series charts / donut slices, from BRAND."""
    return [BRAND["burgundy"], BRAND["rose"], BRAND["pink"], BRAND["green"],
            BRAND["amber"], BRAND["grey"], BRAND["pink_lt"]]


def _resolve_theme(theme: dict | None) -> dict:
    """Merge a partial `theme` dict over DEFAULT_THEME → a complete theme.
    `theme` may set any of: brand_name, logo_path, font, font_heading, colours (partial).
    `font_heading` defaults to `font` when a theme sets only `font`, so a brand that gives
    one font stack gets it everywhere (no stray Space Grotesk leaking into their headings)."""
    t = theme or {}
    colours = dict(DEFAULT_THEME["colours"])
    colours.update(t.get("colours") or {})
    font = t.get("font", DEFAULT_THEME["font"])
    heading_default = font if "font" in t else DEFAULT_THEME["font_heading"]
    return {
        "brand_name": t.get("brand_name", DEFAULT_THEME["brand_name"]),
        "logo_path": t.get("logo_path", DEFAULT_THEME["logo_path"]),
        "font": font,
        "font_heading": t.get("font_heading", heading_default),
        "colours": colours,
    }


def apply_theme(theme: dict | None) -> dict:
    """Rebind the module-level brand state (BRAND/FONT/BRAND_NAME/LOGO_PATH and the
    chart series palette) from a (partial) `theme` dict, so building blocks built
    afterwards pick up the brand. Returns the fully-resolved theme. Pass None to
    reset to the neutral default. Call this BEFORE composing blocks if you want a
    firm's colours in the charts; dashboard() also accepts `theme=` for the shell."""
    global BRAND, FONT, FONT_HEADING, BRAND_NAME, LOGO_PATH, _SERIES, _STATUS_COLOUR
    rt = _resolve_theme(theme)
    BRAND = rt["colours"]
    FONT = rt["font"]
    FONT_HEADING = rt["font_heading"]
    BRAND_NAME = rt["brand_name"]
    LOGO_PATH = Path(rt["logo_path"]) if rt["logo_path"] else None
    _SERIES = _series_palette()
    _STATUS_COLOUR = _status_colour_map()
    return rt


def _logo_html() -> str:
    """The brand mark for the header: a logo image as a base64 data URI if a firm has
    supplied one via theme["logo_path"], else a plain text wordmark of the brand name.
    The toolkit ships with no default logo — the neutral default is the wordmark."""
    return _logo_for({"logo_path": str(LOGO_PATH) if LOGO_PATH else "",
                      "brand_name": BRAND_NAME})


def _logo_for(rt: dict) -> str:
    """Logo mark from a resolved theme: the logo PNG as a base64 data URI if present,
    else a plain text wordmark of the brand name."""
    path = rt.get("logo_path")
    name = rt.get("brand_name", "")
    try:
        if path:
            p = Path(path)
            allowed = {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
                       ".gif": "image/gif", ".svg": "image/svg+xml"}
            mime = allowed.get(p.suffix.lower())
            if not mime:
                sys.stderr.write(f"[WARN] logo ignored: unsupported image type {p.suffix or '(none)'}\n")
                return f'<span class="mark">{_e(name)}</span>'
            if p.stat().st_size > 1_000_000:
                sys.stderr.write("[WARN] logo ignored: file exceeds 1MB\n")
                return f'<span class="mark">{_e(name)}</span>'
            uri = f"data:{mime};base64," + _b64.b64encode(p.read_bytes()).decode()
            return f'<img class="logo" src="{uri}" alt="{_e(name)}">'
    except Exception:
        pass
    return f'<span class="mark">{_e(name)}</span>'

# Sequence used to colour multi-series charts / donut slices.
_SERIES = _series_palette()

def _status_colour_map():
    return {
        "green": BRAND["green"], "ok": BRAND["green"], "done": BRAND["green"],
        "amber": BRAND["amber"], "warn": BRAND["amber"], "due": BRAND["amber"],
        "red": BRAND["red"], "bad": BRAND["red"], "overdue": BRAND["red"],
        "brand": BRAND["burgundy"], "info": BRAND["burgundy"],
        "grey": BRAND["grey"], "neutral": BRAND["grey"], None: BRAND["grey"],
    }


_STATUS_COLOUR = _status_colour_map()


_UID = [0]


def _next_id() -> str:
    """A render-local unique id (for pairing a filter control to its table)."""
    _UID[0] += 1
    return f"qv{_UID[0]}"


def _e(s) -> str:
    """HTML-escape anything to a safe string."""
    return _html.escape("" if s is None else str(s))


def _status(s) -> str:
    return _STATUS_COLOUR.get(str(s).lower() if s is not None else None, BRAND["grey"])


def _fmt_num(v) -> str:
    """Thousands-separate plain numbers; leave pre-formatted strings alone."""
    if isinstance(v, bool):
        return _e(v)
    if isinstance(v, Decimal):
        return f"{v:,}".rstrip("0").rstrip(".") if "." in f"{v:f}" else f"{v:,}"
    if isinstance(v, int):
        return f"{v:,}"
    if isinstance(v, float):
        return f"{v:,.2f}".rstrip("0").rstrip(".") if v % 1 else f"{int(v):,}"
    return _e(v)


# --------------------------------------------------------------------------- #
# Building blocks — each returns a self-contained HTML fragment.
# --------------------------------------------------------------------------- #
def kpi_card(label, value, sub=None, status="brand") -> str:
    """A single metric card: big value, label, optional sub-line, status accent."""
    col = _status(status)
    sub_h = f'<div class="kpi-sub">{_e(sub)}</div>' if sub else ""
    return (f'<div class="kpi" style="border-top-color:{col}">'
            f'<div class="kpi-val" style="color:{col}">{_fmt_num(value)}</div>'
            f'<div class="kpi-lbl">{_e(label)}</div>{sub_h}</div>')


def kpi_row(cards: list[dict]) -> str:
    """A responsive row of KPI cards. Each card: {label, value, sub?, status?}."""
    inner = "".join(kpi_card(c.get("label"), c.get("value"), c.get("sub"),
                             c.get("status", "brand")) for c in cards)
    return f'<div class="kpi-row">{inner}</div>'


def status_pill(text, status="grey") -> str:
    """A small coloured RAG pill."""
    col = _status(status)
    return (f'<span class="pill" style="background:{col}1a;color:{col};'
            f'border:1px solid {col}55">{_e(text)}</span>')


def _norm_pairs(data) -> list[tuple]:
    """Accept [(label, value)] or [{'label':..,'value':..}] -> [(label, value)]."""
    out = []
    for d in data:
        if isinstance(d, dict):
            out.append((d.get("label"), d.get("value", 0)))
        else:
            out.append((d[0], d[1]))
    return out


def _nice_ticks(lo, hi, n=5):
    """A small set of 'nice' axis ticks spanning [lo, hi]. Floats the axis to the
    data range (not forced to 0), so a high-clustered trend line isn't squashed
    against the top. Returns (ticks, axis_lo, axis_hi)."""
    import math
    if hi <= lo:
        hi = lo + 1
    raw = (hi - lo) / n
    mag = 10 ** math.floor(math.log10(raw)) if raw > 0 else 1
    step = next(m * mag for m in (1, 2, 2.5, 5, 10) if raw <= m * mag)
    nlo = math.floor(lo / step) * step
    nhi = math.ceil(hi / step) * step
    ticks, v = [], nlo
    while v <= nhi + step * 1e-6:
        ticks.append(round(v, 6))
        v += step
    return ticks, nlo, nhi


_PARSE_NUMBER = None       # lazily resolved by _num(); False once known unreachable


def _load_parse_number():
    """The engine's `parse_number` if the shared `scripts/` is reachable, else None.

    Mirrors `_load_ingest`: viz.py stays stdlib-only and standalone, but when the
    toolkit engine *is* present the charts parse values exactly as the rest of the
    toolkit does (`'15%'` → 0.15, `'1.2m'` → 1200000, `'(500)'` → -500) instead of
    inventing a second, divergent numeric dialect."""
    import importlib
    from pathlib import Path as _P
    for p in (_P(__file__).resolve().parents[3] / "scripts",   # toolkit-root engine
              _P(__file__).resolve().parent / "scripts"):      # vendored sibling
        if (p / "dataclean.py").is_file():
            if str(p) not in sys.path:
                sys.path.insert(0, str(p))
            try:
                return importlib.import_module("dataclean").parse_number
            except (ImportError, AttributeError):
                return None
    return None


def _num(v):
    """Coerce a chart input to float, or None when it is not a number.

    Returning None rather than 0.0 is deliberate: a stray header, blank or free-text
    cell must be *skipped*, never plotted at the origin where it reads as a real
    observation of zero."""
    global _PARSE_NUMBER
    if isinstance(v, bool):                       # avoid True/False -> 1/0
        return None
    if isinstance(v, (int, float, Decimal)):
        f = float(v)
        return f if (f == f and f not in (float("inf"), float("-inf"))) else None
    if _PARSE_NUMBER is None:
        _PARSE_NUMBER = _load_parse_number() or False
    if _PARSE_NUMBER:
        val, _note = _PARSE_NUMBER(v)
        return float(val) if val is not None else None
    s = str(v).strip().replace(",", "")           # stdlib fallback: plain + accounting negatives
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try:
        return float(s)
    except ValueError:
        return None


def _skipped_note(n: int, what: str = "value") -> str:
    """Footnote for values a chart could not plot. Surfaced under the chart rather
    than dropped, so a half-empty plot is never mistaken for the whole story."""
    if not n:
        return ""
    return (f'<div class="note">{n} non-numeric {_e(what)}'
            f'{"s" if n != 1 else ""} skipped</div>')


def bar_chart(data, title=None, height=220, unit="") -> str:
    """A horizontal-axis bar chart as inline SVG. data: [(label, value)] or dicts.
    Geometry uses float coordinates; labels preserve Decimal formatting where supplied.
    Negative values are drawn below a zero line."""
    pairs = _norm_pairs(data)
    if not pairs:
        return _empty_block(title, "No data")
    vals = [float(v or 0) for _, v in pairs]
    vmin, vmax = min(vals + [0]), max(vals + [0])
    rng = (vmax - vmin) or 1
    n = len(pairs)
    W, pad_l, pad_b, pad_t = 640, 8, 28, 20  # pad_t leaves room for value labels
    plot_h = height - pad_b - pad_t
    gap = 10
    bw = (W - pad_l) / n - gap
    has_neg = vmin < 0
    y_zero = pad_t + plot_h - (0 - vmin) / rng * plot_h if has_neg else pad_t + plot_h
    bars = []
    for i, ((lab, raw_v), v) in enumerate(zip(pairs, vals)):
        if has_neg:
            h = abs(v) / rng * plot_h
            y = y_zero - h if v >= 0 else y_zero
        else:
            h = (v / rng) * plot_h
            y = pad_t + (plot_h - h)
        x = pad_l + i * (bw + gap)
        col = _SERIES[i % len(_SERIES)] if n <= len(_SERIES) else BRAND["burgundy"]
        lbl_y = y - 4 if v >= 0 else y + h + 12
        bars.append(
            f'<rect x="{x:.1f}" y="{y:.1f}" width="{bw:.1f}" height="{h:.1f}" '
            f'rx="3" fill="{col}"><title>{_e(lab)}: {_fmt_num(raw_v)}{_e(unit)}</title></rect>'
            f'<text x="{x + bw/2:.1f}" y="{lbl_y:.1f}" text-anchor="middle" '
            f'class="v">{_fmt_num(raw_v)}</text>'
            f'<text x="{x + bw/2:.1f}" y="{height - 9:.1f}" text-anchor="middle" '
            f'class="x">{_e(lab)}</text>')
    svg = (f'<svg viewBox="0 0 {W} {height}" class="chart" '
           f'preserveAspectRatio="xMidYMid meet">{" ".join(bars)}</svg>')
    return _chart_block(title, svg)


def line_chart(series, title=None, height=240, unit="", toggle=False) -> str:
    """A line chart. `series` is either [(label, value)] for one line, or
    {name: [(label, value)], ...} for several. X labels come from the first series.
    `toggle=True` makes the legend clickable to show/hide each series (inline JS)."""
    if isinstance(series, dict):
        named = list(series.items())
    else:
        named = [("", _norm_pairs(series))]
    named = [(nm, _norm_pairs(s) if not (s and isinstance(s[0], tuple)) else s)
             for nm, s in named]
    first = named[0][1] if named else []
    if not first:
        return _empty_block(title, "No data")
    labels = [l for l, _ in first]
    allv = [float(v or 0) for _, s in named for _, v in s]
    ticks, vmin, vmax = _nice_ticks(min(allv or [0]), max(allv or [0]))
    W, pad_l, pad_b, pad_t, pad_r = 640, 34, 26, 16, 10
    plot_h = height - pad_b - pad_t
    plot_w = W - pad_l - pad_r
    n = len(labels)
    xstep = plot_w / max(n - 1, 1)

    def _y(v):
        rng = (vmax - vmin) or 1
        return pad_t + plot_h - ((float(v or 0) - vmin) / rng) * plot_h

    # horizontal gridlines + y-axis value labels (so the floated scale is legible)
    gridlines = "".join(
        f'<line x1="{pad_l}" y1="{_y(t):.1f}" x2="{W - pad_r}" y2="{_y(t):.1f}" '
        f'stroke="{BRAND["grey_lt"]}" stroke-width="1"/>'
        f'<text x="{pad_l - 5}" y="{_y(t) + 3:.1f}" text-anchor="end" class="x">'
        f'{_fmt_num(_strip(t))}{_e(unit)}</text>' for t in ticks)

    groups, legend = [], []
    for si, (nm, s) in enumerate(named):
        col = _SERIES[si % len(_SERIES)]
        pts = [(pad_l + i * xstep, _y(v)) for i, (_, v) in enumerate(s)]
        d = "M" + " L".join(f"{x:.1f},{y:.1f}" for x, y in pts)
        path = f'<path d="{d}" fill="none" stroke="{col}" stroke-width="2.5"/>'
        dots = "".join(
            f'<circle cx="{x:.1f}" cy="{y:.1f}" r="3" fill="{col}">'
            f'<title>{_e(nm)} {_fmt_num(_strip(v))}{_e(unit)}</title></circle>'
            for (x, y), (_, v) in zip(pts, s))
        groups.append(f'<g data-series="{si}">{path}{dots}</g>')
        if nm:
            legend.append(f'<span class="lg"><i style="background:{col}"></i>{_e(nm)}</span>')
    xlabs = "".join(
        f'<text x="{pad_l + i * xstep:.1f}" y="{height - 8:.1f}" text-anchor="middle" '
        f'class="x">{_e(l)}</text>' for i, l in enumerate(labels))
    svg = (f'<svg viewBox="0 0 {W} {height}" class="chart" '
           f'preserveAspectRatio="xMidYMid meet">{gridlines}{"".join(groups)}{xlabs}</svg>')
    tg = ' data-toggle="1"' if (toggle and legend) else ''
    leg = f'<div class="legend"{tg}>{"".join(legend)}</div>' if legend else ""
    return _chart_block(title, svg + leg)


def donut_chart(data, title=None, height=240, centre=None) -> str:
    """A donut chart. data: [(label, value)] or dicts. `centre` overrides the
    centre label (defaults to the total)."""
    pairs = _norm_pairs(data)
    vals = [max(float(v or 0), 0) for _, v in pairs]
    total = sum(vals)
    if total <= 0:
        return _empty_block(title, "No data")
    cx, cy, r, rin = 120, 120, 100, 62
    import math
    a0 = -math.pi / 2
    arcs, legend = [], []
    for i, ((lab, _), v) in enumerate(zip(pairs, vals)):
        frac = v / total
        a1 = a0 + frac * 2 * math.pi
        large = 1 if frac > 0.5 else 0
        x0, y0 = cx + r * math.cos(a0), cy + r * math.sin(a0)
        x1, y1 = cx + r * math.cos(a1), cy + r * math.sin(a1)
        col = _SERIES[i % len(_SERIES)]
        arcs.append(f'<path d="M{x0:.1f},{y0:.1f} A{r},{r} 0 {large} 1 {x1:.1f},{y1:.1f} '
                    f'L{cx},{cy} Z" fill="{col}"><title>{_e(lab)}: {_fmt_num(_strip(v))} '
                    f'({frac*100:.0f}%)</title></path>')
        legend.append(f'<span class="lg"><i style="background:{col}"></i>'
                      f'{_e(lab)} <b>{_fmt_num(_strip(v))}</b></span>')
        a0 = a1
    hole = f'<circle cx="{cx}" cy="{cy}" r="{rin}" fill="{BRAND["white"]}"/>'
    centre_t = _e(centre) if centre is not None else _fmt_num(_strip(total))
    ctext = (f'<text x="{cx}" y="{cy-2}" text-anchor="middle" class="dn-c">{centre_t}</text>'
             f'<text x="{cx}" y="{cy+16}" text-anchor="middle" class="dn-l">total</text>')
    svg = (f'<svg viewBox="0 0 240 {max(height,240)}" class="chart donut" '
           f'preserveAspectRatio="xMidYMid meet">{"".join(arcs)}{hole}{ctext}</svg>')
    return _chart_block(title, f'<div class="donut-wrap">{svg}'
                               f'<div class="legend col">{"".join(legend)}</div></div>')


def _hex_to_rgb(h: str) -> tuple[int, int, int]:
    h = (h or "#000000").lstrip("#")
    if len(h) == 3:
        h = "".join(c * 2 for c in h)
    return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)


def _rgb_to_hex(r: float, g: float, b: float) -> str:
    return f"#{max(0, min(255, int(round(r)))):02x}{max(0, min(255, int(round(g)))):02x}{max(0, min(255, int(round(b)))):02x}"


def _colour_at(t: float, scale: str = "sequential") -> str:
    """Map t∈[0,1] onto the active theme. sequential = tint→primary; diverging =
    red→paper→green (association / correlation)."""
    t = 0.0 if t is None else max(0.0, min(1.0, float(t)))
    if scale == "diverging":
        lo, mid, hi = (_hex_to_rgb(BRAND["red"]), _hex_to_rgb(BRAND["white"]),
                       _hex_to_rgb(BRAND["green"]))
        if t < 0.5:
            u = t * 2
            rgb = tuple(lo[i] + (mid[i] - lo[i]) * u for i in range(3))
        else:
            u = (t - 0.5) * 2
            rgb = tuple(mid[i] + (hi[i] - mid[i]) * u for i in range(3))
    else:
        lo, hi = _hex_to_rgb(BRAND["pink_vlt"]), _hex_to_rgb(BRAND["burgundy"])
        rgb = tuple(lo[i] + (hi[i] - lo[i]) * t for i in range(3))
    return _rgb_to_hex(*rgb)


def _contrast_ink(fill: str) -> str:
    r, g, b = _hex_to_rgb(fill)
    # Relative luminance shortcut — dark cells get white labels.
    return BRAND["white"] if (0.299 * r + 0.587 * g + 0.114 * b) < 140 else BRAND["ink"]


def heatmap(matrix, row_labels=None, col_labels=None, title=None,
            scale="sequential", mid=0, unit="", cell_w=48, cell_h=32) -> str:
    """Matrix as an inline-SVG heat map. `matrix` is a list of rows (each a list of
    numbers); missing/None cells render empty. `scale`: ``sequential`` (magnitude)
    or ``diverging`` (centred on `mid`, e.g. 0 for correlations). Built for pivot
    tables, cohort retention and correlation matrices — not a generic Excel dump."""
    if not matrix:
        return _empty_block(title, "No data")
    rows = [list(r) for r in matrix]
    n_r, n_c = len(rows), max(len(r) for r in rows)
    row_labels = list(row_labels) if row_labels is not None else [str(i + 1) for i in range(n_r)]
    col_labels = list(col_labels) if col_labels is not None else [str(i + 1) for i in range(n_c)]
    nums = []
    for r in rows:
        for v in r:
            if v is None or v == "":
                continue
            try:
                nums.append(float(v))
            except (TypeError, ValueError):
                continue
    if not nums:
        return _empty_block(title, "No numeric cells")
    if scale == "diverging":
        span = max(abs(v - float(mid)) for v in nums) or 1.0
    else:
        lo, hi = min(nums), max(nums)
        span = (hi - lo) or 1.0
    pad_l, pad_t = 90, 28
    W = pad_l + n_c * cell_w + 8
    H = pad_t + n_r * cell_h + 8
    cells, xlabs, ylabs = [], [], []
    for j, lab in enumerate(col_labels[:n_c]):
        xlabs.append(
            f'<text x="{pad_l + j * cell_w + cell_w/2:.1f}" y="{pad_t - 8}" '
            f'text-anchor="middle" class="x">{_e(lab)}</text>')
    for i, lab in enumerate(row_labels[:n_r]):
        ylabs.append(
            f'<text x="{pad_l - 6}" y="{pad_t + i * cell_h + cell_h/2 + 4:.1f}" '
            f'text-anchor="end" class="x">{_e(lab)}</text>')
        row = rows[i] + [None] * (n_c - len(rows[i]))
        for j, raw in enumerate(row[:n_c]):
            x, y = pad_l + j * cell_w, pad_t + i * cell_h
            if raw is None or raw == "":
                cells.append(
                    f'<rect x="{x}" y="{y}" width="{cell_w - 2}" height="{cell_h - 2}" '
                    f'rx="3" fill="{BRAND["pink_vlt"]}" opacity=".35"/>')
                continue
            try:
                v = float(raw)
            except (TypeError, ValueError):
                cells.append(
                    f'<rect x="{x}" y="{y}" width="{cell_w - 2}" height="{cell_h - 2}" '
                    f'rx="3" fill="{BRAND["pink_vlt"]}" opacity=".35"/>')
                continue
            if scale == "diverging":
                t = 0.5 + (v - float(mid)) / (2 * span)
            else:
                t = (v - lo) / span
            fill = _colour_at(t, scale=scale)
            ink = _contrast_ink(fill)
            cells.append(
                f'<rect x="{x}" y="{y}" width="{cell_w - 2}" height="{cell_h - 2}" '
                f'rx="3" fill="{fill}"><title>{_e(row_labels[i])} × {_e(col_labels[j])}: '
                f'{_fmt_num(_strip(v))}{_e(unit)}</title></rect>'
                f'<text x="{x + (cell_w - 2)/2:.1f}" y="{y + cell_h/2 + 3:.1f}" '
                f'text-anchor="middle" class="hm-v" fill="{ink}">{_fmt_num(_strip(v))}</text>')
    svg = (f'<svg viewBox="0 0 {W} {H}" class="chart heatmap" '
           f'preserveAspectRatio="xMidYMid meet">{"".join(xlabs)}{"".join(ylabs)}'
           f'{"".join(cells)}</svg>')
    return _chart_block(title, svg)


def sparkline(data, title=None, width=180, height=42, show_last=True, unit="") -> str:
    """Compact inline trend — a single path sized for a KPI strip or tight grid.
    Prefer this when the shape matters more than axis ticks; use `line_chart` for
    a readable scale. data: [(label, value)] or bare values."""
    if data and not isinstance(data[0], (list, tuple, dict)):
        pairs = [(str(i + 1), v) for i, v in enumerate(data)]
    else:
        pairs = _norm_pairs(data)
    if len(pairs) < 2:
        return _empty_block(title, "Need ≥2 points")
    vals = [float(v or 0) for _, v in pairs]
    lo, hi = min(vals), max(vals)
    rng = (hi - lo) or 1.0
    pad_x, pad_y = 4, 6
    plot_w = width - 2 * pad_x
    plot_h = height - 2 * pad_y
    pts = []
    for i, v in enumerate(vals):
        x = pad_x + (i / (len(vals) - 1)) * plot_w
        y = pad_y + plot_h - ((v - lo) / rng) * plot_h
        pts.append((x, y))
    d = "M" + " L".join(f"{x:.1f},{y:.1f}" for x, y in pts)
    area = d + f" L{pts[-1][0]:.1f},{height - pad_y:.1f} L{pts[0][0]:.1f},{height - pad_y:.1f} Z"
    col = BRAND["burgundy"]
    last = pairs[-1]
    delta = vals[-1] - vals[0]
    tip = f'{_e(last[0])}: {_fmt_num(last[1])}{_e(unit)}'
    svg = (
        f'<svg viewBox="0 0 {width} {height}" class="chart spark" '
        f'preserveAspectRatio="none" width="{width}" height="{height}">'
        f'<path d="{area}" fill="{col}" opacity=".12"/>'
        f'<path d="{d}" fill="none" stroke="{col}" stroke-width="2"/>'
        f'<circle cx="{pts[-1][0]:.1f}" cy="{pts[-1][1]:.1f}" r="2.5" fill="{col}">'
        f'<title>{tip}</title></circle></svg>'
    )
    meta = ""
    if show_last:
        sign = "+" if delta > 0 else ""
        meta = (f'<div class="spark-meta"><span class="spark-last">{_fmt_num(last[1])}'
                f'{_e(unit)}</span><span class="spark-delta">'
                f'{sign}{_fmt_num(_strip(delta))} overall</span></div>')
    return _chart_block(title, f'<div class="spark-wrap">{svg}{meta}</div>')


def waterfall(steps, title=None, height=240, unit="") -> str:
    """Bridge chart: how a total moves through signed steps. Each step is
    ``(label, value, kind)`` or ``{"label","value","kind"}`` where kind is
    ``start`` | ``delta`` | ``total``. If kind is omitted: first→start, last→total,
    middle→delta. Deltas float from the running total; start/total rise from zero.
    Descriptive only — never a forecast."""
    norm = []
    for i, s in enumerate(steps or []):
        if isinstance(s, dict):
            lab, val, kind = s.get("label"), s.get("value", 0), s.get("kind")
        else:
            lab = s[0]
            val = s[1] if len(s) > 1 else 0
            kind = s[2] if len(s) > 2 else None
        if kind is None:
            kind = "start" if i == 0 else ("total" if i == len(steps) - 1 else "delta")
        if kind not in ("start", "delta", "total"):
            kind = "delta"
        norm.append((lab, val, kind))
    if not norm:
        return _empty_block(title, "No data")
    running = 0.0
    geometry = []  # (label, raw, y0, y1, kind)
    for lab, raw, kind in norm:
        v = float(raw or 0)
        if kind == "start":
            y0, y1, running = 0.0, v, v
        elif kind == "total":
            y0, y1 = 0.0, v
            running = v
        else:
            y0, y1 = running, running + v
            running = y1
        geometry.append((lab, raw, y0, y1, kind))
    ys = [y for _, _, a, b, _ in geometry for y in (a, b)]
    vmin, vmax = min(ys + [0.0]), max(ys + [0.0])
    ticks, axis_lo, axis_hi = _nice_ticks(vmin, vmax)
    rng = (axis_hi - axis_lo) or 1.0
    W, pad_l, pad_b, pad_t, pad_r = 640, 40, 30, 18, 10
    plot_h = height - pad_b - pad_t
    n = len(geometry)
    gap = 12
    bw = (W - pad_l - pad_r) / n - gap

    def _y(v):
        return pad_t + plot_h - ((float(v) - axis_lo) / rng) * plot_h

    grid = "".join(
        f'<line x1="{pad_l}" y1="{_y(t):.1f}" x2="{W - pad_r}" y2="{_y(t):.1f}" '
        f'stroke="{BRAND["grey_lt"]}" stroke-width="1"/>'
        f'<text x="{pad_l - 5}" y="{_y(t) + 3:.1f}" text-anchor="end" class="x">'
        f'{_fmt_num(_strip(t))}{_e(unit)}</text>' for t in ticks)
    shapes, connectors = [], []
    prev_top = None
    for i, (lab, raw, y0, y1, kind) in enumerate(geometry):
        x = pad_l + i * (bw + gap)
        top, bot = max(y0, y1), min(y0, y1)
        y = _y(top)
        h = abs(_y(y0) - _y(y1))
        if kind == "delta":
            col = BRAND["green"] if float(raw or 0) >= 0 else BRAND["red"]
        elif kind == "total":
            col = BRAND["burgundy"]
        else:
            col = BRAND["rose"]
        shapes.append(
            f'<rect x="{x:.1f}" y="{y:.1f}" width="{bw:.1f}" height="{max(h, 1):.1f}" '
            f'rx="3" fill="{col}"><title>{_e(lab)}: {_fmt_num(raw)}{_e(unit)}</title></rect>'
            f'<text x="{x + bw/2:.1f}" y="{y - 4:.1f}" text-anchor="middle" class="v">'
            f'{_fmt_num(raw)}</text>'
            f'<text x="{x + bw/2:.1f}" y="{height - 9:.1f}" text-anchor="middle" class="x">'
            f'{_e(lab)}</text>')
        if prev_top is not None and kind == "delta":
            x0 = prev_top[0]
            connectors.append(
                f'<line x1="{x0:.1f}" y1="{_y(prev_top[1]):.1f}" x2="{x:.1f}" '
                f'y2="{_y(y0):.1f}" stroke="{BRAND["grey"]}" stroke-width="1" '
                f'stroke-dasharray="3 3"/>')
        prev_top = (x + bw, y1)
    svg = (f'<svg viewBox="0 0 {W} {height}" class="chart" '
           f'preserveAspectRatio="xMidYMid meet">{grid}{"".join(connectors)}'
           f'{"".join(shapes)}</svg>')
    return _chart_block(title, svg)


def scatter_chart(x, y, title=None, height=260, x_label=None, y_label=None,
                  unit_x="", unit_y="", labels=None, trend_line=False) -> str:
    """Scatter plot of paired observations — for correlation and outlier spotting.

    ``x`` and ``y`` are parallel sequences of numbers, or ``[(label, value)]`` pairs
    (the labels are then used for tooltips). On a length mismatch the longer side is
    **trimmed to the shorter** and the unpaired tail is counted and reported — never
    silently dropped. A pair is plotted only when both sides parse as numbers.

    ``trend_line=True`` overlays an ordinary least-squares fit across the plot area.
    It is **descriptive, never a forecast**; the tooltip carries slope, intercept and
    Pearson r so the reader can judge how much the line is worth. A cloud with no
    x-variance has no defensible slope, so no line is drawn at all.

    ``x_label`` / ``y_label`` annotate the axes (y rotated -90°). Both axes float to
    the data via ``_nice_ticks`` (as ``line_chart`` does): a tightly-clustered scatter
    forced against a zero baseline shows nothing.
    """
    def _split(seq):
        """Accept raw numbers or (label, value) pairs; return (values, labels)."""
        seq = list(seq or [])
        if seq and isinstance(seq[0], (tuple, list, dict)):
            pairs = _norm_pairs(seq)
            return [v for _l, v in pairs], [l for l, _v in pairs]
        return seq, []

    xs_raw, xlabs = _split(x)
    ys_raw, ylabs = _split(y)
    labs = list(labels or []) or xlabs or ylabs
    pts, skipped = [], 0
    for i in range(min(len(xs_raw), len(ys_raw))):
        fx, fy = _num(xs_raw[i]), _num(ys_raw[i])
        if fx is None or fy is None:
            skipped += 1
            continue
        pts.append((fx, fy, labs[i] if i < len(labs) else None,
                    xs_raw[i], ys_raw[i]))
    # a length mismatch is a data error, not a plot: trim to the shorter side and
    # count the unpaired tail so the discrepancy is visible under the chart
    skipped += abs(len(xs_raw) - len(ys_raw))
    if not pts:
        return _empty_block(title, "No data")

    xticks, x_lo, x_hi = _nice_ticks(min(p[0] for p in pts), max(p[0] for p in pts))
    yticks, y_lo, y_hi = _nice_ticks(min(p[1] for p in pts), max(p[1] for p in pts))
    W, pad_t, pad_r = 640, 16, 16
    pad_l = 44 if not y_label else 58        # room for the rotated axis label
    pad_b = 36 if not x_label else 48
    plot_h = height - pad_b - pad_t
    plot_w = W - pad_l - pad_r
    xrng = (x_hi - x_lo) or 1.0
    yrng = (y_hi - y_lo) or 1.0

    def _px(v):
        return pad_l + ((float(v) - x_lo) / xrng) * plot_w

    def _py(v):
        return pad_t + plot_h - ((float(v) - y_lo) / yrng) * plot_h

    grid = "".join(
        f'<line x1="{pad_l}" y1="{_py(t):.1f}" x2="{W - pad_r}" y2="{_py(t):.1f}" '
        f'stroke="{BRAND["grey_lt"]}" stroke-width="1"/>'
        f'<text x="{pad_l - 5}" y="{_py(t) + 3:.1f}" text-anchor="end" class="x">'
        f'{_fmt_num(_strip(t))}{_e(unit_y)}</text>' for t in yticks)
    grid += "".join(
        f'<line x1="{_px(t):.1f}" y1="{pad_t}" x2="{_px(t):.1f}" y2="{pad_t + plot_h:.1f}" '
        f'stroke="{BRAND["grey_lt"]}" stroke-width="1"/>'
        f'<text x="{_px(t):.1f}" y="{pad_t + plot_h + 16:.1f}" text-anchor="middle" '
        f'class="x">{_fmt_num(_strip(t))}{_e(unit_x)}</text>' for t in xticks)

    axis_labels = ""
    if x_label:
        axis_labels += (f'<text x="{pad_l + plot_w / 2:.1f}" y="{height - 2}" '
                        f'text-anchor="middle" class="x">{_e(x_label)}</text>')
    if y_label:
        ly = pad_t + plot_h / 2
        axis_labels += (f'<text x="12" y="{ly:.1f}" text-anchor="middle" class="x" '
                        f'transform="rotate(-90 12 {ly:.1f})">{_e(y_label)}</text>')

    col = BRAND["burgundy"]
    dots = "".join(
        f'<circle cx="{_px(fx):.1f}" cy="{_py(fy):.1f}" r="4" fill="{col}" '
        f'fill-opacity="0.72"><title>{_e(lab) + ": " if lab else ""}'
        f'{_fmt_num(rx)}{_e(unit_x)}, {_fmt_num(ry)}{_e(unit_y)}</title></circle>'
        for fx, fy, lab, rx, ry in pts)

    fit = ""
    if trend_line and len(pts) >= 2:
        n = len(pts)
        mx = sum(p[0] for p in pts) / n
        my = sum(p[1] for p in pts) / n
        sxx = sum((p[0] - mx) ** 2 for p in pts)
        sxy = sum((p[0] - mx) * (p[1] - my) for p in pts)
        syy = sum((p[1] - my) ** 2 for p in pts)
        if sxx > 0:                       # vertical cloud has no OLS slope — draw nothing
            slope = sxy / sxx
            intercept = my - slope * mx
            r = sxy / ((sxx * syy) ** 0.5) if syy > 0 else 0.0
            # span the plot area, per spec. The axes are floated to the data, so the
            # edges sit only a tick outside the observed range — not a projection.
            fit = (f'<line x1="{_px(x_lo):.1f}" y1="{_py(slope * x_lo + intercept):.1f}" '
                   f'x2="{_px(x_hi):.1f}" y2="{_py(slope * x_hi + intercept):.1f}" '
                   f'stroke="{BRAND["grey"]}" stroke-width="1.5" stroke-dasharray="4,3">'
                   f'<title>trend: y = {slope:,.4g}x + {intercept:,.4g} '
                   f'(r {r:,.3f}, n {n}) — descriptive, not a forecast</title></line>')
    svg = (f'<svg viewBox="0 0 {W} {height}" class="chart" '
           f'preserveAspectRatio="xMidYMid meet">{grid}{fit}{dots}{axis_labels}</svg>')
    return _chart_block(title, svg + _skipped_note(skipped, "pair"))


def histogram(values, bins=10, title=None, height=240, unit="") -> str:
    """Distribution shape as contiguous frequency bars.

    ``bins`` is either a bin *count* (equal-width across the observed range) or an
    explicit list of *edges* — e.g. ``[0, 30, 60, 90, 365]`` for ageing-style
    buckets. Edges are half-open ``[lo, hi)`` except the last, which includes its
    upper bound so the maximum observation is never dropped.

    Non-numeric values are skipped and reported under the chart rather than silently
    discarded. Unlike the other charts the y-axis is **forced to zero**: a frequency
    axis that floats misstates how tall a bar is relative to nothing, which is the
    entire point of a histogram. Bars touch, signalling a continuous scale.
    """
    nums, skipped = [], 0
    for v in (values or []):
        f = _num(v)
        if f is None:
            if str(v).strip() != "":       # blanks are absence, not bad data
                skipped += 1
            continue
        nums.append(f)
    if len(nums) < 2:
        # a single observation has no distribution to show; say so rather than
        # rendering one lone bar that implies a shape
        return _empty_block(title, "No data" if not nums else "Need at least 2 values")

    if isinstance(bins, (list, tuple)):
        edges = sorted(float(e) for e in bins if _num(e) is not None)
        if len(edges) < 2:
            return _empty_block(title, "Need at least two bin edges")
    else:
        k = max(int(bins or 10), 1)
        lo, hi = min(nums), max(nums)
        if hi <= lo:                       # every value identical — one honest bin
            lo, hi = lo - 0.5, hi + 0.5
        step = (hi - lo) / k
        edges = [lo + i * step for i in range(k + 1)]

    counts = [0] * (len(edges) - 1)
    outside = 0
    for f in nums:
        if f < edges[0] or f > edges[-1]:
            outside += 1                   # numeric, but outside the caller's edges
            continue
        for i in range(len(counts)):
            hi_edge = edges[i + 1]
            if f < hi_edge or (i == len(counts) - 1 and f <= hi_edge):
                counts[i] += 1
                break

    ticks, _lo, axis_hi = _nice_ticks(0, max(counts) or 1)
    axis_hi = axis_hi or 1
    W, pad_l, pad_b, pad_t, pad_r = 640, 40, 34, 18, 12
    plot_h = height - pad_b - pad_t
    plot_w = W - pad_l - pad_r
    bw = plot_w / len(counts)

    def _y(c):
        return pad_t + plot_h - (float(c) / axis_hi) * plot_h

    grid = "".join(
        f'<line x1="{pad_l}" y1="{_y(t):.1f}" x2="{W - pad_r}" y2="{_y(t):.1f}" '
        f'stroke="{BRAND["grey_lt"]}" stroke-width="1"/>'
        f'<text x="{pad_l - 5}" y="{_y(t) + 3:.1f}" text-anchor="end" class="x">'
        f'{_fmt_num(_strip(t))}</text>' for t in ticks)
    bars = []
    for i, c in enumerate(counts):
        x = pad_l + i * bw
        y = _y(c)
        h = max(pad_t + plot_h - y, 0)
        bars.append(
            f'<rect x="{x:.1f}" y="{y:.1f}" width="{bw:.1f}" height="{h:.1f}" rx="1" '
            f'fill="{BRAND["burgundy"]}" stroke="{BRAND["white"]}" stroke-width="0.5">'
            f'<title>{c} value{"s" if c != 1 else ""} in '
            f'[{_fmt_num(_strip(round(edges[i], 6)))}{_e(unit)}, '
            f'{_fmt_num(_strip(round(edges[i + 1], 6)))}{_e(unit)}'
            f'{"]" if i == len(counts) - 1 else ")"}</title></rect>')
    # label the edges, thinned so they stay legible on a many-bin histogram
    every = max(1, (len(edges) + 7) // 8)
    xlabs = "".join(
        f'<text x="{pad_l + i * bw:.1f}" y="{height - 10:.1f}" text-anchor="middle" '
        f'class="x">{_fmt_num(_strip(round(e, 6)))}{_e(unit)}</text>'
        for i, e in enumerate(edges) if i % every == 0 or i == len(edges) - 1)
    svg = (f'<svg viewBox="0 0 {W} {height}" class="chart" '
           f'preserveAspectRatio="xMidYMid meet">{grid}{"".join(bars)}{xlabs}</svg>')
    # two distinct exclusions, reported distinctly: unparseable data vs values the
    # caller's own bin edges do not cover (the latter is a spec gap, not dirty data)
    note = _skipped_note(skipped)
    if outside:
        note += (f'<div class="note">{outside} value{"s" if outside != 1 else ""} '
                 f'outside the bin range '
                 f'({_fmt_num(_strip(round(edges[0], 6)))}'
                 f'–{_fmt_num(_strip(round(edges[-1], 6)))}) excluded</div>')
    return _chart_block(title, svg + note)


def _is_pairish(seq) -> bool:
    """True when a sequence holds (label, value) pairs rather than bare numbers.

    This is what separates ``{segment: [(cat, value)]}`` from the brief's
    ``{category: [v1, v2, v3]}`` — both are dicts of lists, and only the element
    shape tells them apart."""
    for item in (seq or []):
        return isinstance(item, (tuple, list, dict))
    return False


def _norm_stacked(data) -> tuple[list, list, list]:
    """Normalise stacked_bar input to ``(categories, segments, matrix[cat][segment])``.

    Five accepted shapes, because callers arrive from different directions:

    - a ``pivot()`` result — ``row_keys`` become the bars, ``col_keys`` the segments
    - ``{category: [v1, v2, …]}`` — one bar per key, positional segments
    - ``[(category, [v1, v2, …])]`` — the same, order-preserving
    - ``{"categories": [...], "series": {name: [values]}}`` — named segments
    - ``{segment: [(category, value)]}`` — series-first, labels carried per point

    Positional shapes have no segment names of their own, so segments are numbered
    (``Segment 1…n``) rather than left blank — an unlabelled legend swatch is worse
    than an honest placeholder.

    **Missing / unparseable cells become 0.0**, so a sparse pivot (where some
    category×segment combinations have no rows) renders as a zero-height segment
    rather than a gap. This is correct for composition charts — an absent value
    contributes nothing to the stack — but a caller who needs to distinguish
    "zero" from "no data" should filter those cells beforehand or use the
    ``heatmap`` block, which renders missing cells as a visibly empty outline.
    """
    if isinstance(data, dict) and "row_keys" in data and "matrix" in data:
        cats = [_strip(k) for k in data.get("row_keys", [])]
        names = [_strip(k) for k in data.get("col_keys", [])]
        matrix = [[_num(v) or 0.0 for v in row] for row in data.get("matrix", [])]
        return cats, names, matrix
    if isinstance(data, dict) and "series" in data:
        cats = list(data.get("categories", []))
        series = data.get("series") or {}
        names = list(series.keys())
        matrix = [[_num(series[nm][i]) or 0.0 if i < len(series[nm]) else 0.0
                   for nm in names] for i in range(len(cats))]
        return cats, names, matrix

    # list of (category, [values]) tuples
    if isinstance(data, (list, tuple)) and data:
        cats = [_strip(c) for c, _vals in data]
        width = max(len(v or []) for _c, v in data)
        matrix = [[_num((v or [])[i]) or 0.0 if i < len(v or []) else 0.0
                   for i in range(width)] for _c, v in data]
        return cats, [f"Segment {i + 1}" for i in range(width)], matrix

    if isinstance(data, dict) and data:
        first = next(iter(data.values()))
        if not _is_pairish(first):
            # {category: [v1, v2, ...]} — positional segments
            cats = [_strip(k) for k in data]
            width = max(len(v or []) for v in data.values())
            matrix = [[_num((v or [])[i]) or 0.0 if i < len(v or []) else 0.0
                       for i in range(width)] for v in data.values()]
            return cats, [f"Segment {i + 1}" for i in range(width)], matrix
        # {segment: [(category, value)]} — segment-first, categories from the labels
        names = list(data.keys())
        cats, seen = [], set()
        for nm in names:
            for lab, _v in _norm_pairs(data[nm]):
                if lab not in seen:
                    seen.add(lab)
                    cats.append(lab)
        lookup = {nm: dict(_norm_pairs(data[nm])) for nm in names}
        matrix = [[_num(lookup[nm].get(c)) or 0.0 for nm in names] for c in cats]
        return cats, names, matrix
    return [], [], []


def stacked_bar(data, title=None, height=260, unit="", legend=True) -> str:
    """Composition over time: one bar per category, split into stacked segments.

    ``data`` accepts a ``pivot()`` result straight from ``analyse.py`` (``row_keys``
    become the bars, ``col_keys`` the segments), ``{segment: [(category, value)]}``,
    or ``{"categories": [...], "series": {name: [values]}}``.

    **Negative segments stack downward from the zero line** rather than being folded
    into the positive stack — a contra, credit note or reversal stays visible as a
    reduction instead of silently inflating the bar it belongs to.
    """
    cats, names, matrix = _norm_stacked(data)
    if not cats or not names:
        return _empty_block(title, "No data")

    tops = [sum(v for v in row if v > 0) for row in matrix]
    bots = [sum(v for v in row if v < 0) for row in matrix]
    ticks, axis_lo, axis_hi = _nice_ticks(min(bots + [0.0]), max(tops + [0.0]))
    rng = (axis_hi - axis_lo) or 1.0
    W, pad_l, pad_b, pad_t, pad_r = 640, 40, 30, 18, 12
    plot_h = height - pad_b - pad_t
    plot_w = W - pad_l - pad_r
    gap = 6
    bw = plot_w / len(cats) - gap

    def _y(v):
        return pad_t + plot_h - ((float(v) - axis_lo) / rng) * plot_h

    grid = "".join(
        f'<line x1="{pad_l}" y1="{_y(t):.1f}" x2="{W - pad_r}" y2="{_y(t):.1f}" '
        f'stroke="{BRAND["grey_lt"]}" stroke-width="1"/>'
        f'<text x="{pad_l - 5}" y="{_y(t) + 3:.1f}" text-anchor="end" class="x">'
        f'{_fmt_num(_strip(t))}{_e(unit)}</text>' for t in ticks)
    if any(v < 0 for row in matrix for v in row):
        grid += (f'<line x1="{pad_l}" y1="{_y(0):.1f}" x2="{W - pad_r}" y2="{_y(0):.1f}" '
                 f'stroke="{BRAND["grey"]}" stroke-width="1.5"/>')

    shapes = []
    for ci, cat in enumerate(cats):
        x = pad_l + ci * (bw + gap) + gap / 2
        up = 0.0     # running top of the positive stack
        down = 0.0   # running bottom of the negative stack
        for si, nm in enumerate(names):
            v = matrix[ci][si]
            if not v:
                continue
            col = _SERIES[si % len(_SERIES)]
            if v > 0:
                y0, y1 = up, up + v
                up = y1
            else:
                y0, y1 = down + v, down
                down = y0
            y = _y(y1)
            h = abs(_y(y0) - _y(y1))
            shapes.append(
                f'<rect x="{x:.1f}" y="{y:.1f}" width="{bw:.1f}" height="{max(h, 1):.1f}" '
                f'fill="{col}"><title>{_e(cat)} · {_e(nm)}: {_fmt_num(_strip(v))}'
                f'{_e(unit)}</title></rect>')
        shapes.append(
            f'<text x="{x + bw/2:.1f}" y="{height - 9:.1f}" text-anchor="middle" '
            f'class="x">{_e(cat)}</text>')
    svg = (f'<svg viewBox="0 0 {W} {height}" class="chart" '
           f'preserveAspectRatio="xMidYMid meet">{grid}{"".join(shapes)}</svg>')
    leg = ""
    if legend:
        leg = ('<div class="legend">' + "".join(
            f'<span class="lg"><i style="background:{_SERIES[si % len(_SERIES)]}"></i>'
            f'{_e(nm)}</span>' for si, nm in enumerate(names)) + '</div>')
    return _chart_block(title, svg + leg)


def table(rows: list[dict], columns: list[str] | None = None, title=None,
          rag: dict | None = None, sortable=False, filter_by=None) -> str:
    """An on-brand data table. `rows` = list of dicts. `columns` selects/orders
    columns (default: keys of the first row). `rag` maps a column name to a
    function value->status ('green'/'amber'/'red'/...) for a coloured cell.
    `sortable=True` makes column headers click-to-sort; `filter_by=<column>`
    adds a dropdown that filters rows by that column's value. Both are inline JS
    and degrade cleanly for print (sorted order prints; filtered-out rows don't)."""
    if not rows:
        return _empty_block(title, "No rows")
    cols = columns or list(rows[0].keys())
    rag = rag or {}
    head = "".join(f"<th>{_e(c)}</th>" for c in cols)
    body = []
    for r in rows:
        tds = []
        for c in cols:
            val = r.get(c)
            if c in rag:
                try:
                    st = rag[c](val)
                except Exception:
                    st = None
                if st:
                    col = _status(st)
                    tds.append(f'<td style="background:{col}14;'
                               f'border-left:3px solid {col}">{_fmt_num(val)}</td>')
                    continue
            tds.append(f"<td>{_fmt_num(val)}</td>")
        body.append(f"<tr>{''.join(tds)}</tr>")
    tid = _next_id() if filter_by in cols else ""
    sort_attr = ' data-sortable="1"' if sortable else ''
    id_attr = f' id="{tid}"' if tid else ''
    tbl = (f'<table{id_attr} class="grid"{sort_attr}><thead><tr>{head}</tr></thead>'
           f'<tbody>{"".join(body)}</tbody></table>')
    if tid:                                    # filter dropdown bound to this table
        ci = cols.index(filter_by)
        opts = "".join(f'<option value="{_e(o)}">{_e(o)}</option>'
                       for o in sorted({str(r.get(filter_by, "")) for r in rows}))
        bar = (f'<div class="filterbar">{_e(filter_by)}: '
               f'<select data-filter="1" data-col="{ci}" data-target="{tid}">'
               f'<option value="__all">All</option>{opts}</select></div>')
        tbl = bar + tbl
    return _chart_block(title, tbl)


def section(title, *blocks) -> str:
    """A titled section wrapping one or more blocks."""
    return (f'<section class="sec"><h2>{_e(title)}</h2>'
            f'{"".join(blocks)}</section>')


def grid(*blocks, cols=2) -> str:
    """Lay blocks out in an N-column responsive grid."""
    return (f'<div class="lay" style="grid-template-columns:repeat({cols},1fr)">'
            f'{"".join(blocks)}</div>')


# --------------------------------------------------------------------------- #
# Internal layout helpers
# --------------------------------------------------------------------------- #
def _strip(v):
    """Drop a trailing .0 from whole floats so labels read cleanly."""
    return int(v) if isinstance(v, float) and v.is_integer() else v


def _chart_block(title, inner) -> str:
    t = f'<h3 class="bt">{_e(title)}</h3>' if title else ""
    return f'<div class="block">{t}{inner}</div>'


def _empty_block(title, msg) -> str:
    return _chart_block(title, f'<div class="empty">{_e(msg)}</div>')


# --------------------------------------------------------------------------- #
# Internationalisation — conditional font fallback (no fonts shipped)
# --------------------------------------------------------------------------- #
# CJK Unified Ideographs + Hiragana/Katakana + Hangul (Project 3b acceptance).
_CJK_RANGES = (
    (0x4E00, 0x9FFF),   # CJK Unified Ideographs
    (0x3040, 0x30FF),   # Hiragana, Katakana
    (0xAC00, 0xD7AF),   # Hangul Syllables
)

# Broader script coverage for SG/Asia + other markets — still browser-fallback only.
_SCRIPT_RANGES = {
    "cjk": _CJK_RANGES,
    "arabic": (
        (0x0600, 0x06FF), (0x0750, 0x077F), (0x08A0, 0x08FF),
        (0xFB50, 0xFDFF), (0xFE70, 0xFEFF),
    ),
    "hebrew": ((0x0590, 0x05FF),),
    "cyrillic": ((0x0400, 0x04FF), (0x0500, 0x052F),),
    "thai": ((0x0E00, 0x0E7F),),
    "devanagari": ((0x0900, 0x097F),),
    "tamil": ((0x0B80, 0x0BFF),),
    "bengali": ((0x0980, 0x09FF),),
    "gujarati": ((0x0A80, 0x0AFF),),
    "gurmukhi": ((0x0A00, 0x0A7F),),
    "kannada": ((0x0C80, 0x0CFF),),
    "malayalam": ((0x0D00, 0x0D7F),),
    "telugu": ((0x0C00, 0x0C7F),),
    "georgian": ((0x10A0, 0x10FF),),
    "armenian": ((0x0530, 0x058F),),
    "ethiopic": ((0x1200, 0x137F),),
}

# Named fonts commonly present on macOS / Windows / desktop Linux. Never install or
# search the filesystem — the browser picks the first available name.
_CJK_FONT_NAMES = (
    "Noto Sans CJK SC", "Noto Sans CJK JP", "Noto Sans CJK KR",
    "Microsoft YaHei", "PingFang SC", "Hiragino Sans GB",
    "Yu Gothic", "Malgun Gothic", "Apple SD Gothic Neo",
)
_ARABIC_FONT_NAMES = (
    "Noto Naskh Arabic", "Noto Sans Arabic", "Segoe UI", "Tahoma", "Arial",
)
_INDIC_FONT_NAMES = (
    "Noto Sans Devanagari", "Noto Sans Tamil", "Noto Sans Bengali",
    "Nirmala UI", "Mangal", "Latha",
)
_THAI_FONT_NAMES = ("Noto Sans Thai", "Thonburi", "Tahoma", "Segoe UI")
_CYRILLIC_OK_NAMES = ("Segoe UI", "Arial", "Helvetica Neue", "Noto Sans")  # usually covered


def _char_in_ranges(cp: int, ranges) -> bool:
    return any(lo <= cp <= hi for lo, hi in ranges)


def _has_cjk(text) -> bool:
    """True if any character falls in the CJK / Hiragana-Katakana / Hangul ranges."""
    if not text:
        return False
    for ch in str(text):
        if _char_in_ranges(ord(ch), _CJK_RANGES):
            return True
    return False


def _detect_scripts(text) -> set:
    """Return the set of script keys from ``_SCRIPT_RANGES`` present in ``text``."""
    found = set()
    if not text:
        return found
    for ch in str(text):
        cp = ord(ch)
        for name, ranges in _SCRIPT_RANGES.items():
            if name in found:
                continue
            if _char_in_ranges(cp, ranges):
                found.add(name)
        if len(found) == len(_SCRIPT_RANGES):
            break
    return found


def _needs_i18n_fonts(text) -> bool:
    """True when any non-Latin script that needs an extended font stack is present."""
    return bool(_detect_scripts(text))


def _split_font_stack(stack: str) -> list:
    """Split a CSS font-family list on top-level commas (respecting quotes)."""
    parts, buf, in_q = [], [], None
    for ch in stack:
        if in_q:
            buf.append(ch)
            if ch == in_q:
                in_q = None
            continue
        if ch in ("'", '"'):
            in_q = ch
            buf.append(ch)
            continue
        if ch == ",":
            parts.append("".join(buf).strip())
            buf = []
            continue
        buf.append(ch)
    if buf:
        parts.append("".join(buf).strip())
    return [p for p in parts if p]


def _augment_font_stack(base_stack: str, text: str) -> str:
    """Insert script-specific fallback fonts before the generic family when needed.

    English-only (and other Latin-only) text returns ``base_stack`` unchanged.
    CJK detection alone is enough to unlock the CJK stack; other scripts add their
    own named fonts. No fonts are shipped or looked up on disk.
    """
    scripts = _detect_scripts(text)
    if not scripts:
        return base_stack

    extras = []
    if "cjk" in scripts:
        extras.extend(_CJK_FONT_NAMES)
    if "arabic" in scripts:
        extras.extend(_ARABIC_FONT_NAMES)
    if "hebrew" in scripts:
        extras.extend(("Segoe UI", "Tahoma", "Arial Hebrew", "Noto Sans Hebrew"))
    if scripts & {"devanagari", "tamil", "bengali", "gujarati", "gurmukhi",
                  "kannada", "malayalam", "telugu"}:
        extras.extend(_INDIC_FONT_NAMES)
    if "thai" in scripts:
        extras.extend(_THAI_FONT_NAMES)
    if "cyrillic" in scripts:
        extras.extend(_CYRILLIC_OK_NAMES)

    parts = _split_font_stack(base_stack)
    generics = {"serif", "sans-serif", "monospace", "cursive", "fantasy", "system-ui"}
    named, generic_tail = [], []
    for p in parts:
        bare = p.strip().strip("'\"").lower()
        if bare in generics:
            generic_tail.append(p)
        else:
            named.append(p)

    existing = {p.strip().strip("'\"").lower() for p in named}
    inserted = []
    for name in extras:
        if name.lower() not in existing:
            inserted.append(f"'{name}'")
            existing.add(name.lower())

    merged = named + inserted + (generic_tail or ["sans-serif"])
    return ", ".join(merged)


def _text_direction(text: str) -> str:
    """Page ``dir``: ``rtl`` only when Arabic/Hebrew outweighs basic Latin letters.

    Avoids flipping a mostly-English dashboard to RTL because of one Arabic KPI
    label (review m2). Fully Arabic/Hebrew pages still get ``rtl``.
    """
    rtl_chars = ltr_chars = 0
    for ch in str(text or ""):
        cp = ord(ch)
        if _char_in_ranges(cp, _SCRIPT_RANGES["arabic"]) or _char_in_ranges(
            cp, _SCRIPT_RANGES["hebrew"]
        ):
            rtl_chars += 1
        elif ch.isalpha() and cp < 0x0300:
            ltr_chars += 1
    if rtl_chars == 0:
        return "ltr"
    return "rtl" if rtl_chars >= ltr_chars else "ltr"


def _strip_tags(html: str) -> str:
    """Drop HTML/SVG tags so script/direction scans see visible text only."""
    import re
    return re.sub(r"<[^>]+>", " ", html or "")


def _scan_dashboard_text(title, subtitle, body, footnote, brand_name) -> str:
    """Concatenate visible strings used to decide font / direction augmentation."""
    parts = [title or "", subtitle or "", footnote or "", brand_name or "",
             _strip_tags(body)]
    return "\n".join(parts)

def dashboard(title, blocks, subtitle=None, as_of=None, out_path=None,
              footnote=None, theme=None) -> str:
    """Assemble blocks into a full, self-contained HTML page.

    title     — page heading.
    blocks    — list of HTML fragments from the building blocks above.
    subtitle  — optional sub-heading under the title.
    as_of     — "as-of" stamp shown top-right (e.g. '14 Jun 2026'); defaults to today.
    out_path  — if given, write the file and return its path; else return the HTML.
    footnote  — optional extra line in the footer (above the standard disclaimer).
    theme     — optional (partial) theme dict {brand_name, logo_path, font,
                font_heading, colours} overriding the neutral default for the page
                shell (header rule, wordmark/logo, fonts, footer brand line). To also
                re-skin the chart colours, call apply_theme(theme) before the blocks.

    When any label/title/body text contains CJK (or other non-Latin scripts), the
    font stacks are augmented with a browser fallback chain so glyphs render
    instead of □□□. English-only dashboards keep the theme fonts unchanged.
    """
    rt = _resolve_theme(theme)
    brand, font, font_heading = rt["colours"], rt["font"], rt["font_heading"]
    logo_html = _logo_html() if theme is None else _logo_for(rt)
    as_of = as_of or _today_str()
    sub_h = f'<div class="sub">{_e(subtitle)}</div>' if subtitle else ""
    body = "".join(blocks) if isinstance(blocks, (list, tuple)) else str(blocks)
    foot_extra = f'<div>{_e(footnote)}</div>' if footnote else ""
    # the interaction script is inert unless a block opted in (sortable/filter/toggle)
    script = _INTERACT_JS if any(k in body for k in
             ('data-sortable', 'data-filter', 'data-toggle')) else ""

    scan = _scan_dashboard_text(title, subtitle, body, footnote, rt["brand_name"])
    font = _augment_font_stack(font, scan)
    font_heading = _augment_font_stack(font_heading, scan)
    text_dir = _text_direction(scan)
    # Prefer a more specific lang when CJK-only content is clear; otherwise keep en.
    html_lang = "zh" if (_has_cjk(scan) and not (_detect_scripts(scan) - {"cjk"})) else "en"

    doc = _PAGE.format(
        title=_e(title), subtitle=sub_h, asof=_e(as_of), body=body,
        foot_extra=foot_extra, brand=brand, font=font, font_heading=font_heading,
        logo=logo_html, brand_name=_e(rt["brand_name"]), script=script,
        year=as_of.split()[-1] if as_of and as_of[-1].isdigit() else "",
        html_lang=html_lang, text_dir=text_dir)
    if out_path:
        Path(out_path).write_text(doc, encoding="utf-8")
        return str(out_path)
    return doc


def open_in_browser(path) -> None:
    """Open a rendered dashboard in the default browser (for review / print-to-PDF)."""
    webbrowser.open(Path(path).resolve().as_uri())


def _today_str() -> str:
    d = dt.date.today()
    return d.strftime("%d %b %Y")


# The page shell. Note: CSS braces are doubled for str.format; {brand[..]} /
# {font} / {title} etc. are the only single-brace fields.
_PAGE = """<!DOCTYPE html><html lang="{html_lang}" dir="{text_dir}"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{title}</title><style>
:root{{--burg:{brand[burgundy]};--rose:{brand[rose]};--ink:{brand[ink]};
--grey:{brand[grey]};--line:{brand[grey_lt]};--bg:{brand[bg]}}}
*{{box-sizing:border-box}}
body{{margin:0;background:var(--bg);color:var(--ink);font:14px/1.45 {font}}}
.wrap{{max-width:1040px;margin:0 auto;padding:0 20px 40px}}
header{{display:flex;align-items:flex-end;justify-content:space-between;
gap:16px;border-bottom:3px solid var(--burg);padding:22px 0 12px;margin-bottom:18px}}
.brand{{display:flex;align-items:center;gap:13px}}
.brand .logo{{height:46px;width:auto;display:block}}
.mark{{font-family:{font_heading};font-weight:700;letter-spacing:2px;color:var(--ink);font-size:24px}}
h1{{font-family:{font_heading};font-size:21px;margin:0;font-weight:700}}
.sub{{color:var(--grey);font-size:13px;margin-top:2px}}
.asof{{text-align:right;color:var(--grey);font-size:12px;white-space:nowrap}}
.asof b{{display:block;color:var(--ink);font-size:13px}}
.btn{{margin-top:8px;font:12px {font};background:var(--burg);color:#fff;border:0;
border-radius:5px;padding:6px 12px;cursor:pointer}}
.sec{{margin:22px 0}}
.sec h2{{font-family:{font_heading};font-size:15px;color:var(--burg);margin:0 0 10px;
border-left:4px solid var(--burg);padding-left:9px}}
.block{{background:#fff;border:1px solid var(--line);border-radius:9px;
padding:14px 16px;margin:0 0 14px}}
.bt{{font-size:13px;margin:0 0 10px;color:var(--ink)}}
.kpi-row{{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));
gap:14px;margin:0 0 14px}}
.kpi{{background:#fff;border:1px solid var(--line);border-top:4px solid var(--burg);
border-radius:9px;padding:14px 16px}}
.kpi-val{{font-family:{font_heading};font-size:30px;font-weight:700;line-height:1.1}}
.kpi-lbl{{color:var(--grey);font-size:12px;margin-top:4px;text-transform:uppercase;
letter-spacing:.4px}}
.kpi-sub{{color:var(--ink);font-size:12px;margin-top:6px}}
.lay{{display:grid;gap:14px}}
.lay .block{{margin:0}}
.chart{{width:100%;height:auto;display:block}}
.chart .v{{font:11px {font};fill:var(--ink)}}
.chart .x{{font:11px {font};fill:var(--grey)}}
.donut-wrap{{display:flex;gap:18px;align-items:center;flex-wrap:wrap}}
.donut{{max-width:240px}}
.dn-c{{font:700 26px {font};fill:var(--burg)}}
.dn-l{{font:11px {font};fill:var(--grey)}}
.heatmap .hm-v{{font:10px {font};pointer-events:none}}
.spark-wrap{{display:flex;align-items:center;gap:12px}}
.spark{{flex:0 0 auto}}
.spark-meta{{display:flex;flex-direction:column;gap:2px;font-size:12px}}
.spark-last{{font-family:{font_heading};font-size:18px;font-weight:700;color:var(--ink)}}
.spark-delta{{color:var(--grey)}}
.note{{color:var(--grey);font-size:11px;margin-top:6px;font-style:italic}}
.legend{{display:flex;flex-wrap:wrap;gap:10px 16px;margin-top:8px;font-size:12px}}
.legend.col{{flex-direction:column;gap:6px}}
.lg{{display:inline-flex;align-items:center;gap:6px;color:var(--ink)}}
.lg i{{width:11px;height:11px;border-radius:2px;display:inline-block}}
.pill{{display:inline-block;padding:2px 9px;border-radius:11px;font-size:11px;
font-weight:600}}
table.grid{{border-collapse:collapse;width:100%;font-size:13px}}
table.grid th{{background:var(--burg);color:#fff;text-align:left;padding:7px 9px;
font-weight:600;font-size:12px}}
table.grid td{{border-bottom:1px solid var(--line);padding:6px 9px}}
table.grid tbody tr:nth-child(even){{background:{brand[pink_vlt]}55}}
table.grid[data-sortable] th{{cursor:pointer;user-select:none}}
.so{{font-size:9px;opacity:.85}}
.legend[data-toggle] .lg{{cursor:pointer}}
.legend[data-toggle] .lg.off{{opacity:.35}}
.filterbar{{font-size:12px;color:var(--grey);margin:0 0 10px}}
.filterbar select{{font:12px {font};padding:3px 7px;border:1px solid var(--line);
border-radius:5px;color:var(--ink);background:#fff}}
.empty{{color:var(--grey);text-align:center;padding:18px;font-style:italic}}
footer{{margin-top:26px;border-top:1px solid var(--line);padding-top:10px;
color:var(--grey);font-size:11px}}
@media print{{
 body{{background:#fff}}
 .btn{{display:none}}
 .wrap{{max-width:none;padding:0}}
 .block,.kpi,.sec{{break-inside:avoid}}
 header{{padding-top:0}}
}}
</style></head><body><div class="wrap">
<header>
 <div>
  <div class="brand">{logo}<h1>{title}</h1></div>
  {subtitle}
 </div>
 <div class="asof">as of<b>{asof}</b>
  <button class="btn" onclick="window.print()">Print / Save&nbsp;PDF</button></div>
</header>
{body}
<footer>{foot_extra}
 {brand_name}. Generated {asof}. A draft for review by a
 qualified person, not advice.</footer>
</div>{script}</body></html>"""


# Optional inline interactivity (vanilla JS, no library/CDN). Injected by
# dashboard() only when a block opted in. Powers: click-to-sort tables,
# click-legend to toggle line series, and a dropdown row-filter. Kept tiny.
_INTERACT_JS = """<script>
(function(){
 document.querySelectorAll('table.grid[data-sortable]').forEach(function(t){
  var ths=t.tHead.rows[0].cells;
  [].forEach.call(ths,function(th,ci){
   th.addEventListener('click',function(){
    var tb=t.tBodies[0],rows=[].slice.call(tb.rows),asc=th.getAttribute('data-asc')!=='1';
    [].forEach.call(ths,function(o){o.removeAttribute('data-asc');var s=o.querySelector('.so');if(s)s.remove();});
    th.setAttribute('data-asc',asc?'1':'0');
    var num=rows.every(function(r){var x=(r.cells[ci].textContent||'').replace(/[^0-9.\\-]/g,'');return x!==''&&!isNaN(x);});
    rows.sort(function(a,b){var x=a.cells[ci].textContent.trim(),y=b.cells[ci].textContent.trim();
     if(num){x=parseFloat(x.replace(/[^0-9.\\-]/g,''))||0;y=parseFloat(y.replace(/[^0-9.\\-]/g,''))||0;}
     return (x>y?1:x<y?-1:0)*(asc?1:-1);});
    rows.forEach(function(r){tb.appendChild(r);});
    var a=document.createElement('span');a.className='so';a.textContent=asc?' \\u25B2':' \\u25BC';th.appendChild(a);
   });
  });
 });
 document.querySelectorAll('.legend[data-toggle]').forEach(function(lg){
  var svg=lg.parentNode.querySelector('svg');
  [].forEach.call(lg.querySelectorAll('.lg'),function(item,i){
   item.addEventListener('click',function(){
    var g=svg.querySelector('[data-series="'+i+'"]');if(!g)return;
    var off=g.getAttribute('data-off')!=='1';
    g.setAttribute('data-off',off?'1':'0');g.style.display=off?'none':'';
    item.classList.toggle('off',off);
   });
  });
 });
 document.querySelectorAll('select[data-filter]').forEach(function(sel){
  var ci=+sel.getAttribute('data-col'),t=document.getElementById(sel.getAttribute('data-target'));
  sel.addEventListener('change',function(){var v=sel.value;
   [].forEach.call(t.tBodies[0].rows,function(r){
    r.style.display=(v==='__all'||r.cells[ci].textContent.trim()===v)?'':'none';});
  });
 });
})();
</script>"""


# --------------------------------------------------------------------------- #
# Optional adapters — feed the engine from existing toolkit .xlsx stores.
# Kept light and dependency-optional so the engine stands alone.
# --------------------------------------------------------------------------- #
def _load_ingest():
    """Import the shared ingest engine (toolkit-root scripts/, or a vendored sibling) if it's
    reachable — for its multi-sheet-safe .xlsx reader. Returns the module or None."""
    import importlib
    from pathlib import Path as _P
    for p in (_P(__file__).resolve().parents[3] / "scripts",   # toolkit-root engine
              _P(__file__).resolve().parent / "scripts"):       # vendored sibling
        if (p / "ingest.py").is_file():
            if str(p) not in sys.path:
                sys.path.insert(0, str(p))
            try:
                return importlib.import_module("ingest")
            except ImportError:
                return None
    return None


def _rows_to_dicts(rows) -> list[dict]:
    if not rows:
        return []
    header = [str(h) if h is not None else "" for h in rows[0]]
    out = []
    for row in rows[1:]:
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        out.append({header[i]: row[i] for i in range(min(len(header), len(row)))})
    return out


def rows_from_xlsx(path, sheet=None) -> list[dict]:
    """Read a simple header+rows .xlsx into a list of dicts (needs openpyxl).
    Use to point the engine at any header+rows .xlsx (e.g. a clean table from
    data-tidy / data-extract).

    Multi-tab safe: it won't silently read the 'active' sheet. When the shared `ingest` engine
    is reachable it uses `ingest.read_xlsx` (auto-selects the single data sheet; raises
    `SheetSelectionRequired` if several tabs hold data). Standalone (ingest absent), the same
    rule is applied locally. Pass `sheet=<name>` to read a specific tab."""
    ing = _load_ingest()
    if ing is not None:
        rows, _ = ing.read_xlsx(path, sheet=sheet)
        return _rows_to_dicts(rows)

    # Standalone fallback — mirror ingest's single-sheet safety without importing it.
    import openpyxl  # optional dependency
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        def _has_data(ws):
            for r in ws.iter_rows(values_only=True):
                if any(c is not None and str(c).strip() != "" for c in r):
                    return True
            return False
        if sheet is not None:
            ws = wb[sheet]
        else:
            cands = [ws for ws in wb.worksheets
                     if ws.sheet_state == "visible" and _has_data(ws)]
            if len(cands) == 1:
                ws = cands[0]
            elif not cands:
                ws = wb.active
            else:
                names = ", ".join(repr(ws.title) for ws in cands)
                raise ValueError(
                    f"{Path(path).name} has {len(cands)} non-empty sheets ({names}); "
                    f"pass sheet=<name> to choose one.")
        rows = [["" if c is None else c for c in r] for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()
    return _rows_to_dicts(rows)


# --------------------------------------------------------------------------- #
# Analyse → visualise handoff
# --------------------------------------------------------------------------- #
def _as_number(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float, Decimal)):
        return v
    try:
        return Decimal(str(v))
    except Exception:
        try:
            return float(v)
        except (TypeError, ValueError):
            return None


def _analysis_results(analysis) -> list[dict]:
    if isinstance(analysis, list):
        return analysis
    if isinstance(analysis, dict):
        return list(analysis.get("results") or [])
    return []


def _blocks_for_analysis_op(op: str, name: str, result: dict, *, max_groups: int = 10) -> list[dict]:
    """Map one analyse engine result onto declarative visualise blocks.

    Numbers stay whatever the engine produced — this never recomputes metrics.
    Unknown ops are skipped (so older/newer analysis.json files degrade cleanly).
    """
    result = result or {}
    title = name or op
    out: list[dict] = []

    if op == "numeric_summary":
        items = [
            {"label": "n", "value": result.get("n"), "status": "brand"},
            {"label": "Total", "value": result.get("total"), "status": "brand"},
            {"label": "Mean", "value": result.get("mean"), "status": "grey"},
            {"label": "Median", "value": result.get("median"), "status": "grey"},
        ]
        out.append({"type": "kpi_row", "items": [i for i in items if i["value"] is not None]})
    elif op == "breakdown":
        groups = (result.get("groups") or [])[:max_groups]
        measure = "total" if groups and "total" in groups[0] else "count"
        data = [{"label": g.get("key"), "value": g.get(measure)} for g in groups]
        if data:
            out.append({"type": "bar_chart", "title": f"{title} — by {result.get('by', 'group')}",
                        "data": data})
            if measure == "total" and any(g.get("share") is not None for g in groups):
                out.append({"type": "donut_chart", "title": f"{title} — share", "data": data})
    elif op == "period_series":
        periods = result.get("periods") or []
        measure = "total" if periods and "total" in periods[0] else "count"
        series = [(p.get("period"), p.get(measure)) for p in periods]
        if series:
            out.append({"type": "line_chart", "title": f"{title} — {result.get('grain', 'period')}",
                        "data": [{"label": a, "value": b} for a, b in series]})
            out.append({"type": "sparkline", "title": f"{title} — shape",
                        "data": [{"label": a, "value": b} for a, b in series]})
            if len(series) >= 2:
                steps = [{"label": series[0][0], "value": series[0][1], "kind": "start"}]
                for i in range(1, len(series)):
                    lab, val = series[i]
                    delta = periods[i].get("delta")
                    if delta is None:
                        cur, base = _as_number(val), _as_number(series[i - 1][1])
                        delta = (cur - base) if cur is not None and base is not None else 0
                    steps.append({"label": f"Δ {lab}", "value": delta, "kind": "delta"})
                steps.append({"label": series[-1][0], "value": series[-1][1], "kind": "total"})
                out.append({"type": "waterfall", "title": f"{title} — period bridge", "steps": steps})
    elif op == "ageing":
        buckets = [b for b in (result.get("buckets") or []) if b.get("count")]
        measure = "total" if buckets and "total" in buckets[0] else "count"
        data = [{"label": b.get("bucket"), "value": b.get(measure)} for b in buckets]
        if data:
            out.append({"type": "bar_chart", "title": f"{title} — ageing", "data": data})
    elif op == "outliers_iqr":
        items = [
            {"label": "Outliers", "value": result.get("n_outliers", len(result.get("outliers") or [])),
             "status": "amber"},
            {"label": "Fence low", "value": result.get("fence_low"), "status": "grey"},
            {"label": "Fence high", "value": result.get("fence_high"), "status": "grey"},
        ]
        out.append({"type": "kpi_row", "items": [i for i in items if i["value"] is not None]})
    elif op == "currency_mix":
        currencies = result.get("currencies") or []
        out.append({"type": "kpi_row", "items": [
            {"label": "Currencies", "value": len(currencies), "sub": ", ".join(currencies) or "—",
             "status": "amber" if len(currencies) > 1 else "green"}
        ]})
    elif op == "concentration":
        out.append({"type": "kpi_row", "items": [
            {"label": "HHI", "value": result.get("hhi"), "status": "brand"},
            {"label": "Top-N share", "value": result.get("top_n_share"), "status": "brand"},
            {"label": "Groups to 80%", "value": result.get("groups_to_80"), "status": "grey"},
            {"label": "Class", "value": result.get("classification"), "status": "amber"},
        ]})
    elif op == "gini":
        out.append({"type": "kpi_row", "items": [
            {"label": "Gini", "value": result.get("gini"), "status": "brand"},
            {"label": "Class", "value": result.get("classification"), "status": "grey"},
            {"label": "n", "value": result.get("n"), "status": "grey"},
        ]})
    elif op == "distribution":
        out.append({"type": "kpi_row", "items": [
            {"label": "Skewness", "value": result.get("skewness"), "status": "brand"},
            {"label": "Kurtosis", "value": result.get("kurtosis"), "status": "brand"},
            {"label": "Shape", "value": result.get("classification"), "status": "grey"},
        ]})
        # A histogram shows the shape the skew/kurt coefficients describe — but
        # only when the engine carried the raw values. distribution() doesn't, so
        # the block is a hint the agent must re-supply the column, not a ready render.
        if result.get("values"):
            out.append({"type": "histogram", "title": f"{title} — shape",
                        "values": result["values"]})
    elif op == "percentile":
        if "value" in result:
            out.append({"type": "kpi_row", "items": [
                {"label": "Percentile", "value": result.get("value"), "status": "brand"},
                {"label": "n", "value": result.get("n"), "status": "grey"},
            ]})
        else:
            items = [{"label": f"p{float(k)*100:g}", "value": v, "status": "brand"}
                     for k, v in result.items() if _as_number(k) is not None]
            if items:
                out.append({"type": "kpi_row", "items": items})
    elif op == "trend":
        out.append({"type": "kpi_row", "items": [
            {"label": "Slope", "value": result.get("slope"), "status": "brand"},
            {"label": "R²", "value": result.get("r_squared"), "status": "grey"},
            {"label": "Direction", "value": result.get("classification"), "status": "amber"},
        ]})
    elif op == "rolling":
        series = result.get("series") or []
        data = [{"label": p.get("period"), "value": p.get("value")}
                for p in series if p.get("value") is not None]
        if data:
            out.append({"type": "line_chart",
                        "title": f"{title} — rolling {result.get('func', 'mean')}",
                        "data": data})
    elif op == "seasonality":
        seasons = result.get("seasons") or []
        grain = result.get("grain", "month")
        label = (lambda s: f"Q{s}" if grain == "quarter" else f"M{s:02d}")
        data = [{"label": label(s.get("season")), "value": s.get("average")}
                for s in seasons if s.get("count")]
        if data:
            out.append({"type": "bar_chart", "title": f"{title} — seasonal average", "data": data})
    elif op == "pivot":
        matrix = result.get("matrix") or []
        if matrix:
            out.append({"type": "heatmap", "title": title,
                        "matrix": matrix,
                        "row_labels": result.get("row_keys") or result.get("rows"),
                        "col_labels": result.get("col_keys") or result.get("cols"),
                        "scale": "sequential"})
            # A stacked bar shows composition per row — more readable than a heatmap
            # when there are few columns (segments) and the question is "what makes
            # up each row", not "which cell is hottest". Pass the pivot result straight
            # through; stacked_bar accepts the {row_keys, col_keys, matrix} shape.
            if len(result.get("col_keys") or []) <= 8:
                out.append({"type": "stacked_bar", "title": f"{title} — composition",
                            "data": result})
    elif op == "correlation_matrix":
        matrix = result.get("matrix") or []
        if matrix:
            out.append({"type": "heatmap", "title": f"{title} — correlation",
                        "matrix": matrix, "row_labels": result.get("columns"),
                        "col_labels": result.get("columns"), "scale": "diverging", "mid": 0})
    elif op == "cohort":
        matrix = result.get("retention") or result.get("matrix") or []
        if matrix:
            max_off = result.get("max_offset", len(matrix[0]) - 1 if matrix else 0)
            out.append({"type": "heatmap", "title": f"{title} — retention",
                        "matrix": matrix,
                        "row_labels": result.get("cohorts"),
                        "col_labels": [f"+{i}" for i in range(int(max_off) + 1)],
                        "scale": "sequential"})
    elif op == "compare_series":
        points = result.get("points") or []
        a_label = result.get("a_label", "A")
        b_label = result.get("b_label", "B")
        if points:
            out.append({"type": "line_chart", "title": f"{title} — {a_label} vs {b_label}",
                        "data": {
                            a_label: [{"label": p.get("key"), "value": p.get(a_label)} for p in points],
                            b_label: [{"label": p.get("key"), "value": p.get(b_label)} for p in points],
                        }, "toggle": True})
        out.append({"type": "kpi_row", "items": [
            {"label": "n paired", "value": result.get("n"), "status": "brand"},
            {"label": "Correlation", "value": result.get("correlation"), "status": "grey"},
            {"label": "Best align", "value": result.get("best_alignment"), "status": "amber"},
        ]})
    elif op == "join_on":
        report = result.get("report") or result
        out.append({"type": "kpi_row", "items": [
            {"label": "Matched", "value": report.get("matched"), "status": "green"},
            {"label": "Left only", "value": report.get("left_only"), "status": "amber"},
            {"label": "Right only", "value": report.get("right_only"), "status": "amber"},
        ]})
    return [b for b in out if b.get("type") != "kpi_row" or b.get("items")]


def suggest_blocks_from_analysis(analysis, *, ops=None, max_groups: int = 10) -> list[dict]:
    """Turn a data-analyse ``analysis.json`` (or its ``results`` list) into
    declarative dashboard blocks the agent can show, edit, then render.

    This is the handoff contract: analyse computes; visualise proposes drawings.
    Pass ``ops`` to keep only named operations (by ``op`` or ``name``).
    """
    wanted = set(ops) if ops else None
    blocks: list[dict] = []
    for item in _analysis_results(analysis):
        if not isinstance(item, dict):
            continue
        op = item.get("op")
        name = item.get("name") or op
        if wanted is not None and op not in wanted and name not in wanted:
            continue
        mapped = _blocks_for_analysis_op(op, name, item.get("result") or {}, max_groups=max_groups)
        if mapped:
            blocks.append({"type": "section", "title": str(name), "blocks": mapped})
    return blocks


def blocks_from_analysis(analysis, *, ops=None, max_groups: int = 10) -> list[str]:
    """Convenience: ``suggest_blocks_from_analysis`` rendered to HTML fragments."""
    specs = suggest_blocks_from_analysis(analysis, ops=ops, max_groups=max_groups)

    def _render(spec: dict) -> str:
        kind = spec.get("type")
        if kind == "kpi_row":
            return kpi_row(spec.get("items", []))
        if kind == "bar_chart":
            return bar_chart(spec.get("data", []), title=spec.get("title"), unit=spec.get("unit", ""))
        if kind == "line_chart":
            return line_chart(spec.get("data", []), title=spec.get("title"),
                              unit=spec.get("unit", ""), toggle=spec.get("toggle", False))
        if kind == "donut_chart":
            return donut_chart(spec.get("data", []), title=spec.get("title"), centre=spec.get("centre"))
        if kind == "heatmap":
            return heatmap(spec.get("matrix", []), row_labels=spec.get("row_labels"),
                           col_labels=spec.get("col_labels"), title=spec.get("title"),
                           scale=spec.get("scale", "sequential"), mid=spec.get("mid", 0),
                           unit=spec.get("unit", ""))
        if kind == "sparkline":
            return sparkline(spec.get("data", []), title=spec.get("title"),
                             show_last=spec.get("show_last", True), unit=spec.get("unit", ""))
        if kind == "waterfall":
            return waterfall(spec.get("steps", []), title=spec.get("title"), unit=spec.get("unit", ""))
        if kind == "section":
            return section(spec.get("title", ""), *[_render(c) for c in spec.get("blocks", [])])
        if kind == "grid":
            return grid(*[_render(c) for c in spec.get("blocks", [])], cols=int(spec.get("cols", 2)))
        raise ValueError(f"unsupported block type from analysis: {kind!r}")

    return [_render(spec) for spec in specs]


# --------------------------------------------------------------------------- #
# Self-test — builds a demo dashboard (no external data, no network).
# --------------------------------------------------------------------------- #
if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else "dashboard-selftest.html"
    tasks = [
        {"ID": 1, "Task": "Circulate weekly tracker", "Owner": "Alex", "Days late": 12},
        {"ID": 2, "Task": "Confirm vendor quotes", "Owner": "Jordan", "Days late": 0},
        {"ID": 3, "Task": "Draft planning outline", "Owner": "Sam", "Days late": 0},
    ]
    blocks = [
        kpi_row([
            {"label": "Open tasks", "value": 12, "status": "brand"},
            {"label": "Overdue", "value": 3, "sub": "oldest 12 days", "status": "red"},
            {"label": "Due this week", "value": 5, "status": "amber"},
            {"label": "Done (7d)", "value": 9, "status": "green"},
        ]),
        section("Throughput",
            grid(
                bar_chart([("Mon", 4), ("Tue", 7), ("Wed", 5), ("Thu", 6), ("Fri", 3)],
                          title="Tasks completed by day"),
                donut_chart([("Compliance", 8), ("Finance", 5), ("Marketing", 3),
                             ("Legal", 4)], title="Open tasks by function"),
            )),
        section("Trend",
            grid(
                line_chart({"Opened": [("W1", 10), ("W2", 14), ("W3", 9), ("W4", 12)],
                            "Closed": [("W1", 8), ("W2", 11), ("W3", 13), ("W4", 10)]},
                           title="Opened vs closed (4 weeks)"),
                sparkline([("W1", 10), ("W2", 14), ("W3", 9), ("W4", 12)],
                          title="Opened — sparkline"),
                cols=2)),
        section("Composition",
            grid(
                heatmap([[8, 3], [2, 5]], row_labels=["Compliance", "Finance"],
                        col_labels=["Open", "Done"], title="Load matrix"),
                waterfall([("Open", 12, "start"), ("Done", -5, "delta"),
                           ("New", 4, "delta"), ("Close", 11, "total")],
                          title="Open → close bridge"),
                cols=2)),
        section("Detail",
            table(tasks, columns=["ID", "Task", "Owner", "Days late"],
                  title="Outstanding tasks",
                  rag={"Days late": lambda v: "red" if (v or 0) > 7 else
                       ("amber" if (v or 0) > 0 else "green")})),
    ]
    path = dashboard("Operations dashboard", blocks, subtitle="Demo / self-test",
                     as_of="14 Jun 2026", out_path=out)
    print(f"[self-test] built {path} ({Path(path).stat().st_size:,} bytes)")
    print("[self-test] pills:", status_pill("On track", "green"),
          status_pill("At risk", "amber"), status_pill("Breach", "red"))
