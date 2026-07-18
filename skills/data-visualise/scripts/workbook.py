"""workbook.py — native Excel chart packager for data-visualise.

Chart-only. Writes a real .xlsx with worksheet data ranges + embedded Excel
charts (openpyxl). HTML dashboards stay in ``viz.py``; this module never emits
HTML or SVG.

Vocabulary is aligned with AionUi / OfficeCLI chart props where they map cleanly
onto openpyxl (see ``references/workbook-charts.md``):

    chartType: column | bar | line | pie | doughnut
    title, categories, seriesN / data, legend, colors, anchor, width, height

Waterfall uses the classic stacked-bar bridge (invisible base + rise/fall),
with increase / decrease / total colours in the OfficeCLI spirit
(``increaseColor`` / ``decreaseColor`` / ``totalColor``).

Quick start
-----------
    from workbook import write_charts_xlsx, charts_from_analysis
    write_charts_xlsx("out.xlsx", [
        {"chart_type": "column", "title": "By region",
         "categories": ["North", "South"],
         "series": [{"name": "Amount", "values": [120, 80]}]},
        {"chart_type": "line", "title": "Monthly",
         "categories": ["2026-01", "2026-02"],
         "series": [{"name": "Total", "values": [100, 130]}]},
    ], workbook_title="Insight charts")
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import Any

# OfficeCLI / AionUi chartType names we support via openpyxl.
SUPPORTED_CHART_TYPES = frozenset({
    "column", "bar", "line", "pie", "doughnut", "waterfall",
})

# Fallback palette — used only if the visualise theme can't be resolved (hex without '#').
DEFAULT_COLORS = ("163F3A", "4FB3A0", "20574F", "2E7D57", "B26B00", "9B2226")

# Series order mirrors viz._series_palette() so an Excel workbook and an HTML dashboard
# built from the same theme use the same colours in the same order.
_SERIES_TOKENS = ("burgundy", "rose", "pink", "green", "amber", "grey", "pink_lt")


def _hex(value: Any) -> str:
    return str(value).lstrip("#").upper()


def _theme_colors(theme: Any = None) -> tuple[list[str], str, str, str]:
    """Resolve (series palette, increase, decrease, total) from the **visualise theme**.

    The workbook is branded by the same `theme` dict that skins the HTML dashboard, so a
    white-label firm gets its colours in both artefacts and there is one palette to maintain.
    `viz` is imported lazily (viz imports this module lazily too) and any failure degrades to
    DEFAULT_COLORS rather than breaking chart generation.
    """
    try:
        import viz

        colours = viz._resolve_theme(theme)["colours"]
        series = [_hex(colours[t]) for t in _SERIES_TOKENS if colours.get(t)]
        return (
            series or list(DEFAULT_COLORS),
            _hex(colours.get("green", "2E7D57")),
            _hex(colours.get("red", "9B2226")),
            _hex(colours.get("burgundy", "163F3A")),
        )
    except Exception:  # noqa: BLE001 - theming must never break the workbook
        return list(DEFAULT_COLORS), "2E7D57", "9B2226", "163F3A"


def _as_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return float(int(v))
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, Decimal):
        return float(v)
    try:
        return float(Decimal(str(v)))
    except Exception:
        try:
            return float(v)
        except (TypeError, ValueError):
            return None


def _sheet_title(name: str, used: set[str]) -> str:
    base = "".join(ch if ch not in r'[]:*?/\\' else "_" for ch in str(name or "Chart"))[:31] or "Chart"
    title, n = base, 2
    while title in used:
        suffix = f"_{n}"
        title = (base[: 31 - len(suffix)] + suffix)
        n += 1
    used.add(title)
    return title


def _norm_series(series: Any) -> list[tuple[str, list[float | None]]]:
    """Accept [{name, values}], {name: values}, or [[name, ...values]]."""
    out: list[tuple[str, list[float | None]]] = []
    if series is None:
        return out
    if isinstance(series, dict) and "values" not in series:
        for name, values in series.items():
            out.append((str(name), [_as_float(v) for v in values]))
        return out
    if isinstance(series, dict):
        series = [series]
    for item in series:
        if isinstance(item, dict):
            name = str(item.get("name") or item.get("label") or "Series")
            values = item.get("values") or item.get("data") or []
            out.append((name, [_as_float(v) for v in values]))
        elif isinstance(item, (list, tuple)) and item:
            out.append((str(item[0]), [_as_float(v) for v in item[1:]]))
    return out


def _apply_series_colors(chart, colors: list[str] | None) -> None:
    from openpyxl.chart.shapes import GraphicalProperties

    palette = [c.lstrip("#") for c in (colors or list(DEFAULT_COLORS))]
    for i, ser in enumerate(getattr(chart, "series", []) or []):
        hex_color = palette[i % len(palette)]
        try:
            ser.graphicalProperties = GraphicalProperties(solidFill=hex_color)
        except Exception:
            try:
                ser.graphicalProperties = GraphicalProperties()
                ser.graphicalProperties.solidFill = hex_color
            except Exception:
                pass


def _make_chart(chart_type: str, title: str | None, legend: bool):
    from openpyxl.chart import BarChart, DoughnutChart, LineChart, PieChart
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.legend import Legend

    ctype = (chart_type or "column").lower()
    if ctype not in SUPPORTED_CHART_TYPES:
        raise ValueError(
            f"unsupported chart_type {chart_type!r}; "
            f"supported (OfficeCLI-aligned): {sorted(SUPPORTED_CHART_TYPES)}"
        )
    if ctype == "column":
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
    elif ctype == "bar":
        chart = BarChart()
        chart.type = "bar"
        chart.grouping = "clustered"
    elif ctype == "line":
        chart = LineChart()
        chart.style = 10
    elif ctype == "pie":
        chart = PieChart()
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showVal = False
        chart.dataLabels.showCatName = False
    elif ctype == "doughnut":
        chart = DoughnutChart()
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
    else:
        # waterfall built separately
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
    if title:
        chart.title = title
    if legend:
        chart.legend = Legend(legendPos="b")
    else:
        chart.legend = None
    return chart


def add_chart(
    ws,
    *,
    chart_type: str,
    title: str | None = None,
    categories: list[Any] | None = None,
    series: Any = None,
    data_start_row: int = 1,
    data_start_col: int = 1,
    write_data: bool = True,
    legend: bool = True,
    colors: list[str] | None = None,
    anchor: str = "E2",
    width: float = 15,
    height: float = 10,
    increase_color: str = "2E7D57",
    decrease_color: str = "9B2226",
    total_color: str = "163F3A",
) -> Any:
    """Write category/series cells (optional) and embed a chart.

    ``chart_type`` uses AionUi/OfficeCLI names: ``column``, ``bar``, ``line``,
    ``pie``, ``doughnut``, ``waterfall``. Series values are stored on the sheet
    and the chart references those cells (OfficeCLI ``dataRange`` style) — not
    free-floating arrays.
    """
    from openpyxl.chart import Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.chart.shapes import GraphicalProperties

    raw_categories = categories
    categories = list(categories or [])
    series_list = _norm_series(series)
    if chart_type.lower() == "waterfall":
        steps = None
        if categories and isinstance(categories[0], dict):
            steps = categories
            categories = []
        return _add_waterfall_chart(
            ws, title=title, steps=steps,
            series=series_list, categories=categories or list(raw_categories or []),
            data_start_row=data_start_row, data_start_col=data_start_col,
            write_data=write_data, anchor=anchor, width=width, height=height,
            increase_color=increase_color, decrease_color=decrease_color,
            total_color=total_color,
        )
    if not categories or not series_list:
        raise ValueError("add_chart requires categories and at least one series")

    r0, c0 = data_start_row, data_start_col
    if write_data:
        ws.cell(r0, c0, "Category")
        for j, (name, _) in enumerate(series_list):
            ws.cell(r0, c0 + 1 + j, name)
        for i, cat in enumerate(categories):
            ws.cell(r0 + 1 + i, c0, cat)
            for j, (_, values) in enumerate(series_list):
                val = values[i] if i < len(values) else None
                ws.cell(r0 + 1 + i, c0 + 1 + j, val)

    n_cats = len(categories)
    n_series = len(series_list)
    chart = _make_chart(chart_type, title, legend=legend and chart_type.lower() not in {"pie", "doughnut"})
    data_ref = Reference(ws, min_col=c0 + 1, min_row=r0,
                         max_col=c0 + n_series, max_row=r0 + n_cats)
    cats_ref = Reference(ws, min_col=c0, min_row=r0 + 1, max_row=r0 + n_cats)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    if chart_type.lower() not in {"pie", "doughnut"}:
        _apply_series_colors(chart, colors)
    else:
        # Per-point colours on the first series.
        palette = [c.lstrip("#") for c in (colors or list(DEFAULT_COLORS))]
        if chart.series:
            pts = []
            for i in range(n_cats):
                pt = DataPoint(idx=i)
                pt.graphicalProperties = GraphicalProperties(solidFill=palette[i % len(palette)])
                pts.append(pt)
            chart.series[0].data_points = pts
    chart.width = width
    chart.height = height
    ws.add_chart(chart, anchor)
    return chart


def _add_waterfall_chart(
    ws, *, title, steps, series, categories, data_start_row, data_start_col,
    write_data, anchor, width, height, increase_color, decrease_color, total_color,
):
    """Stacked-column bridge: Base (hidden) + Rise + Fall + Total markers.

    Accepts either OfficeCLI-like flow steps via ``series`` values with
    ``categories`` labels, or ``steps`` as ``[{label, value, kind}]``.
    """
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.chart.shapes import GraphicalProperties

    # Normalise to (label, value, kind)
    norm = []
    if steps and isinstance(steps[0], dict):
        for i, s in enumerate(steps):
            kind = s.get("kind")
            if kind is None:
                kind = "start" if i == 0 else ("total" if i == len(steps) - 1 else "delta")
            norm.append((s.get("label"), _as_float(s.get("value")) or 0.0, kind))
    elif categories and series:
        values = series[0][1]
        for i, lab in enumerate(categories):
            kind = "start" if i == 0 else ("total" if i == len(categories) - 1 else "delta")
            norm.append((lab, values[i] if i < len(values) and values[i] is not None else 0.0, kind))
    else:
        raise ValueError("waterfall requires steps or categories+series")

    # Build helper columns: Base, Rise, Fall, Total
    rows = []
    running = 0.0
    for lab, val, kind in norm:
        if kind == "start":
            base, rise, fall, total = 0.0, 0.0, 0.0, val
            running = val
        elif kind == "total":
            base, rise, fall, total = 0.0, 0.0, 0.0, val
            running = val
        else:
            total = None
            if val >= 0:
                base, rise, fall = running, val, 0.0
            else:
                base, rise, fall = running + val, 0.0, abs(val)
            running = running + val
        rows.append((lab, base, rise, fall, total if total is not None else None))

    r0, c0 = data_start_row, data_start_col
    if write_data:
        headers = ["Category", "Base", "Rise", "Fall", "Total"]
        for j, h in enumerate(headers):
            ws.cell(r0, c0 + j, h)
        for i, (lab, base, rise, fall, total) in enumerate(rows):
            ws.cell(r0 + 1 + i, c0, lab)
            ws.cell(r0 + 1 + i, c0 + 1, base)
            ws.cell(r0 + 1 + i, c0 + 2, rise)
            ws.cell(r0 + 1 + i, c0 + 3, fall)
            ws.cell(r0 + 1 + i, c0 + 4, total if total is not None else 0)

    n = len(rows)
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.title = title
    chart.legend = None
    data_ref = Reference(ws, min_col=c0 + 1, min_row=r0, max_col=c0 + 4, max_row=r0 + n)
    cats_ref = Reference(ws, min_col=c0, min_row=r0 + 1, max_row=r0 + n)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    # Base invisible; Rise / Fall / Total coloured (OfficeCLI increase/decrease/total).
    fills = {
        0: "FFFFFF",  # Base — invisible against white plot
        1: increase_color.lstrip("#"),
        2: decrease_color.lstrip("#"),
        3: total_color.lstrip("#"),
    }
    for idx, ser in enumerate(chart.series):
        ser.graphicalProperties = GraphicalProperties(solidFill=fills.get(idx, "163F3A"))
        if idx == 0:
            # Force base series transparent-ish via white + no border where possible.
            ser.graphicalProperties.solidFill = "FFFFFF"

    chart.width = width
    chart.height = height
    ws.add_chart(chart, anchor)
    return chart


def write_charts_xlsx(
    out_path: str | Path,
    charts: list[dict[str, Any]],
    *,
    workbook_title: str | None = None,
    theme: Any = None,
) -> Path:
    """Write one sheet (+ embedded chart) per chart spec. Returns the path.

    `theme` is the same (partial) visualise theme dict that skins an HTML dashboard; it sets the
    default series palette and the waterfall increase/decrease/total colours. A per-chart
    `colors` / `increase_color` / … still overrides it.
    """
    from openpyxl import Workbook

    if not charts:
        raise ValueError("write_charts_xlsx requires at least one chart spec")
    series_colors, inc_default, dec_default, total_default = _theme_colors(theme)
    wb = Workbook()
    wb.remove(wb.active)
    used: set[str] = set()
    for i, spec in enumerate(charts):
        if not isinstance(spec, dict):
            raise ValueError(f"charts[{i}] must be an object")
        ctype = spec.get("chart_type") or spec.get("type") or "column"
        if ctype in {"chart", "from_analysis"}:
            ctype = spec.get("chart_type", "column")
        title = spec.get("title") or f"Chart {i + 1}"
        ws = wb.create_sheet(title=_sheet_title(title, used))
        categories = spec.get("steps") if ctype == "waterfall" and spec.get("steps") else spec.get("categories")
        add_chart(
            ws,
            chart_type=ctype,
            title=title,
            categories=categories,
            series=None if (ctype == "waterfall" and spec.get("steps")) else spec.get("series"),
            legend=spec.get("legend", True),
            colors=spec.get("colors") or series_colors,
            anchor=spec.get("anchor", "E2"),
            width=float(spec.get("width", 15)),
            height=float(spec.get("height", 10)),
            increase_color=spec.get("increase_color", spec.get("increaseColor", inc_default)),
            decrease_color=spec.get("decrease_color", spec.get("decreaseColor", dec_default)),
            total_color=spec.get("total_color", spec.get("totalColor", total_default)),
        )
    if workbook_title:
        try:
            wb.properties.title = workbook_title
        except Exception:
            pass
    path = Path(out_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def _analysis_results(analysis) -> list[dict]:
    if isinstance(analysis, list):
        return analysis
    if isinstance(analysis, dict):
        return list(analysis.get("results") or [])
    return []


def suggest_charts_from_analysis(analysis, *, ops=None, max_groups: int = 12) -> list[dict]:
    """Map data-analyse ``analysis.json`` results to Excel chart specs.

    Parallel to ``viz.suggest_blocks_from_analysis`` — numbers are not recomputed.
    Chart types follow OfficeCLI names (column / line / pie / waterfall).
    """
    wanted = set(ops) if ops else None
    charts: list[dict] = []
    for item in _analysis_results(analysis):
        if not isinstance(item, dict):
            continue
        op = item.get("op")
        name = item.get("name") or op
        if wanted is not None and op not in wanted and name not in wanted:
            continue
        result = item.get("result") or {}
        title = str(name)

        if op == "breakdown":
            groups = (result.get("groups") or [])[:max_groups]
            measure = "total" if groups and "total" in groups[0] else "count"
            cats = [g.get("key") for g in groups]
            vals = [g.get(measure) for g in groups]
            if cats:
                charts.append({
                    "chart_type": "column", "title": f"{title} — by {result.get('by', 'group')}",
                    "categories": cats,
                    "series": [{"name": measure, "values": vals}],
                })
                if measure == "total":
                    charts.append({
                        "chart_type": "pie", "title": f"{title} — share",
                        "categories": cats,
                        "series": [{"name": measure, "values": vals}],
                    })
        elif op == "period_series":
            periods = result.get("periods") or []
            measure = "total" if periods and "total" in periods[0] else "count"
            cats = [p.get("period") for p in periods]
            vals = [p.get(measure) for p in periods]
            if len(cats) >= 2:
                charts.append({
                    "chart_type": "line", "title": f"{title} — {result.get('grain', 'period')}",
                    "categories": cats,
                    "series": [{"name": measure, "values": vals}],
                })
                steps = [{"label": cats[0], "value": vals[0], "kind": "start"}]
                for i in range(1, len(cats)):
                    delta = periods[i].get("delta")
                    if delta is None:
                        a, b = _as_float(vals[i]), _as_float(vals[i - 1])
                        delta = (a - b) if a is not None and b is not None else 0
                    steps.append({"label": f"Δ {cats[i]}", "value": delta, "kind": "delta"})
                steps.append({"label": cats[-1], "value": vals[-1], "kind": "total"})
                charts.append({
                    "chart_type": "waterfall", "title": f"{title} — period bridge",
                    "steps": steps,
                })
        elif op == "ageing":
            buckets = [b for b in (result.get("buckets") or []) if b.get("count")]
            measure = "total" if buckets and "total" in buckets[0] else "count"
            cats = [b.get("bucket") for b in buckets]
            vals = [b.get(measure) for b in buckets]
            if cats:
                charts.append({
                    "chart_type": "column", "title": f"{title} — ageing",
                    "categories": cats,
                    "series": [{"name": measure, "values": vals}],
                })
        elif op == "seasonality":
            seasons = result.get("seasons") or []
            grain = result.get("grain", "month")
            label = (lambda s: f"Q{s}" if grain == "quarter" else f"M{int(s):02d}")
            cats = [label(s.get("season")) for s in seasons if s.get("count")]
            vals = [s.get("average") for s in seasons if s.get("count")]
            if cats:
                charts.append({
                    "chart_type": "column", "title": f"{title} — seasonal average",
                    "categories": cats,
                    "series": [{"name": "average", "values": vals}],
                })
        elif op == "rolling":
            series = result.get("series") or []
            cats = [p.get("period") for p in series if p.get("value") is not None]
            vals = [p.get("value") for p in series if p.get("value") is not None]
            if len(cats) >= 2:
                charts.append({
                    "chart_type": "line",
                    "title": f"{title} — rolling {result.get('func', 'mean')}",
                    "categories": cats,
                    "series": [{"name": result.get("func", "mean"), "values": vals}],
                })
        elif op == "compare_series":
            points = result.get("points") or []
            a_label = result.get("a_label", "A")
            b_label = result.get("b_label", "B")
            if len(points) >= 2:
                charts.append({
                    "chart_type": "line",
                    "title": f"{title} — {a_label} vs {b_label}",
                    "categories": [p.get("key") for p in points],
                    "series": [
                        {"name": a_label, "values": [p.get(a_label) for p in points]},
                        {"name": b_label, "values": [p.get(b_label) for p in points]},
                    ],
                })
    return charts


def charts_from_analysis(analysis, out_path, *, ops=None, max_groups: int = 12,
                         workbook_title: str | None = None, theme: Any = None) -> Path:
    """Convenience: ``suggest_charts_from_analysis`` → ``write_charts_xlsx`` (theme-aware)."""
    charts = suggest_charts_from_analysis(analysis, ops=ops, max_groups=max_groups)
    if not charts:
        raise ValueError("no chartable operations found in analysis payload")
    title = workbook_title
    if title is None and isinstance(analysis, dict):
        title = analysis.get("source") or "Analysis charts"
    return write_charts_xlsx(out_path, charts, workbook_title=str(title) if title else None,
                             theme=theme)


if __name__ == "__main__":
    import sys
    out = sys.argv[1] if len(sys.argv) > 1 else "workbook-selftest.xlsx"
    path = write_charts_xlsx(out, [
        {"chart_type": "column", "title": "Tasks by day",
         "categories": ["Mon", "Tue", "Wed", "Thu", "Fri"],
         "series": [{"name": "Done", "values": [4, 7, 5, 6, 3]}]},
        {"chart_type": "pie", "title": "Open by function",
         "categories": ["Compliance", "Finance", "Marketing", "Legal"],
         "series": [{"name": "Open", "values": [8, 5, 3, 4]}]},
        {"chart_type": "line", "title": "Opened vs closed",
         "categories": ["W1", "W2", "W3", "W4"],
         "series": [
             {"name": "Opened", "values": [10, 14, 9, 12]},
             {"name": "Closed", "values": [8, 11, 13, 10]},
         ]},
        {"chart_type": "waterfall", "title": "Open → close bridge",
         "steps": [
             {"label": "Open", "value": 12, "kind": "start"},
             {"label": "Done", "value": -5, "kind": "delta"},
             {"label": "New", "value": 4, "kind": "delta"},
             {"label": "Close", "value": 11, "kind": "total"},
         ]},
    ], workbook_title="Workbook self-test")
    print(f"[self-test] wrote {path} ({path.stat().st_size:,} bytes)")
