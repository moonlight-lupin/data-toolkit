"""convert.py — data-convert: interoperability, not cleaning.

Map data onto a TARGET system's contract (system A's export -> system B's import) and
RESHAPE its structure (long<->wide, nest<->flat, split, union). The counterpart to
data-tidy: **tidy** makes a messy source clean (quality); **convert** takes a clean-enough
source and re-expresses it in a *different* structure/format a downstream system requires.

Design notes:
- **Deterministic engine.** The model designs the mapping (intent-first, from the target
  contract + purpose); this engine APPLIES it. It never invents a value and never fetches:
  live inputs (e.g. an FX rate) are **pinned** into the spec by the user and applied as a
  recorded constant, so a re-run is reproducible.
- **The reusable artefact is a declarative spec** — a Markdown *conversion card* with an
  embedded ```convert-spec fenced JSON block (the machine source-of-truth). No per-conversion
  .py runner: a future agent reads the card, SENSE-CHECKS today's source against it, and
  re-runs this engine. Human reads the prose + mapping table; engine reads the block.
- **Cleaning is delegated to data-tidy** — if the source is messy, tidy it first, then convert
  the clean output (or pin a `tidy_recipe` in the spec's source block).

Run `python convert.py --self-test` (offline) for a worked check.
"""
from __future__ import annotations

import argparse
import csv as _csv
import json
import re
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path

# --------------------------------------------------------------------------- #
# Shared engine (dataclean / ingest) — loaded lazily so the pure reshape/map
# helpers and the self-test run offline. Coercion uses dataclean when present
# (so convert's dates/amounts match the rest of the toolkit), else a fallback.
# --------------------------------------------------------------------------- #
def _load_engine():
    cands = [Path(__file__).resolve().parents[3] / "scripts",   # toolkit-root engine
             Path(__file__).resolve().parent / "scripts"]       # vendored sibling
    for p in cands:
        if (p / "dataclean.py").is_file() and str(p) not in sys.path:
            sys.path.insert(0, str(p))
            break
    try:
        import dataclean, ingest  # noqa: F401
        return dataclean, ingest
    except Exception:  # noqa: BLE001
        return None, None


def _s(v) -> str:
    return "" if v is None else str(v).strip()


def _num(v):
    """Best-effort Decimal (exact, for finance). Uses dataclean when available."""
    if v is None or v == "":
        return None
    if isinstance(v, Decimal):
        return v
    dc, _ = _load_engine()
    if dc is not None:
        n, _note = dc.parse_number(v)
        return n
    try:
        return Decimal(str(v).replace(",", ""))
    except InvalidOperation:
        return None


def _fmt_number(v, spec):
    if v is None:
        return ""
    dp = spec.get("dp")
    if dp is not None:
        v = Decimal(str(v)).quantize(Decimal(1).scaleb(-int(dp)))
    return str(v)


def _fmt_date(v, spec):
    dc, _ = _load_engine()
    out_fmt = spec.get("format", "%d %b %Y")
    if dc is not None:
        d, _note = dc.parse_date(v, dayfirst=spec.get("dayfirst", True))
        return d.strftime(out_fmt) if d is not None else _s(v)
    return _s(v)


# --------------------------------------------------------------------------- #
# Stage 1 — reshape (structure change), pure functions on (header, rows).
# rows are list[dict]; header keeps column order.
# --------------------------------------------------------------------------- #
def unpivot(header, rows, id_cols, value_cols=None, var_name="variable", value_name="value"):
    """Wide -> long. Each `value_cols` cell becomes a row (id_cols + var_name/value_name)."""
    value_cols = value_cols or [c for c in header if c not in id_cols]
    out = []
    for r in rows:
        for vc in value_cols:
            row = {c: r.get(c) for c in id_cols}
            row[var_name] = vc
            row[value_name] = r.get(vc)
            out.append(row)
    return list(id_cols) + [var_name, value_name], out


def pivot(header, rows, index_cols, var_col, value_col, agg="first"):
    """Long -> wide. Distinct `var_col` values become columns carrying `value_col`."""
    variants, groups = [], {}
    for r in rows:
        key = tuple(_s(r.get(c)) for c in index_cols)
        var = _s(r.get(var_col))
        if var not in variants:
            variants.append(var)
        cell = groups.setdefault(key, {c: r.get(c) for c in index_cols})
        if agg == "sum":
            cell[var] = (_num(cell.get(var)) or Decimal(0)) + (_num(r.get(value_col)) or Decimal(0))
        elif var not in cell:                       # "first"
            cell[var] = r.get(value_col)
    variants.sort()
    out_header = list(index_cols) + variants
    return out_header, [{c: g.get(c, "") for c in out_header} for g in groups.values()]


def flatten(records, sep="."):
    """Nested dicts (parsed JSON) -> flat (header, rows). Lists index as key[0], key[1]…"""
    def walk(obj, prefix, into):
        if isinstance(obj, dict):
            for k, v in obj.items():
                walk(v, f"{prefix}{sep}{k}" if prefix else str(k), into)
        elif isinstance(obj, list):
            for i, v in enumerate(obj):
                walk(v, f"{prefix}{sep}{i}", into)
        else:
            into[prefix] = obj
    flat, header = [], []
    for rec in records:
        d = {}
        walk(rec, "", d)
        flat.append(d)
        for k in d:
            if k not in header:
                header.append(k)
    return header, [{c: r.get(c, "") for c in header} for r in flat]


def nest(header, rows, key_cols, into, child_cols):
    """Flat table -> nested records: group by key_cols, gather child_cols under `into`."""
    order, groups = [], {}
    for r in rows:
        key = tuple(_s(r.get(c)) for c in key_cols)
        if key not in groups:
            groups[key] = {c: r.get(c) for c in key_cols}
            groups[key][into] = []
            order.append(key)
        groups[key][into].append({c: r.get(c) for c in child_cols})
    return [groups[k] for k in order]


def split(header, rows, by):
    """One table -> {key: (header, rows)} partitioned on the value of column `by`."""
    parts = {}
    for r in rows:
        parts.setdefault(_s(r.get(by)), []).append(r)
    return {k: (header, rs) for k, rs in parts.items()}


def union(tables, how="outer"):
    """Many (header, rows) -> one, aligning columns. outer = union of headers; inner = shared."""
    headers = [h for h, _ in tables]
    if how == "inner":
        cols = [c for c in headers[0] if all(c in h for h in headers)]
    else:
        cols = []
        for h in headers:
            for c in h:
                if c not in cols:
                    cols.append(c)
    out = []
    for h, rs in tables:
        for r in rs:
            out.append({c: r.get(c, "") for c in cols})
    return cols, out


_RESHAPE = {"unpivot": unpivot, "pivot": pivot, "flatten": flatten, "nest": nest}


def apply_reshape(header, rows, ops):
    """Run a list of row-reshape ops in order. `split`/`nest` are terminal-ish and handled by
    the caller; here we cover the linear ones (unpivot/pivot/flatten)."""
    for op in ops or []:
        name = op["op"]
        kw = {k: v for k, v in op.items() if k != "op"}
        if name == "unpivot":
            header, rows = unpivot(header, rows, **kw)
        elif name == "pivot":
            header, rows = pivot(header, rows, **kw)
        elif name == "flatten":
            header, rows = flatten(rows, **kw)
        else:
            raise ValueError(f"reshape op '{name}' is applied by convert_file(), not apply_reshape()")
    return header, rows


# --------------------------------------------------------------------------- #
# Stage 1.5 — row filter (drop rows before mapping, e.g. only Posted entries).
# --------------------------------------------------------------------------- #
def _match(row, cond):
    v = row.get(cond["col"])
    op = cond.get("op", "eq")
    val = cond.get("value")
    sv = _s(v)
    if op == "in":
        return sv in [_s(x) for x in val]
    if op == "not_in":
        return sv not in [_s(x) for x in val]
    if op == "eq":
        return sv == _s(val)
    if op == "ne":
        return sv != _s(val)
    if op == "contains":
        return _s(val).lower() in sv.lower()
    if op == "blank":
        return sv == ""
    if op == "nonblank":
        return sv != ""
    n, m = _num(v), _num(val)                        # numeric comparisons
    if n is None or m is None:
        return False
    return {"gt": n > m, "ge": n >= m, "lt": n < m, "le": n <= m}.get(op, True)


def apply_filter(rows, conditions):
    """Keep only rows where ALL conditions match. `conditions` = [{col, op, value}]."""
    if not conditions:
        return rows
    return [r for r in rows if all(_match(r, c) for c in conditions)]


# --------------------------------------------------------------------------- #
# Lookup / enrich — translate a source value via a reference table (2nd input)
# or an inline map (e.g. internal account code -> target chart-of-accounts code).
# --------------------------------------------------------------------------- #
def _load_lookup_table(path, key, value):
    _, ingest = _load_engine()
    _h, rows = _records(ingest.read_any(path))
    return {_s(r.get(key)): r.get(value) for r in rows}


def load_lookups(mapping, base_dir=None):
    """Preload every `compute:"lookup"` field's table (once, not per row) -> {target_col: {k: v}}.
    A field carries either an inline `map_values` dict or a `table` file (+ `key`/`value` cols)."""
    out = {}
    for tcol, sp in (mapping or {}).items():
        if sp.get("compute") != "lookup":
            continue
        if "map_values" in sp:
            out[tcol] = {_s(k): v for k, v in sp["map_values"].items()}
        elif sp.get("table"):
            path = str(Path(base_dir) / sp["table"]) if base_dir else sp["table"]
            out[tcol] = _load_lookup_table(path, sp["key"], sp["value"])
    return out


def _fx_rate(fx, row, spec):
    """The rate to apply: a single pinned `fx.rate`, or the row-date-effective rate from a pinned
    `fx.rates` table (latest as_of <= the row's `on_date`). The engine never fetches."""
    if not fx:
        return None
    if fx.get("rate") not in (None, ""):
        return Decimal(str(fx["rate"]))
    rates = fx.get("rates")
    if rates:
        on = spec.get("on_date")
        d = _fmt_date(row.get(on), {"format": "%Y-%m-%d"}) if on else None
        eligible = [r for r in rates if (not d or _s(r.get("as_of")) <= d)]
        chosen = (max(eligible, key=lambda r: _s(r.get("as_of")), default=None)
                  or min(rates, key=lambda r: _s(r.get("as_of")), default=None))
        return Decimal(str(chosen["rate"])) if chosen else None
    return None


# --------------------------------------------------------------------------- #
# Stage 2 — map onto the target contract.
# mapping: {target_col: {from, compute?, type?, format?, dp?, const?, sep?, ...}}
# --------------------------------------------------------------------------- #
def _field(row, spec, fx=None, lookup=None):
    comp = spec.get("compute", "as_is")
    frm = spec.get("from")
    if "const" in spec:
        raw = spec["const"]
    elif comp == "debit_minus_credit":
        d, c = frm
        return _fmt_number((_num(row.get(d)) or Decimal(0)) - (_num(row.get(c)) or Decimal(0)), spec)
    elif comp == "sum":
        return _fmt_number(sum((_num(row.get(x)) or Decimal(0)) for x in frm), spec)
    elif comp == "concat":
        return spec.get("sep", " ").join(_s(row.get(x)) for x in frm)
    elif comp == "fx_convert":
        rate = _fx_rate(fx, row, spec)
        if rate is None:
            raise ValueError("compute 'fx_convert' needs a pinned fx.rate (or fx.rates + on_date)")
        n = _num(row.get(frm))
        return _fmt_number((n * rate) if n is not None else None, spec)
    elif comp == "lookup":
        k = _s(row.get(frm))
        if lookup is not None and k in lookup:
            raw = lookup[k]
        else:
            om = spec.get("on_missing", "keep")      # keep (source value passes through) | blank | error
            if om == "error":
                raise ValueError(f"lookup: no match for {k!r} (target left untranslated)")
            raw = "" if om == "blank" else row.get(frm)
    else:                                            # as_is
        raw = row.get(frm if isinstance(frm, str) else (frm[0] if frm else ""))
    typ = spec.get("type")
    if typ == "date":
        return _fmt_date(raw, spec)
    if typ in ("number", "currency"):
        return _fmt_number(_num(raw), spec)
    return _s(raw)


def apply_map(header, rows, mapping, fx=None, lookups=None):
    """Produce (target_header, target_rows) by applying `mapping` to each source row."""
    lookups = lookups or {}
    tcols = list(mapping.keys())
    out = [{tc: _field(r, sp, fx, lookups.get(tc)) for tc, sp in mapping.items()} for r in rows]
    return tcols, out


def check_contract(target_header, target_rows, columns, source_header, mapping, rules=None):
    """Report contract issues: required target fields never populated, and source columns that
    the mapping never consumed (so nothing is silently dropped)."""
    rules = rules or {}
    issues = []
    required = [c["name"] for c in (columns or []) if c.get("required")]
    for req in required:
        if req not in target_header:
            issues.append({"kind": "missing_required", "detail": f"target field '{req}' is not mapped",
                           "severity": "error"})
        elif all(_s(r.get(req)) == "" for r in target_rows):
            issues.append({"kind": "required_empty", "detail": f"required field '{req}' is empty on every row",
                           "severity": "error"})
    consumed = set()
    for sp in mapping.values():
        frm = sp.get("from")
        consumed.update([frm] if isinstance(frm, str) else (frm or []))
    for c in source_header:
        if c not in consumed:
            issues.append({"kind": "unmapped_source", "detail": f"source column '{c}' is not used",
                           "severity": "info" if rules.get("on_unmapped_source", "report") != "error" else "error"})
    issues.extend(_validate(target_rows, columns))
    return issues


def _valid_iban(v):
    s = re.sub(r"\s", "", _s(v)).upper()
    if not re.fullmatch(r"[A-Z]{2}\d{2}[A-Z0-9]{1,30}", s):
        return False
    rearranged = s[4:] + s[:4]                       # move the country+check to the back
    digits = "".join(str(int(ch, 36)) for ch in rearranged)  # A..Z -> 10..35
    return int(digits) % 97 == 1


def _valid_bic(v):
    return bool(re.fullmatch(r"[A-Z]{6}[A-Z0-9]{2}([A-Z0-9]{3})?", _s(v).upper()))


_CHECKERS = {"iban": _valid_iban, "bic": _valid_bic}


def _validate(target_rows, columns):
    """Validate output values against each target column's contract: `allowed` (value set),
    `pattern` (regex), `max_len`, and `check` (iban/bic). Aggregated to one issue per
    (column, rule) with a count + first offending value — so a bad column doesn't flood."""
    issues, seen = [], {}
    for c in (columns or []):
        name = c["name"]
        for r in target_rows:
            v = _s(r.get(name))
            if v == "":
                continue
            fails = []
            if c.get("allowed") is not None and v not in [_s(x) for x in c["allowed"]]:
                fails.append(("not_allowed", "not in the allowed set"))
            if c.get("pattern") and not re.search(c["pattern"], v):
                fails.append(("pattern", f"fails pattern {c['pattern']}"))
            if c.get("max_len") and len(v) > int(c["max_len"]):
                fails.append(("too_long", f"exceeds max_len {c['max_len']}"))
            chk = c.get("check")
            if chk in _CHECKERS and not _CHECKERS[chk](v):
                fails.append((f"bad_{chk}", f"is not a valid {chk.upper()}"))
            for kind, why in fails:
                key = (name, kind)
                if key not in seen:
                    seen[key] = {"kind": kind, "severity": "error", "n": 0,
                                 "detail": f"'{name}' {why} (e.g. '{v}')"}
                    issues.append(seen[key])
                seen[key]["n"] += 1
    for i in issues:
        if i.get("n", 0) > 1:
            i["detail"] += f" — {i['n']} rows"
        i.pop("n", None)
    return issues


# --------------------------------------------------------------------------- #
# Stage 3 — write the target format.
# --------------------------------------------------------------------------- #
def _rows_as_lists(header, rows):
    return [[r.get(c, "") for c in header] for r in rows]


def write_csv(header, rows, path, delimiter=",", encoding="utf-8"):
    with open(path, "w", newline="", encoding=encoding) as f:
        w = _csv.writer(f, delimiter=delimiter)
        w.writerow(header)
        w.writerows(_rows_as_lists(header, rows))
    return path


def write_json(rows, path, indent=2):
    def _default(o):
        return str(o) if isinstance(o, Decimal) else o
    Path(path).write_text(json.dumps(rows, indent=indent, default=_default, ensure_ascii=False),
                          encoding="utf-8")
    return path


def write_xlsx(header, rows, path):
    dc, _ = _load_engine()
    if dc is None:
        raise RuntimeError("xlsx output needs the shared dataclean engine (openpyxl)")
    dc.write_xlsx(header, _rows_as_lists(header, rows), path)
    return path


def write_fixedwidth(header, rows, path, columns, write_header=False, encoding="utf-8"):
    """Fixed-width flat file (legacy bank/ERP imports). Each target column carries `width`,
    `align` (`l`/`r`, default left) and `pad` (default space); values are clipped to width."""
    specs = {c["name"]: c for c in (columns or [])}

    def fld(name, val):
        c = specs.get(name, {})
        w = int(c.get("width", len(_s(val))))
        pad = (c.get("pad") or " ")[:1]
        s = _s(val)[:w]
        return s.rjust(w, pad) if c.get("align") == "r" else s.ljust(w, pad)

    lines = []
    if write_header:
        lines.append("".join(fld(c, c) for c in header))
    lines += ["".join(fld(c, r.get(c, "")) for c in header) for r in rows]
    Path(path).write_text("\n".join(lines) + ("\n" if lines else ""), encoding=encoding)
    return path


def write_template(target_header, rows, template_path, out_path):
    """Populate a PROVIDED target template (its first row is the header). The template's own
    column order/format is preserved; mapped rows are written under it. XLSX keeps the template
    workbook (header styling, other sheets); CSV takes the template's header line."""
    ext = Path(template_path).suffix.lower()
    if ext == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(template_path)
        ws = wb.active
        thdr = [_s(c.value) for c in ws[1]]
        start = ws.max_row + 1
        for i, r in enumerate(rows):
            for j, col in enumerate(thdr, start=1):
                ws.cell(row=start + i, column=j, value=r.get(col, ""))
        wb.save(out_path)
        return out_path
    with open(template_path, encoding="utf-8") as f:
        thdr = next(_csv.reader(f))
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = _csv.writer(f)
        w.writerow(thdr)
        w.writerows([[r.get(c, "") for c in thdr] for r in rows])
    return out_path


def write_output(target_header, target_rows, target_spec, out_path, base_dir=None):
    ts = target_spec or {}
    if ts.get("template"):
        tpath = str(Path(base_dir) / ts["template"]) if base_dir else ts["template"]
        return write_template(target_header, target_rows, tpath, out_path)
    fmt = ts.get("format", "csv")
    if fmt == "csv":
        return write_csv(target_header, target_rows, out_path,
                         delimiter=ts.get("delimiter", ","), encoding=ts.get("encoding", "utf-8"))
    if fmt in ("json", "jsonl"):
        return write_json(target_rows, out_path)
    if fmt == "xlsx":
        return write_xlsx(target_header, target_rows, out_path)
    if fmt == "fixedwidth":
        return write_fixedwidth(target_header, target_rows, out_path, ts.get("columns"),
                                write_header=ts.get("header", False))
    raise ValueError(f"unsupported target format '{fmt}'")


# --------------------------------------------------------------------------- #
# Sense-check — verify today's source against the card BEFORE applying.
# --------------------------------------------------------------------------- #
def fingerprint(header, rows):
    """A compact 'expected source' signature stored in the card for drift detection."""
    return {"columns": list(header), "rows_magnitude": _magnitude(len(rows))}


def _magnitude(n):
    if n < 100:
        return "~10s"
    if n < 1000:
        return "~100s"
    if n < 10000:
        return "~1000s"
    return "~10000s+"


def sense_check(spec, header, rows, run_dates_after=None):
    """Compare a live source against the card's expectations. Returns a list of discrepancies —
    the caller FLAGS these to the user and does NOT auto-apply over drift."""
    issues = []
    src = spec.get("source", {})
    expected = src.get("expected_columns") or (src.get("fingerprint") or {}).get("columns") or []
    mapping = spec.get("map", {})
    mapped = set()
    for sp in mapping.values():
        frm = sp.get("from")
        mapped.update([frm] if isinstance(frm, str) else (frm or []))
    for col in mapped:
        if col not in header:
            issues.append({"kind": "mapped_column_missing", "severity": "error",
                           "detail": f"mapping needs source column '{col}', which is absent"})
    for col in header:
        if expected and col not in expected:
            issues.append({"kind": "new_column", "severity": "warn",
                           "detail": f"source has a new column '{col}' not in the card"})
    for col in expected:
        if col not in header:
            issues.append({"kind": "expected_column_gone", "severity": "warn",
                           "detail": f"card expects '{col}', absent from this source"})
    fx = spec.get("fx")
    if fx and fx.get("as_of") and run_dates_after and fx["as_of"] < run_dates_after:
        issues.append({"kind": "stale_fx", "severity": "warn",
                       "detail": f"pinned {fx.get('pair','rate')} is as of {fx['as_of']} — "
                                 f"older than this run; refresh if the rate should be current"})
    return issues


# --------------------------------------------------------------------------- #
# The reusable artefact: a Markdown conversion CARD with an embedded JSON spec.
# --------------------------------------------------------------------------- #
_SPEC_FENCE = re.compile(r"```(?:convert-spec|json)\s*\n(.*?)\n```", re.DOTALL)


def load_spec(path):
    """Load a spec from a .json file or from the ```convert-spec block of a .md card."""
    text = Path(path).read_text(encoding="utf-8")
    if str(path).lower().endswith(".json"):
        return json.loads(text)
    m = _SPEC_FENCE.search(text)
    if not m:
        raise ValueError(f"no ```convert-spec (or ```json) block found in {path}")
    return json.loads(m.group(1))


def _md_table(headers, rows):
    L = ["| " + " | ".join(headers) + " |", "|" + "|".join(["---"] * len(headers)) + "|"]
    for r in rows:
        L.append("| " + " | ".join(str(c).replace("|", "\\|") for c in r) + " |")
    return "\n".join(L)


def render_card(spec):
    """Generate the human+machine conversion card (Markdown prose + mapping table + the
    ```convert-spec JSON block that the engine executes)."""
    name = spec.get("name", "Conversion")
    src, tgt = spec.get("source", {}), spec.get("target", {})
    L = [f"# Convert: {name}", ""]
    if spec.get("purpose"):
        L += [spec["purpose"], ""]
    L += [f"- **Source:** {src.get('format','?')}" + (f" (sheet `{src['sheet']}`)" if src.get("sheet") else ""),
          f"- **Target:** {tgt.get('format','?')}"
          + (f" — contract `{tgt['contract']}`" if tgt.get("contract") else ""), ""]
    fx = spec.get("fx")
    if fx:
        L += [f"> **Pinned rate:** {fx.get('pair','?')} = {fx.get('rate','?')} "
              f"(as of {fx.get('as_of','?')}, source: {fx.get('source','user-supplied')}). "
              f"Applied as a constant — refresh deliberately.", ""]
    L += ["## Mapping (target ← source)", ""]
    rows = []
    for tcol, sp in spec.get("map", {}).items():
        frm = sp.get("const", sp.get("from", ""))
        frm = ", ".join(frm) if isinstance(frm, list) else str(frm)
        rule = sp.get("compute", "as_is")
        if sp.get("type"):
            rule += f" → {sp['type']}"
        rows.append([tcol, frm, rule])
    L += [_md_table(["Target column", "Source", "Rule"], rows), ""]
    exp = src.get("expected_columns") or (src.get("fingerprint") or {}).get("columns")
    if exp:
        L += ["## Expected source (verify before running)",
              "Columns: " + ", ".join(f"`{c}`" for c in exp), "",
              "> Before applying, sense-check today's source against this. Flag any missing, "
              "renamed or new columns to the user; don't blind-apply over drift.", ""]
    L += ["## Spec (machine source of truth)", "", "```convert-spec",
          json.dumps(spec, indent=2, ensure_ascii=False), "```", ""]
    L += ["> Deterministic conversion — a draft for a qualified person to review. The engine "
          "applies this spec; it never invents a value and never fetches (pinned inputs only)."]
    return "\n".join(L)


# --------------------------------------------------------------------------- #
# Orchestration
# --------------------------------------------------------------------------- #
def convert_rows(spec, header, rows, lookups=None):
    """Pure core: filter -> reshape -> map (+ lookups) -> contract-check + validate. Returns
    (target_header, target_rows, report). `split`/`nest` are applied by convert_file since they
    change the output cardinality. `lookups` may be preloaded by convert_file; else inline
    `map_values` are resolved here."""
    n_in = len(rows)
    rows = apply_filter(rows, spec.get("filter"))
    reshape = [o for o in spec.get("reshape", []) if o["op"] in ("unpivot", "pivot", "flatten")]
    header, rows = apply_reshape(header, rows, reshape)
    mapping, tgt = spec.get("map", {}), spec.get("target", {})
    if lookups is None:
        lookups = load_lookups(mapping)              # inline map_values (file tables need a base_dir)
    if mapping:
        t_header, t_rows = apply_map(header, rows, mapping, fx=spec.get("fx"), lookups=lookups)
    else:
        t_header, t_rows = header, rows
    issues = check_contract(t_header, t_rows, tgt.get("columns"), header, mapping,
                            spec.get("rules"))
    report = {"rows_in": n_in, "rows_filtered": n_in - len(rows), "rows_out": len(t_rows),
              "issues": issues, "target_columns": t_header}
    return t_header, t_rows, report


def _records(raw):
    """ingest.read_any -> list[dict] using the detected header (mirrors reconcile._records)."""
    dc, _ = _load_engine()
    rows = raw[0] if isinstance(raw, tuple) else raw
    rows = [r for r in rows if any(_s(c) for c in r)]
    if not rows:
        return [], []
    hr = dc.detect_header(rows) if dc and hasattr(dc, "detect_header") else 0
    header = [_s(h) for h in rows[hr]]
    out = []
    for r in rows[hr + 1:]:
        r = list(r) + [None] * (len(header) - len(r))
        out.append({header[i]: r[i] for i in range(len(header))})
    return header, out


def convert_file(spec, in_path, out_path, run_dates_after=None, base_dir=None):
    """Read source (any format via ingest), sense-check, convert, write. Returns a report dict
    (with `sense_check` issues surfaced separately so the agent can halt on drift). `base_dir`
    (defaults to the card's folder) resolves relative lookup tables and output templates."""
    _, ingest = _load_engine()
    if spec.get("source", {}).get("clean_with_tidy") and spec["source"].get("tidy_recipe"):
        dc, _ = _load_engine()
        raw = ingest.read_any(in_path, sheet=spec["source"].get("sheet"))
        rows = raw[0] if isinstance(raw, tuple) else raw
        header, body, _log = dc.apply_recipe(rows, spec["source"]["tidy_recipe"])
        rows = [dict(zip(header, r)) for r in body]
    else:
        header, rows = _records(ingest.read_any(in_path, sheet=spec.get("source", {}).get("sheet")))

    drift = sense_check(spec, header, rows, run_dates_after=run_dates_after)
    lookups = load_lookups(spec.get("map", {}), base_dir=base_dir)   # resolve file tables here
    t_header, t_rows, report = convert_rows(spec, header, rows, lookups=lookups)
    report["sense_check"] = drift

    split_op = next((o for o in spec.get("reshape", []) if o["op"] == "split"), None)
    nest_op = next((o for o in spec.get("reshape", []) if o["op"] == "nest"), None)
    outp = Path(out_path)
    if nest_op:
        kw = {k: v for k, v in nest_op.items() if k != "op"}
        nested = nest(t_header, t_rows, **kw)
        out_path_json = outp.with_suffix(".json")
        outp.write_text(json.dumps(nested, indent=2, default=str), encoding="utf-8")
        report["written"] = [str(out_path_json)]
    elif split_op:
        parts = split(t_header, t_rows, split_op["by"])
        written = []
        for key, (h, rs) in parts.items():
            safe = re.sub(r"[^\w.-]+", "_", key) or "blank"
            p = outp.with_name(f"{outp.stem}_{safe}{outp.suffix}")
            written.append(write_output(h, rs, spec.get("target"), p, base_dir=base_dir))
        report["written"] = [str(p) for p in written]
    else:
        report["written"] = [str(write_output(t_header, t_rows, spec.get("target"), out_path,
                                               base_dir=base_dir))]
    return report


def render_report(report, *, name="Conversion"):
    L = [f"# {name} — conversion report", ""]
    filt = report.get("rows_filtered") or 0
    L.append(f"- Rows in: {report['rows_in']}"
             + (f" · filtered out: {filt}" if filt else "")
             + f" · rows out: {report['rows_out']}")
    L.append(f"- Written: {', '.join(report.get('written', [])) or '(not written)'}")
    drift = report.get("sense_check") or []
    if drift:
        L += ["", "## ⚠️ Source sense-check — confirm before relying on this run"]
        for d in drift:
            L.append(f"- **{d['severity']}** — {d['detail']}")
    issues = report.get("issues") or []
    if issues:
        L += ["", "## Contract issues"]
        for i in issues:
            L.append(f"- **{i['severity']}** — {i['detail']}")
    if not drift and not [i for i in issues if i["severity"] == "error"]:
        L += ["", "> Clean: source matched the card and every required target field is populated."]
    return "\n".join(L)


# --------------------------------------------------------------------------- #
# Self-test (offline)
# --------------------------------------------------------------------------- #
def _self_test():
    # reshape: unpivot (wide months -> long)
    wide_h = ["Entity", "Jan", "Feb"]
    wide = [{"Entity": "A", "Jan": "10", "Feb": "20"}, {"Entity": "B", "Jan": "5", "Feb": "7"}]
    lh, lr = unpivot(wide_h, wide, ["Entity"], var_name="Month", value_name="Amount")
    assert lh == ["Entity", "Month", "Amount"] and len(lr) == 4, (lh, lr)
    # pivot back
    ph, pr = pivot(lh, lr, ["Entity"], "Month", "Amount")
    assert ph == ["Entity", "Feb", "Jan"] and len(pr) == 2, (ph, pr)
    print("reshape unpivot/pivot: PASS")

    # split + union round-trip
    parts = split(wide_h, wide, "Entity")
    assert set(parts) == {"A", "B"}
    uh, ur = union([parts["A"], parts["B"]])
    assert len(ur) == 2 and uh == wide_h, (uh, ur)
    print("reshape split/union: PASS")

    # flatten nested JSON, then nest back
    nested = [{"id": 1, "party": {"name": "Acme"}, "lines": [{"amt": 10}, {"amt": 20}]}]
    fh, fr = flatten(nested)
    assert "party.name" in fh and fr[0]["party.name"] == "Acme", (fh, fr)
    flat_rows = [{"id": 1, "sku": "X", "amt": 10}, {"id": 1, "sku": "Y", "amt": 20}]
    nrec = nest(["id", "sku", "amt"], flat_rows, ["id"], "lines", ["sku", "amt"])
    assert len(nrec) == 1 and len(nrec[0]["lines"]) == 2, nrec
    print("reshape flatten/nest: PASS")

    # contract mapping: GL (Debit/Credit, signed) -> journal import (Date/Amount/Description)
    gl_h = ["Date", "Account", "Debit", "Credit", "Memo"]
    gl = [{"Date": "13/06/2026", "Account": "6000", "Debit": "1,200.00", "Credit": "", "Memo": "Rent"},
          {"Date": "14/06/2026", "Account": "1000", "Debit": "", "Credit": "1200", "Memo": "Rent pmt"}]
    spec = {
        "name": "GL -> journal import",
        "source": {"format": "csv", "expected_columns": gl_h},
        "target": {"format": "csv", "contract": "journal_import",
                   "columns": [{"name": "JournalDate", "required": True},
                               {"name": "Amount", "required": True}, {"name": "Narration"}]},
        "map": {
            "JournalDate": {"from": "Date", "type": "date", "format": "%Y-%m-%d"},
            "Amount": {"from": ["Debit", "Credit"], "compute": "debit_minus_credit", "dp": 2},
            "Narration": {"from": "Memo", "type": "text"},
        },
        "rules": {"on_unmapped_source": "report"},
    }
    th, tr, rep = convert_rows(spec, gl_h, gl)
    assert th == ["JournalDate", "Amount", "Narration"], th
    assert tr[0]["JournalDate"] == "2026-06-13" and tr[0]["Amount"] == "1200.00", tr[0]
    assert tr[1]["Amount"] == "-1200.00", tr[1]                       # credit -> negative
    # 'Account' was not mapped -> reported (not silently dropped)
    assert any(i["kind"] == "unmapped_source" and "Account" in i["detail"] for i in rep["issues"]), rep
    print("contract mapping + debit_minus_credit: PASS")

    # required target field empty -> error
    spec2 = json.loads(json.dumps(spec))
    spec2["map"]["Amount"] = {"const": "", "compute": "as_is"}
    _, _, rep2 = convert_rows(spec2, gl_h, gl)
    assert any(i["kind"] in ("required_empty",) for i in rep2["issues"]), rep2
    print("contract required-field check: PASS")

    # pinned FX applied as a recorded constant (engine never fetches)
    fx_spec = {"map": {"GBP": {"from": "USD", "compute": "fx_convert", "dp": 2}},
               "fx": {"pair": "USD/GBP", "rate": "0.80", "as_of": "2026-07-15"}, "target": {}}
    _, fr2, _ = convert_rows(fx_spec, ["USD"], [{"USD": "100"}])
    assert fr2[0]["GBP"] == "80.00", fr2
    try:
        convert_rows({"map": {"GBP": {"from": "USD", "compute": "fx_convert"}}, "target": {}},
                     ["USD"], [{"USD": "100"}])
        raise AssertionError("fx_convert without a pinned rate must error")
    except ValueError:
        pass
    print("pinned FX: PASS")

    # card round-trip: render -> load -> same spec
    card = render_card(spec)
    assert "```convert-spec" in card and "## Mapping" in card
    import tempfile, os
    fd, p = tempfile.mkstemp(suffix=".md")
    os.close(fd)
    Path(p).write_text(card, encoding="utf-8")
    assert load_spec(p)["map"]["Amount"]["compute"] == "debit_minus_credit"
    os.unlink(p)
    print("card render/load round-trip: PASS")

    # sense-check flags drift: a renamed source column + a stale pinned rate
    drift_spec = {"source": {"expected_columns": ["Date", "Amount"]},
                  "map": {"D": {"from": "Date"}, "A": {"from": "Amount"}},
                  "fx": {"pair": "USD/GBP", "rate": "0.8", "as_of": "2026-06-01"}}
    issues = sense_check(drift_spec, ["Date", "Amt"], [{"Date": "x", "Amt": "1"}],
                         run_dates_after="2026-07-01")
    kinds = {i["kind"] for i in issues}
    assert "mapped_column_missing" in kinds and "new_column" in kinds and "stale_fx" in kinds, issues
    print("sense-check drift: PASS")

    # ------------------------------ v2 ------------------------------ #
    # v2 (1) row filter: keep only Posted, non-zero rows before mapping
    f_rows = [{"Status": "Posted", "Amt": "10"}, {"Status": "Draft", "Amt": "5"},
              {"Status": "Posted", "Amt": "0"}]
    fspec = {"filter": [{"col": "Status", "op": "eq", "value": "Posted"},
                        {"col": "Amt", "op": "gt", "value": "0"}],
             "map": {"Amount": {"from": "Amt", "type": "number", "dp": 2}}, "target": {}}
    _, fr, frep = convert_rows(fspec, ["Status", "Amt"], f_rows)
    assert len(fr) == 1 and fr[0]["Amount"] == "10.00" and frep["rows_filtered"] == 2, (fr, frep)
    print("v2 row filter: PASS")

    # v2 (2) lookup / enrich: translate an internal code via an inline map
    lspec = {"map": {"TargetCode": {"from": "Acct", "compute": "lookup",
                                    "map_values": {"1000": "CASH", "6000": "RENT"},
                                    "on_missing": "keep"}}, "target": {}}
    _, lr2, _ = convert_rows(lspec, ["Acct"], [{"Acct": "1000"}, {"Acct": "9999"}])
    assert lr2[0]["TargetCode"] == "CASH" and lr2[1]["TargetCode"] == "9999", lr2  # miss -> kept
    print("v2 lookup/enrich: PASS")

    # v2 (3) output validation: allowed set, IBAN checksum, max_len
    vcols = [{"name": "Ccy", "allowed": ["GBP", "USD"]},
             {"name": "IBAN", "check": "iban"}, {"name": "Ref", "max_len": 6}]
    vrows = [{"Ccy": "GBP", "IBAN": "GB82WEST12345698765432", "Ref": "OK"},
             {"Ccy": "EUR", "IBAN": "GB00WEST12345698765432", "Ref": "TOOLONGREF"}]
    vissues = {i["kind"] for i in _validate(vrows, vcols)}
    assert {"not_allowed", "bad_iban", "too_long"} <= vissues, vissues
    assert _valid_iban("GB82 WEST 1234 5698 7654 32") and not _valid_iban("GB00WEST12345698765432")
    assert _valid_bic("NWBKGB2L") and _valid_bic("NWBKGB2LXXX") and not _valid_bic("BADBIC")
    print("v2 output validation (allowed/IBAN/BIC/max_len): PASS")

    # v2 (4) date-keyed FX: per-row rate by transaction date
    fxspec = {"map": {"GBP": {"from": "USD", "compute": "fx_convert", "on_date": "Date", "dp": 2}},
              "fx": {"pair": "USD/GBP", "rates": [{"as_of": "2026-01-01", "rate": "0.80"},
                                                  {"as_of": "2026-06-01", "rate": "0.75"}]},
              "target": {}}
    _, fxr, _ = convert_rows(fxspec, ["USD", "Date"],
                             [{"USD": "100", "Date": "15/03/2026"},   # -> 0.80
                              {"USD": "100", "Date": "20/06/2026"}])  # -> 0.75
    assert fxr[0]["GBP"] == "80.00" and fxr[1]["GBP"] == "75.00", fxr
    print("v2 date-keyed FX: PASS")

    # v2 (5) fixed-width output
    import tempfile as _tf, os as _os
    fd2, fw = _tf.mkstemp(suffix=".txt"); _os.close(fd2)
    write_fixedwidth(["Code", "Amount"], [{"Code": "AB", "Amount": "12.50"}], fw,
                     [{"name": "Code", "width": 4}, {"name": "Amount", "width": 8, "align": "r", "pad": "0"}])
    line = Path(fw).read_text(encoding="utf-8").splitlines()[0]
    assert line == "AB  00012.50", repr(line)
    _os.unlink(fw)
    print("v2 fixed-width output: PASS")

    # v2 (6) template population: fill a provided CSV template (its header/order wins)
    fd3, tpl = _tf.mkstemp(suffix=".csv"); _os.close(fd3)
    Path(tpl).write_text("Ref,Amount,Ccy\n", encoding="utf-8")
    fd4, tout = _tf.mkstemp(suffix=".csv"); _os.close(fd4)
    write_template(["Amount", "Ref"], [{"Amount": "9.99", "Ref": "X1", "Ccy": "GBP"}], tpl, tout)
    got = Path(tout).read_text(encoding="utf-8").splitlines()
    assert got[0] == "Ref,Amount,Ccy" and got[1] == "X1,9.99,GBP", got   # template order preserved
    _os.unlink(tpl); _os.unlink(tout)
    print("v2 template population: PASS")

    print("self-test: PASS")
    return 0


def main(argv=None):
    ap = argparse.ArgumentParser(description="Convert a source onto a target contract / reshape it.")
    ap.add_argument("source", nargs="?", help="source file (any format ingest reads)")
    ap.add_argument("--card", help="conversion card (.md) or spec (.json) to apply")
    ap.add_argument("-o", "--out", help="output path (a split writes <stem>_<key><suffix>)")
    ap.add_argument("--check-only", action="store_true",
                    help="sense-check the source against the card and report; write nothing")
    ap.add_argument("--as-of", help="this run's date, to flag a stale pinned rate")
    ap.add_argument("--self-test", action="store_true")
    a = ap.parse_args(argv)
    if a.self_test:
        return _self_test()
    if not (a.card and a.source):
        ap.error("need a --card and a source (or --self-test)")
    spec = load_spec(a.card)
    if a.check_only:
        _, ingest = _load_engine()
        header, rows = _records(ingest.read_any(a.source, sheet=spec.get("source", {}).get("sheet")))
        issues = sense_check(spec, header, rows, run_dates_after=a.as_of)
        print("Sense-check:", "clean" if not issues else "")
        for i in issues:
            print(f"  [{i['severity']}] {i['detail']}")
        return 0
    out = a.out or (Path(a.source).stem + "_converted." + spec.get("target", {}).get("format", "csv"))
    base_dir = str(Path(a.card).resolve().parent)   # resolve relative lookup tables / templates
    report = convert_file(spec, a.source, out, run_dates_after=a.as_of, base_dir=base_dir)
    print(render_report(report, name=spec.get("name", "Conversion")))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
