"""Large-file streaming helpers for Excel → Parquet with constant-memory paths.

Opt-in from skills via ``ingest.read_large`` when a source may have 10k+ rows.
openpyxl ``read_only`` iteration keeps the row-count probe and the stream path
off the full-workbook load. ``pyarrow`` (+ ``pandas``) are optional — without
them callers fall back to a direct openpyxl read with a warning.

Thresholds
----------
- ``direct``         < 10k rows
- ``parquet_cache``  10k–100k rows (read once, cache Parquet, reload from cache)
- ``stream``         100k+ rows (chunked openpyxl → ParquetWriter)

All Parquet columns are written as ``pa.string()`` so mixed Excel types cannot
produce cross-chunk schema mismatches; numeric conversion is deferred to
``pandas.to_numeric`` after load.
"""

from __future__ import annotations

import gc
import hashlib
import importlib.util
import sys
import warnings
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

DIRECT_THRESHOLD = 10_000
PARQUET_CACHE_THRESHOLD = 100_000
DEFAULT_CHUNK_SIZE = 50_000


def pyarrow_available() -> bool:
    """True when both pyarrow and pandas are importable (large-file path ready)."""
    return (
        importlib.util.find_spec("pyarrow") is not None
        and importlib.util.find_spec("pandas") is not None
    )


def count_rows(path: str, sheet=None) -> dict:
    """Count rows per sheet with constant memory (openpyxl read_only + iter_rows).

    Never uses ``pd.read_excel`` — that would load the workbook into memory.

    Returns
    -------
    dict
        ``{"sheets": {sheet_name: row_count, ...}, "total": int}``
        ``row_count`` includes the header row. Empty worksheets report 0.
    """
    from openpyxl import load_workbook

    path = str(path)
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        sheets: dict[str, int] = {}
        for ws in wb.worksheets:
            if sheet is not None and ws.title != sheet:
                continue
            n = 0
            for _ in ws.iter_rows(values_only=True):
                n += 1
            sheets[ws.title] = n
        if sheet is not None and sheet not in sheets:
            raise ValueError(
                f"Sheet {sheet!r} not found in {Path(path).name}. "
                f"Available: {', '.join(ws.title for ws in wb.worksheets) or '(none)'}"
            )
        return {"sheets": sheets, "total": sum(sheets.values())}
    finally:
        wb.close()


def choose_strategy(path: str, sheet=None) -> str:
    """Pick a read strategy from row count: ``direct`` | ``parquet_cache`` | ``stream``.

    When ``sheet`` is None the largest non-empty sheet's row count is used (multi-sheet
    workbooks still require an explicit sheet at read time via ingest).
    """
    counts = count_rows(path, sheet=sheet)
    if sheet is not None:
        n = counts["sheets"].get(sheet, 0)
    else:
        n = max(counts["sheets"].values()) if counts["sheets"] else 0
    # Strategy gates are on *data* scale; header-only sheets stay direct.
    data_rows = max(0, n - 1) if n else 0
    if data_rows < DIRECT_THRESHOLD:
        return "direct"
    if data_rows < PARQUET_CACHE_THRESHOLD:
        return "parquet_cache"
    return "stream"


def _resolve_sheet(path: str, sheet=None) -> str:
    """Mirror ingest.read_xlsx sheet selection without importing ingest (cycle-safe)."""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        names = [ws.title for ws in wb.worksheets]

        def _has_data(ws):
            for row in ws.iter_rows(values_only=True):
                if any(c is not None and str(c).strip() != "" for c in row):
                    return True
            return False

        if sheet is not None:
            if sheet not in names:
                raise ValueError(
                    f"Sheet {sheet!r} not found in {Path(path).name}. "
                    f"Available: {', '.join(names) or '(none)'}"
                )
            return sheet

        non_empty = []
        for ws in wb.worksheets:
            if ws.sheet_state != "visible":
                continue
            if _has_data(ws):
                non_empty.append(ws.title)
        if len(non_empty) == 1:
            return non_empty[0]
        if len(non_empty) > 1:
            # Defer the typed SheetSelectionRequired to the caller (ingest).
            raise ValueError(
                f"{Path(path).name} has {len(non_empty)} non-empty sheets: "
                f"{', '.join(non_empty)}. Specify sheet=."
            )
        return names[0] if names else "Sheet1"
    finally:
        wb.close()


def _cell_str(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, str):
        return v
    return str(v)


def stream_excel_to_parquet(
    path: str,
    parquet_path: str,
    sheet=None,
    chunk_size: int = DEFAULT_CHUNK_SIZE,
) -> int:
    """Stream one Excel sheet to Parquet in ``chunk_size`` row batches.

    Opens the workbook in ``read_only`` mode, reads the header, then writes each
    chunk via ``pq.ParquetWriter`` with every column typed as ``pa.string()``.
    Calls ``gc.collect()`` between chunks. Returns total *data* rows written
    (header excluded).
    """
    if not pyarrow_available():
        raise ImportError(
            "stream_excel_to_parquet requires pyarrow and pandas. "
            "Install with: pip install pyarrow pandas"
        )

    import pyarrow as pa
    import pyarrow.parquet as pq
    from openpyxl import load_workbook

    path = str(path)
    parquet_path = str(parquet_path)
    target = _resolve_sheet(path, sheet=sheet)
    Path(parquet_path).parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(path, data_only=True, read_only=True)
    writer = None
    total = 0
    try:
        ws = wb[target]
        row_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(row_iter)
        except StopIteration:
            return 0

        colnames = []
        seen: dict[str, int] = {}
        for i, h in enumerate(header_row):
            name = str(h).strip() if h is not None and str(h).strip() else f"col_{i}"
            if name in seen:
                seen[name] += 1
                name = f"{name}_{seen[name]}"
            else:
                seen[name] = 0
            colnames.append(name)

        schema = pa.schema([(c, pa.string()) for c in colnames])
        writer = pq.ParquetWriter(parquet_path, schema)
        chunk: list[list] = []

        def _flush():
            nonlocal total, chunk
            if not chunk:
                return
            cols = {
                colnames[i]: [_cell_str(r[i]) if i < len(r) else None for r in chunk]
                for i in range(len(colnames))
            }
            table = pa.table(cols, schema=schema)
            writer.write_table(table)
            total += len(chunk)
            chunk = []
            gc.collect()

        for row in row_iter:
            # Pad / trim to header width
            cells = list(row) if row is not None else []
            if len(cells) < len(colnames):
                cells = cells + [None] * (len(colnames) - len(cells))
            elif len(cells) > len(colnames):
                cells = cells[: len(colnames)]
            chunk.append(cells)
            if len(chunk) >= chunk_size:
                _flush()
        _flush()
    finally:
        if writer is not None:
            writer.close()
        wb.close()
    return total


def _cache_path(path: str, sheet: str, cache_dir=None) -> Path:
    src = Path(path).resolve()
    key = hashlib.sha1(f"{src}|{sheet}|{src.stat().st_mtime_ns}".encode()).hexdigest()[:12]
    safe_sheet = "".join(c if c.isalnum() or c in "-_" else "_" for c in sheet)[:40]
    base = Path(cache_dir) if cache_dir else src.parent / ".data_toolkit_cache"
    base.mkdir(parents=True, exist_ok=True)
    return base / f"{src.stem}_{safe_sheet}_{key}.parquet"


def load_via_strategy(path: str, sheet=None, cache_dir=None, strategy=None):
    """Dispatch on strategy and return ``(DataFrame, note)``.

    Requires pandas + pyarrow. Caller should check ``pyarrow_available()`` first.
    """
    import pandas as pd

    path = str(path)
    target = _resolve_sheet(path, sheet=sheet)
    strat = strategy or choose_strategy(path, sheet=target)
    cache = _cache_path(path, target, cache_dir=cache_dir)

    if strat == "direct":
        df = pd.read_excel(path, sheet_name=target, dtype=str, engine="openpyxl")
        df = optimize_dtypes(_coerce_numerics(df))
        return df, f"direct read '{target}', {len(df)} rows (strategy=direct)"

    if strat == "parquet_cache":
        if cache.is_file():
            df = pd.read_parquet(cache)
            df = optimize_dtypes(_coerce_numerics(df))
            return df, (
                f"parquet cache hit '{target}', {len(df)} rows "
                f"(strategy=parquet_cache, cache={cache.name})"
            )
        # dtype=str already yields None for blank cells — do not .astype(str).replace(
        # {"None"/"nan": None}), which would also null real values like "None Corp".
        df = pd.read_excel(path, sheet_name=target, dtype=str, engine="openpyxl")
        df.to_parquet(cache, index=False)
        df = optimize_dtypes(_coerce_numerics(df))
        return df, (
            f"parquet cache miss → wrote '{target}', {len(df)} rows "
            f"(strategy=parquet_cache, cache={cache.name})"
        )

    # stream
    n = stream_excel_to_parquet(path, str(cache), sheet=target)
    df = pd.read_parquet(cache)
    df = optimize_dtypes(_coerce_numerics(df))
    return df, (
        f"streamed '{target}' → parquet, {n} rows "
        f"(strategy=stream, cache={cache.name})"
    )


def _coerce_numerics(df):
    """Best-effort numeric conversion after string-typed Parquet load."""
    import pandas as pd

    out = df.copy()
    for col in out.columns:
        if out[col].dtype != object and not str(out[col].dtype).startswith("string"):
            continue
        converted = pd.to_numeric(out[col], errors="coerce")
        # Only adopt if a meaningful share converted (avoid nuking id-like columns).
        non_null = out[col].notna().sum()
        if non_null == 0:
            continue
        ok = converted.notna().sum()
        if ok / non_null >= 0.8:
            out[col] = converted
    return out


def optimize_dtypes(df):
    """Downcast ints/floats and categorise low-cardinality object columns.

    Prints memory before/after and percent saved. Returns the optimised frame.
    Typical saving on mixed finance exports: 50–80%.
    """
    import pandas as pd

    before = df.memory_usage(deep=True).sum()
    out = df.copy()

    def _is_stringy(s) -> bool:
        if pd.api.types.is_object_dtype(s):
            return True
        # pandas 2.1+ / 3.x may infer Python ``str`` or StringDtype instead of object
        dt = s.dtype
        if isinstance(dt, pd.StringDtype):
            return True
        return str(dt) in ("string", "str", "string[python]", "string[pyarrow]")

    for col in out.columns:
        s = out[col]
        if pd.api.types.is_integer_dtype(s):
            out[col] = pd.to_numeric(s, downcast="integer")
        elif pd.api.types.is_float_dtype(s):
            out[col] = pd.to_numeric(s, downcast="float")
        elif _is_stringy(s):
            n = len(s)
            if n == 0:
                continue
            nunique = s.nunique(dropna=True)
            if nunique / n < 0.5:
                out[col] = s.astype("category")

    after = out.memory_usage(deep=True).sum()
    saved = (1 - after / before) * 100 if before else 0.0
    print(
        f"[optimize_dtypes] memory {before:,} → {after:,} bytes "
        f"({saved:.1f}% saved)"
    )
    return out


def warn_no_pyarrow() -> None:
    warnings.warn(
        "pyarrow/pandas not installed — read_large falling back to a direct openpyxl "
        "read. Large files (10k+ rows) may OOM. Install with: pip install pyarrow pandas",
        UserWarning,
        stacklevel=3,
    )
