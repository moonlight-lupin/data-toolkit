"""Shared data-cleanup engine — deterministic primitives used by `data-tidy` and
`data-extract`. Lives at the toolkit-root `scripts/` (with `ingest.py`/`envcheck.py`);
skills import it by adding `../../scripts` to sys.path (run from the skill directory).

Design: INTENT (asked upfront, in the skill) -> INGEST (see ingest.py) -> PROFILE ->
PROPOSE a recipe -> CONFIRM -> APPLY (here, deterministic + logged) -> REPORT + save recipe.

The transforms here are **deterministic and logged** — the model's job is to read the
profile and propose the recipe, not to rewrite cells. Conversion failures are KEPT RAW and
FLAGGED for human review, never silently nulled. House output: dates DD MMM YYYY; currency
amount + code.

    from dataclean import profile_table, apply_recipe, render_report, write_xlsx
    prof = profile_table(header, rows)                       # understand the mess
    clean_h, clean_rows, log = apply_recipe(raw_rows, recipe)  # source -> declared target
    write_xlsx(clean_h, clean_rows, "clean.xlsx"); print(render_report(log))

DATA HANDLING: this engine runs on your machine and makes no network calls (no cloud OCR, no
external APIs). But the AI agent driving it sends whatever it reads into its context to your AI
provider — so "your data never leaves the machine" is NOT claimed. If the data is sensitive or
confidential business/financial data, keep it on your synced or shared file store and never send
it, or OCR of it, to a third-party tool. See ../PRINCIPLES.md (§ Data handling).
"""

from __future__ import annotations

import re
import sys
import datetime as dt
import unicodedata
from collections import Counter
from decimal import Decimal, InvalidOperation
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

from openpyxl import Workbook, load_workbook  # noqa: F401 (load_workbook used by ingest)

DATE_OUT = "%d %b %Y"  # house format: 13 Jun 2026

# Currency CODE resolution. Note: a BARE "$" is deliberately NOT here — in SG/AU/HK/US
# contexts it is ambiguous (USD? SGD? AUD? HKD?), so it is never silently resolved to USD.
# A disambiguated dollar (US$, S$, A$, HK$, NZ$, C$, R$) and the ISO codes are resolved.
# Longest signs first so "US$" wins over "$" and "S$" isn't shadowed.
CURRENCY_SIGNS = {
    "US$": "USD", "S$": "SGD", "A$": "AUD", "HK$": "HKD", "NZ$": "NZD",
    "C$": "CAD", "R$": "BRL",
    "SGD": "SGD", "USD": "USD", "GBP": "GBP", "JPY": "JPY", "EUR": "EUR",
    "AUD": "AUD", "HKD": "HKD", "NZD": "NZD", "CAD": "CAD", "CHF": "CHF",
    "CNY": "CNY", "RMB": "CNY", "INR": "INR",
    "£": "GBP", "€": "EUR", "¥": "JPY",
}
# Order to probe signs in: multi-char (disambiguated) first, then unambiguous symbols.
_CURRENCY_SIGN_ORDER = ("US$", "S$", "A$", "HK$", "NZ$", "C$", "R$",
                        "SGD", "USD", "GBP", "JPY", "EUR", "AUD", "HKD", "NZD",
                        "CAD", "CHF", "CNY", "RMB", "INR", "£", "€", "¥")
# A bare "$" still SIGNALS "this is money" for type inference, even though its code is unknown.
CURRENCY_DETECT = set(CURRENCY_SIGNS) | {"$"}

_DATE_FORMATS = ["%Y-%m-%d", "%Y%m%d", "%d/%m/%Y", "%d-%m-%Y", "%d-%b-%Y", "%d.%m.%Y",
                 "%d %b %Y", "%d %B %Y", "%b %d, %Y", "%B %d, %Y", "%d/%m/%y",
                 "%Y/%m/%d"]
_AMBIGUOUS = re.compile(r"^\s*(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2,4})\s*$")


# --------------------------------------------------------------------------- #
# Parsers — each returns (value_or_None, note). note != "" means flag it.
# --------------------------------------------------------------------------- #
def _s(v) -> str:
    return "" if v is None else str(v).strip()


def parse_number(v):
    """'£1,234.50' '(500)' '1.2m' '15%' -> Decimal. Returns (Decimal|None, note).

    Uses Decimal (not float) so finance amounts stay exact — no binary-float drift in sums,
    reconciliation or currency tables. Inputs that are already int/Decimal pass through; a
    float is taken via its str() so 0.1 stays 0.1, not 0.1000000000000000055."""
    if isinstance(v, bool):                       # avoid True/False -> 1/0
        return None, f"not a number: {v!r}"
    if isinstance(v, Decimal):
        return v, ""
    if isinstance(v, int):
        return Decimal(v), ""
    if isinstance(v, float):
        return Decimal(str(v)), ""
    s = _s(v)
    if s == "":
        return None, "empty"
    if re.search(r"-?\d{1,3}(\.\d{3})+,\d+", s):
        return None, f"ambiguous European number format: {s!r}"
    neg = s.startswith("(") and s.endswith(")")
    s2 = s.strip("()")
    mult = 1
    low = s2.lower()
    if low.endswith("m"):
        mult, s2 = 1_000_000, s2[:-1]
    elif low.endswith("k"):
        mult, s2 = 1_000, s2[:-1]
    pct = s2.rstrip().endswith("%")
    s2 = re.sub(r"[^\d.\-]", "", s2.replace(",", ""))
    if s2 in ("", "-", ".", "-."):
        return None, f"not a number: {s!r}"
    try:
        n = Decimal(s2) * mult
    except InvalidOperation:
        return None, f"not a number: {s!r}"
    if neg:
        n = -abs(n)
    if pct:
        n /= 100                                  # exact: division by a power of ten
    return n, ""


def parse_currency(v):
    """'£1,000,000' 'S$ 2.5m' -> (Decimal amount, code). Returns ((amount, code)|None, note).

    The code is detected SEPARATELY from the amount so multi-market data keeps its currency
    (never collapsed into a bare number). A disambiguated dollar (US$/S$/A$/HK$/…) and ISO
    codes resolve; a BARE "$" is flagged ambiguous and left unresolved (code=None) unless the
    caller supplies an expected currency — see `_convert`."""
    s = _s(v)
    if s == "":
        return None, "empty"
    code = None
    su = s.upper()
    for sign in _CURRENCY_SIGN_ORDER:
        if sign.upper() in su:
            code = CURRENCY_SIGNS[sign]
            break
    amt, note = parse_number(s)
    if amt is None:
        return None, note
    if code:
        return (amt, code), ""
    if "$" in s:                                   # bare dollar — prefix not recognised above
        return (amt, None), "ambiguous '$' — could be USD/SGD/AUD/HKD…; specify expected currency"
    return (amt, None), "no currency symbol — code unknown"


def _detect_code(v):
    """Normalised ISO code from a stand-alone currency cell ('GBP', 'gbp', '£', 'S$') -> code
    or None. Same vocabulary/order as `parse_currency`, but needs NO amount — for a separate
    Currency column carried alongside the amount (a bare '$' stays unknown, as elsewhere)."""
    su = _s(v).upper()
    if not su:
        return None
    for sign in _CURRENCY_SIGN_ORDER:
        if sign.upper() in su:
            return CURRENCY_SIGNS[sign]
    return None


def _excel_serial(s):
    if re.fullmatch(r"\d{5}(\.\d+)?", s):
        try:
            base = dt.date(1899, 12, 30)
            whole, _, frac = s.partition(".")
            note = "parsed from Excel serial"
            if frac and int(frac) != 0:
                note += "; date-time truncated to date"
            return base + dt.timedelta(days=int(whole)), note
        except (ValueError, OverflowError):
            return None
    return None


def parse_date(v, dayfirst=True):
    """Many formats + Excel serial + datetime -> date. Returns (date|None, note).
    Ambiguous dd/mm vs mm/dd is resolved by `dayfirst` (UK/SG default) and flagged."""
    if isinstance(v, (dt.datetime, dt.date)):
        return (v.date() if isinstance(v, dt.datetime) else v), ""
    s = _s(v)
    if s == "":
        return None, "empty"
    ser = _excel_serial(s)
    if ser:
        return ser
    m = _AMBIGUOUS.match(s)
    note = ""
    if m:
        a, b, _ = (int(x) for x in m.groups())
        if a <= 12 and b <= 12 and a != b:
            note = f"ambiguous date {s!r} — read as {'day/month' if dayfirst else 'month/day'}"
    fmts = _DATE_FORMATS if dayfirst else (["%m/%d/%Y", "%m-%d-%Y", "%m/%d/%y"] + _DATE_FORMATS)
    for fmt in fmts:
        try:
            return dt.datetime.strptime(s, fmt).date(), note
        except ValueError:
            continue
    return None, f"unrecognised date: {s!r}"


# --------------------------------------------------------------------------- #
# Structure
# --------------------------------------------------------------------------- #
def detect_header(rows, scan=15):
    """Best guess at the real header row index: the row (in the first `scan`) with the most
    non-empty, mostly-textual, distinct cells. Junk/title/banner rows score low."""
    best, best_score = 0, -1.0
    for i, row in enumerate(rows[:scan]):
        cells = [_s(c) for c in row]
        filled = [c for c in cells if c]
        if len(filled) < 2:
            continue
        texty = sum(1 for c in filled if not parse_number(c)[0] and not parse_date(c)[0])
        score = len(filled) + texty + 0.5 * len(set(filled)) - 0.3 * (len(cells) - len(filled))
        if score > best_score:
            best, best_score = i, score
    return best


def _is_blank(row):
    return all(_s(c) == "" for c in row)


def _looks_total(row):
    return any(_s(c).lower() in ("total", "totals", "subtotal", "grand total") for c in row)


# --------------------------------------------------------------------------- #
# Profile (read-only) — inspect ONLY against what the target needs is the skill's job;
# this gives the full picture the recipe is proposed from.
# --------------------------------------------------------------------------- #
# Ordered vocabularies for ORDINAL detection. Lowercased; detection is by set-membership,
# the returned order follows the scale. Extend freely — keep each scale internally ordered.
ORDINAL_SCALES = [
    ("low", "medium", "high"),
    ("low", "med", "high"),
    ("low", "moderate", "high", "critical"),
    ("xs", "s", "m", "l", "xl", "xxl", "xxxl"),
    ("extra small", "small", "medium", "large", "extra large"),
    ("poor", "fair", "good", "very good", "excellent"),
    ("very poor", "poor", "average", "good", "very good"),
    ("strongly disagree", "disagree", "neutral", "agree", "strongly agree"),
    ("never", "rarely", "sometimes", "often", "always"),
    ("cold", "warm", "hot"),
    ("bronze", "silver", "gold", "platinum", "diamond"),
    ("q1", "q2", "q3", "q4"),
    ("monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"),
    ("mon", "tue", "wed", "thu", "fri", "sat", "sun"),
]


def match_ordinal(distinct_lower):
    """If the distinct (lowercased) values fit a single ordered scale, return them in scale
    order; else None. Used to tag a column 'ordinal' and to suggest a sort order."""
    vals = {_s(v).lower() for v in distinct_lower if _s(v) != ""}
    if not (2 <= len(vals) <= 12):
        return None
    for scale in ORDINAL_SCALES:
        if vals <= set(scale):
            return [s for s in scale if s in vals]
    return None


# Type CLASS is decided by the fraction of cells matching a parser against a threshold, so a
# bounded, evenly-strided sample estimates that fraction to well within the decision margin —
# and avoids a full-column parse. `parse_date` alone tries ~13 strptime formats per cell, which
# dominates profiling at scale (250k rows × 12 cols ≈ minutes). Strided (not head) sampling so a
# sorted/blocked column can't bias the estimate. Cardinality checks below stay on the full column.
_INFER_SAMPLE = 5000


def _infer_type(values):
    nonempty = [v for v in values if _s(v) != ""]
    if not nonempty:
        return "empty"
    if len(nonempty) > _INFER_SAMPLE:
        step = len(nonempty) // _INFER_SAMPLE
        sample = nonempty[::step][:_INFER_SAMPLE]
    else:
        sample = nonempty
    def frac(fn):
        return sum(1 for v in sample if fn(_s(v))) / len(sample)
    if frac(lambda s: parse_date(s)[0] is not None) > 0.8:
        return "date"
    if frac(lambda s: parse_currency(s)[0] is not None and any(k in s for k in CURRENCY_DETECT)) > 0.5:
        return "currency"
    if frac(lambda s: parse_number(s)[0] is not None) > 0.8:
        return "number"
    if frac(lambda s: s.lower() in ("yes", "no", "true", "false", "y", "n")) > 0.8:
        return "bool"
    # text-like: separate ORDINAL / CATEGORICAL from free text (advisory — informs the recipe)
    distinct_lower = {_s(v).lower() for v in nonempty}
    if match_ordinal(distinct_lower):
        return "ordinal"
    n = len(nonempty)
    distinct = len(distinct_lower)
    # categorical = few distinct values that REPEAT (low cardinality + low distinct/row ratio)
    if n >= 8 and 2 <= distinct <= 20 and distinct <= n / 2:
        return "categorical"
    return "text"


def profile_table(header, rows):
    cols = []
    n = len(rows)
    for j, name in enumerate(header):
        vals = [r[j] if j < len(r) else "" for r in rows]
        nonempty = [v for v in vals if _s(v) != ""]
        seen, samples = set(), []
        for v in nonempty:
            sv = _s(v)
            if sv not in seen:
                seen.add(sv)
                if len(samples) < 4:
                    samples.append(sv)
        cols.append({"name": _s(name) or f"(col {j+1})", "type": _infer_type(vals),
                     "missing_pct": round(100 * (n - len(nonempty)) / n, 1) if n else 0.0,
                     "distinct": len(seen), "samples": samples})
    keyed = ["|".join(_s(c) for c in r) for r in rows]
    dups = len(keyed) - len(set(keyed))
    return {"rows": n, "columns": cols, "duplicate_rows": dups}


def render_profile(prof):
    out = [f"## Profile — {prof['rows']} rows, {len(prof['columns'])} columns",
           f"_duplicate rows: {prof['duplicate_rows']}_", "",
           "| Column | Type | Missing | Distinct | Samples |", "|---|---|---|---|---|"]
    for c in prof["columns"]:
        out.append(f"| {c['name']} | {c['type']} | {c['missing_pct']}% | {c['distinct']} | "
                   f"{'; '.join(c['samples'])} |")
    return "\n".join(out)


# --------------------------------------------------------------------------- #
# Quality / health scoring (A) — grade a dataset BEFORE deciding what to clean, and again
# after, to show the lift. Read-only; deterministic. Surfaces issues with a severity so the
# user acts on the few that matter. (Folds in semantic types (D), hygiene noise (C) and
# standardisation candidates (B).)
# --------------------------------------------------------------------------- #
def _grade(pct):
    return ("A" if pct >= 95 else "B" if pct >= 85 else "C" if pct >= 70
            else "D" if pct >= 50 else "F")


def _type_consistency(nonempty, t):
    """% of non-empty cells that match the column's inferred type. Text-like types score on
    cleanliness (no whitespace/encoding noise) rather than parseability."""
    if not nonempty:
        return 100.0
    if t in ("number", "currency", "date", "bool"):
        fn = {"number": lambda s: parse_number(s)[0] is not None,
              "currency": lambda s: parse_currency(s)[0] is not None,
              "date": lambda s: parse_date(s)[0] is not None,
              "bool": lambda s: _s(s).lower() in ("yes", "no", "true", "false", "y", "n")}[t]
        return round(100 * sum(1 for v in nonempty if fn(_s(v))) / len(nonempty), 1)
    clean = sum(1 for v in nonempty if not _has_noise(v))
    return round(100 * clean / len(nonempty), 1)


def _has_noise(v):
    raw = str(v)
    collapsed = re.sub(r"\s+", " ", raw).strip()
    return raw != collapsed or any(m in raw for m in ("Ã", "â€", "Â"))


def score_quality(header, rows, prof=None):
    """Score a table's health. -> {overall_score, overall_grade, rows, duplicate_rows,
    columns:[{name,type,completeness,completeness_grade,consistency,issues:[{severity,msg}]}],
    issues:[...]} . Issue severities: 'critical' | 'warn' | 'info'."""
    prof = prof or profile_table(header, rows)
    n = prof["rows"]
    cols_out, all_issues, scores = [], [], []
    for j, c in enumerate(prof["columns"]):
        vals = [r[j] if j < len(r) else "" for r in rows]
        nonempty = [v for v in vals if _s(v) != ""]
        completeness = round(100 - c["missing_pct"], 1)
        consistency = _type_consistency(nonempty, c["type"])
        issues = []
        if c["missing_pct"] > 50:
            issues.append(("critical", f"{c['missing_pct']}% missing"))
        elif c["missing_pct"] > 20:
            issues.append(("warn", f"{c['missing_pct']}% missing"))
        if c["type"] in ("number", "currency", "date", "bool") and consistency < 90:
            issues.append(("warn", f"{round(100 - consistency, 1)}% don't parse as {c['type']}"))
        if c["type"] in ("text", "categorical", "ordinal"):
            clusters = propose_value_map(nonempty)
            if clusters:
                sev = "warn" if len(clusters) > 2 else "info"
                issues.append((sev, f"{len(clusters)} value variant cluster(s) could be standardised"))
        noise = sum(1 for v in nonempty if _has_noise(v))
        if noise:
            issues.append(("info", f"{noise} cell(s) with whitespace/encoding noise"))
        cols_out.append({"name": c["name"], "type": c["type"], "completeness": completeness,
                         "completeness_grade": _grade(completeness), "consistency": consistency,
                         "distinct": c["distinct"], "issues": issues})
        all_issues += [{"column": c["name"], "severity": s, "msg": m} for s, m in issues]
        scores.append((completeness + consistency) / 2)
    overall = round(sum(scores) / len(scores), 1) if scores else 0.0
    dup_pct = round(100 * prof["duplicate_rows"] / n, 1) if n else 0.0
    overall = round(max(0.0, overall - min(15.0, dup_pct)), 1)  # penalise duplicate rows
    if prof["duplicate_rows"]:
        all_issues.append({"column": "(table)", "severity": "warn" if dup_pct > 5 else "info",
                           "msg": f"{prof['duplicate_rows']} duplicate row(s) ({dup_pct}%)"})
    return {"overall_score": overall, "overall_grade": _grade(overall), "rows": n,
            "duplicate_rows": prof["duplicate_rows"], "columns": cols_out, "issues": all_issues}


def render_quality_report(q, title="Data quality report"):
    _RANK = {"critical": 0, "warn": 1, "info": 2}
    _ICON = {"critical": "🔴", "warn": "🟠", "info": "🔵"}
    out = [f"# {title}",
           f"**Overall: {q['overall_score']}/100 — grade {q['overall_grade']}**  "
           f"({q['rows']} rows, {len(q['columns'])} columns, {q['duplicate_rows']} duplicate)",
           "", "| Column | Type | Complete | Consistency | Issues |", "|---|---|---|---|---|"]
    for c in q["columns"]:
        top = c["issues"][0][1] if c["issues"] else "—"
        out.append(f"| {c['name']} | {c['type']} | {c['completeness']}% ({c['completeness_grade']}) "
                   f"| {c['consistency']}% | {top} |")
    if q["issues"]:
        out.append(f"\n## Issues ({len(q['issues'])}) — worst first")
        for it in sorted(q["issues"], key=lambda x: _RANK.get(x["severity"], 3)):
            out.append(f"- {_ICON.get(it['severity'], '')} **{it['severity']}** — "
                       f"{it['column']}: {it['msg']}")
    else:
        out.append("\n_No issues detected._")
    return "\n".join(out)


# --------------------------------------------------------------------------- #
# Apply a recipe (deterministic) -> (header, rows, log)
# --------------------------------------------------------------------------- #
def _col_index(header, name):
    exact = [_s(h) for h in header]
    n_raw = _s(name)
    if n_raw in exact:
        return exact.index(n_raw)
    norm = [h.strip().lower() for h in exact]
    n = n_raw.strip().lower()
    if n in norm:
        return norm.index(n)
    matches = [i for i, h in enumerate(norm) if n and (n in h or h in n)]
    if len(matches) == 1:
        return matches[0]
    return None


def apply_recipe(raw_rows, recipe, masters=None):
    """recipe = {header_row?, drop?, columns:[{source,target,type,trim,currency,format}],
    dedup_keys?, validate?:[{col,required?,regex?,in_master?,unique?}]}. masters = {name:set}."""
    masters = masters or {}
    log = {"header": None, "dropped": {}, "transforms": [], "flagged": [],
           "duplicates": [], "validation": [], "rows_in": len(raw_rows), "rows_out": 0}
    if not raw_rows:
        log["header"] = None
        log["dropped"] = {"blank": 0, "totals": 0}
        log["message"] = "empty input: no rows to clean"
        return [], [], log

    hr = recipe.get("header_row")
    if hr is None:
        hr = detect_header(raw_rows)
    log["header"] = hr
    header_src = [_s(c) for c in raw_rows[hr]]
    body = raw_rows[hr + 1:]

    drop = recipe.get("drop", {"blank": True, "totals": True})
    kept = []
    nb = nt = 0
    for r in body:
        if drop.get("blank", True) and _is_blank(r):
            nb += 1
            continue
        if drop.get("totals", True) and _looks_total(r):
            nt += 1
            continue
        kept.append(r)
    log["dropped"] = {"blank": nb, "totals": nt}

    cols = recipe.get("columns")
    if not cols:  # passthrough: keep all source columns as text
        cols = [{"source": h, "target": h, "type": "text"} for h in header_src]
    # A currency column may emit its detected code into a SEPARATE column (multi-market data):
    # set "code_target" on a type:"currency" spec. The code column follows its amount column.
    # `code_source` (optional) names a separate Currency column to read the code from when the
    # amount cell carries no symbol — so an unparseable/symbol-less amount keeps its currency.
    code_tgt = [c.get("code_target") if c.get("type") == "currency" else None for c in cols]
    code_src_idx = [_col_index(header_src, c.get("code_source"))
                    if (c.get("type") == "currency" and c.get("code_source")) else None
                    for c in cols]
    out_header = []
    for c, ct in zip(cols, code_tgt):
        out_header.append(c["target"])
        if ct:
            out_header.append(ct)
    idx = [(_col_index(header_src, c["source"]), c) for c in cols]
    _missing_src = [c["source"] for j, c in idx if j is None]

    out_rows = []
    conv = {c["target"]: {"ok": 0, "flagged": 0} for c in cols}
    for ri, r in enumerate(kept):
        out = []
        for (j, spec), ct, csi in zip(idx, code_tgt, code_src_idx):
            raw = "" if j is None or j >= len(r) else r[j]
            code_src_raw = "" if csi is None or csi >= len(r) else r[csi]
            if j is None:
                conv[spec["target"]]["flagged"] += 1
                log["flagged"].append({"row": ri + 1, "column": spec["target"],
                                       "value": "", "reason": f"source column '{spec['source']}' not found"})
                out.append("")
                if ct:
                    out.append("")
                continue
            val, note, kept_raw = _convert(raw, spec, code_src_raw)  # kept_raw=True only on hard failure
            if _s(raw) != "":
                if not kept_raw:
                    conv[spec["target"]]["ok"] += 1     # value converted (may also warn)
                if note:
                    conv[spec["target"]]["flagged"] += 1
                    log["flagged"].append({"row": ri + 1, "column": spec["target"],
                                           "value": _s(raw), "reason": note})
            out.append(val)
            if ct:                                  # separate currency-code column
                out.append(_currency_code(raw, spec, code_src_raw))
        out_rows.append(out)
    log["transforms"] = [{"column": t, "converted": conv[t]["ok"], "flagged": conv[t]["flagged"]}
                         for t in conv]

    keys = recipe.get("dedup_keys")
    if keys:
        out_rows, dlog = _dedup(out_header, out_rows, keys)
        log["duplicates"] = dlog

    for rule in recipe.get("validate", []):
        log["validation"].extend(_validate_rule(out_header, out_rows, rule, masters))

    log["rows_out"] = len(out_rows)
    return out_header, out_rows, log


# --------------------------------------------------------------------------- #
# String hygiene (opt-in, on the `text` type). Deterministic; every change is logged
# (soft-flagged) so nothing is mangled silently. All stdlib.
# --------------------------------------------------------------------------- #
# Common UTF-8-as-Latin-1 mojibake, as a fallback when the clean round-trip can't be applied.
_MOJIBAKE = {
    "â€™": "'", "â€˜": "'", "â€œ": '"', "â€\x9d": '"', "â€”": "—", "â€“": "–",
    "â€¦": "…", "Ã©": "é", "Ã¨": "è", "Ã ": "à", "Ã¢": "â", "Ã§": "ç", "Ã±": "ñ",
    "Ã¼": "ü", "Ã¶": "ö", "Ã¤": "ä", "Ã‰": "É", "Ã¡": "á", "Ã³": "ó", "Ãº": "ú",
    "Â£": "£", "Â ": " ", "Â": "",
}
# strip_specials default: keep word chars, whitespace and common business punctuation.
_SPECIALS_DEFAULT = re.compile(r"[^\w\s.,;:&@/()\-+%'\"£$€]", re.UNICODE)


def _fix_encoding(s):
    """Repair common mojibake then NFC-normalise. Tries the principled UTF-8/Latin-1
    round-trip first, falls back to a small replacement map."""
    if any(seq in s for seq in ("Ã", "â€", "Â")):
        try:
            s = s.encode("latin-1").decode("utf-8")
        except (UnicodeEncodeError, UnicodeDecodeError):
            for bad, good in _MOJIBAKE.items():
                s = s.replace(bad, good)
    return unicodedata.normalize("NFC", s)


def _text_hygiene(s, spec):
    """Apply opt-in casing / special-strip / encoding fixes. -> (s, note)."""
    notes = []
    if spec.get("fix_encoding"):
        fixed = _fix_encoding(s)
        if fixed != s:
            notes.append("encoding repaired")
        s = fixed
    sp = spec.get("strip_specials")
    if sp:
        pat = _SPECIALS_DEFAULT if sp is True else re.compile(sp)
        stripped = re.sub(r"\s+", " ", pat.sub("", s)).strip()
        if stripped != s:
            notes.append("special chars stripped")
        s = stripped
    case = spec.get("case")
    if case == "lower":
        s = s.lower()
    elif case == "upper":
        s = s.upper()
    elif case == "title":
        s = s.title()
    elif case == "sentence":
        s = (s[:1].upper() + s[1:].lower()) if s else s
    return s, "; ".join(notes)


# --------------------------------------------------------------------------- #
# Categorical value standardisation (B). PROPOSE clusters of inconsistent variants of the
# SAME category and a canonical form; the user CONFIRMS; the confirmed map is baked into the
# recipe as `value_map` and applied here. Never auto-applied without confirmation (HITL).
# --------------------------------------------------------------------------- #
def _fold(s):
    """Aggressive match key: lowercase, strip accents (NFKD), drop non-alphanumerics, collapse
    spaces. 'U.S.A.' / 'usa' / 'U S A' -> 'usa'; 'Café' / 'cafe' -> 'cafe'."""
    s = unicodedata.normalize("NFKD", _s(s))
    s = "".join(c for c in s if not unicodedata.combining(c)).lower()
    return re.sub(r"[^a-z0-9]", "", s)


def propose_value_map(values, master=None):
    """Cluster near-variant values of one column and propose a canonical per cluster.
    Returns clusters sorted by frequency: [{canonical, variants:[surface...], counts, n}].
    Canonical = the matching master entry if `master` is given, else the most frequent surface
    (ties → longest). A cluster is only proposed when there is something to standardise
    (>1 surface form, or a single form that differs from its master canonical).
    PROPOSAL ONLY — show it, get confirmation, then bake the accepted map into the recipe."""
    counts = Counter(_s(v) for v in values if _s(v) != "")
    groups = {}
    for surf, c in counts.items():
        groups.setdefault(_fold(surf), Counter())[surf] += c
    master_lookup = {_fold(m): _s(m) for m in master} if master else {}
    clusters = []
    for fold, surfaces in groups.items():
        mcanon = master_lookup.get(fold)
        canonical = mcanon or sorted(surfaces.items(), key=lambda kv: (-kv[1], -len(kv[0])))[0][0]
        if len(surfaces) > 1 or (mcanon and canonical not in surfaces):
            ordered = [s for s, _ in sorted(surfaces.items(), key=lambda kv: -kv[1])]
            clusters.append({"canonical": canonical, "variants": ordered,
                             "counts": dict(surfaces), "n": sum(surfaces.values()),
                             "from_master": bool(mcanon)})
    return sorted(clusters, key=lambda c: -c["n"])


def render_value_map_proposals(clusters, col_name):
    """Markdown for the user to CONFIRM before applying (like reconcile's aggregation review)."""
    if not clusters:
        return f"_No value-standardisation candidates in **{col_name}**._"
    out = [f"## Proposed value standardisation — **{col_name}** ({len(clusters)} cluster(s))",
           "Confirm before applying — nothing is changed until you accept. Canonical ← variants.",
           "", "| Canonical | Variants (count) | Source |", "|---|---|---|"]
    for c in clusters:
        vs = ", ".join(f"{v} ({c['counts'][v]})" for v in c["variants"])
        out.append(f"| {c['canonical']} | {vs} | {'master' if c['from_master'] else 'most frequent'} |")
    return "\n".join(out)


def value_map_from_clusters(clusters, accepted=None):
    """Turn confirmed clusters into a recipe-ready value_map: {canonical: [variants...]}.
    `accepted` = indices to apply (default all)."""
    idx = range(len(clusters)) if accepted is None else accepted
    return {clusters[i]["canonical"]: clusters[i]["variants"] for i in idx}


def _vmap_norm(spec):
    """Build (and cache) a fold->canonical lookup from spec['value_map'], which may be
    {canonical: [variants]} (readable) or {variant: canonical} (flat)."""
    norm = spec.get("_vmap_norm")
    if norm is not None:
        return norm
    vm = spec.get("value_map") or {}
    norm = {}
    if any(isinstance(v, (list, tuple)) for v in vm.values()):
        for canon, variants in vm.items():
            norm[_fold(canon)] = canon
            for v in variants:
                norm[_fold(v)] = canon
    else:
        for variant, canon in vm.items():
            norm[_fold(variant)] = canon
    spec["_vmap_norm"] = norm
    return norm


def _apply_value_map(s, spec):
    """If the cell matches a confirmed cluster, replace with the canonical. -> (s, note)."""
    if not spec.get("value_map"):
        return s, ""
    mapped = _vmap_norm(spec).get(_fold(s))
    if mapped and mapped != s:
        return mapped, f"standardised → {mapped}"
    return s, ""


def _currency_code(raw, spec, code_source_raw=""):
    """Resolve the currency code for a cell -> code string (or "" if genuinely unknown).
    Priority: a code embedded in the amount cell -> a separate `code_source` column ->
    the recipe's expected `currency`. Independent of whether the AMOUNT parsed, so an
    unparseable amount ('pending', blank) never costs the row its currency."""
    res, _ = parse_currency(raw)
    code = res[1] if res else None
    if not code:
        code = _detect_code(code_source_raw)
    return code or spec.get("currency") or ""


def _convert(raw, spec, code_source_raw=""):
    """-> (value, note, kept_raw). kept_raw=True means a HARD failure (raw kept for review);
    a note with kept_raw=False is a soft WARNING (value converted, but flag it too).
    `code_source_raw` is the cell of a separate Currency column (see `code_source`), used to
    resolve — and to quiet the 'code unknown' warning on — a currency amount with no symbol."""
    if spec.get("trim", True) and isinstance(raw, str):
        raw = raw.strip()
    t = spec.get("type", "text")
    if _s(raw) == "":
        return "", "", False
    if t in ("text", "categorical", "ordinal"):
        s = re.sub(r"\s+", " ", _s(raw))
        s, note = _text_hygiene(s, spec)
        s, vnote = _apply_value_map(s, spec)
        return s, "; ".join(n for n in (note, vnote) if n), False
    if t == "number":
        n, note = parse_number(raw)
        return (n, "", False) if n is not None else (_s(raw), note, True)
    if t == "currency":
        res, note = parse_currency(raw)
        if res is None:
            return _s(raw), note, True
        amt, code = res
        want = spec.get("currency")
        if want:
            if code and code != want:
                note = (note + f"; currency {code} != expected {want}").strip("; ")
            elif code is None:
                note = ""                          # ambiguous/unknown resolved by expected currency
        elif code is None and _detect_code(code_source_raw):
            note = ""                              # code supplied by a separate Currency column
        # if no expected currency and the code is truly unknown, `note` already explains it
        return amt, note, False
    if t == "date":
        d, note = parse_date(raw, dayfirst=spec.get("dayfirst", True))
        if d is None:
            return _s(raw), note, True
        return d.strftime(spec.get("format", DATE_OUT)), note, False
    if t == "bool":
        low = _s(raw).lower()
        if low in ("yes", "true", "y", "1"):
            return True, "", False
        if low in ("no", "false", "n", "0"):
            return False, "", False
        return _s(raw), f"unrecognised bool: {_s(raw)!r}", True
    return _s(raw), "", False


def convert_value(raw, spec, code_source_raw=""):
    """Public wrapper around the deterministic cell converter. `code_source_raw` (optional) is
    a separate Currency cell used to resolve a symbol-less currency amount's code."""
    return _convert(raw, spec, code_source_raw)


def currency_code(raw, spec, code_source_raw=""):
    """Public: the currency code for a cell (amount symbol -> `code_source` cell -> expected
    `currency`), independent of whether the amount parsed. "" when genuinely unknown."""
    return _currency_code(raw, spec, code_source_raw)


def _norm_key(v):
    return re.sub(r"[^a-z0-9]", "", _s(v).lower())


def _dedup(header, rows, keys):
    ki = [_col_index(header, k) for k in keys]
    seen, out, log = {}, [], []
    for ri, r in enumerate(rows):
        sig = tuple(_s(r[i]) if i is not None and i < len(r) else "" for i in ki)
        fuzzy = tuple(_norm_key(s) for s in sig)
        if sig in seen:
            log.append({"row": ri + 1, "kind": "exact", "key": " | ".join(sig),
                        "kept_row": seen[sig]})
            continue
        # flag fuzzy near-dup but KEEP it (never auto-merge)
        for s2, kr in seen.items():
            if tuple(_norm_key(x) for x in s2) == fuzzy and s2 != sig:
                log.append({"row": ri + 1, "kind": "possible (review)",
                            "key": " | ".join(sig), "kept_row": kr})
                break
        seen[sig] = ri + 1
        out.append(r)
    return out, log


def _validate_rule(header, rows, rule, masters):
    j = _col_index(header, rule["col"])
    if j is None:
        return [{"col": rule["col"], "issue": "column not found"}]
    fails = []
    seen = set()
    try:
        rgx = re.compile(rule["regex"]) if rule.get("regex") else None
    except re.error as e:
        return [{"col": rule["col"], "issue": f"invalid regex {rule['regex']!r}: {e}"}]
    master = masters.get(rule.get("in_master"))
    master_norm = {_norm_key(m) for m in master} if master is not None else None
    for ri, r in enumerate(rows):
        v = _s(r[j]) if j < len(r) else ""
        if rule.get("required") and v == "":
            fails.append({"row": ri + 1, "col": rule["col"], "issue": "required, empty"})
        if v and rgx and not rgx.search(v):
            fails.append({"row": ri + 1, "col": rule["col"], "issue": f"fails regex {rule['regex']}"})
        if v and master_norm is not None and _norm_key(v) not in master_norm:
            fails.append({"row": ri + 1, "col": rule["col"], "issue": "not in master list"})
        if rule.get("unique"):
            if v in seen and v:
                fails.append({"row": ri + 1, "col": rule["col"], "issue": "duplicate (not unique)"})
            seen.add(v)
    return fails


# --------------------------------------------------------------------------- #
# Output + report
# --------------------------------------------------------------------------- #
def write_xlsx(header, rows, out_path, sheet="Clean"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append([_s(h) for h in header])
    for r in rows:
        ws.append([(_s(c) if isinstance(c, str) else c) for c in r])
    wb.save(out_path)
    return out_path


def render_report(log):
    out = ["# Data-tidy change report",
           f"- Rows in: **{log['rows_in']}** → out: **{log['rows_out']}**",
           f"- Header row used: index {log['header']}",
           f"- Dropped: {log['dropped'].get('blank', 0)} blank, "
           f"{log['dropped'].get('totals', 0)} total/subtotal"]
    if log.get("message"):
        out.append(f"- Note: {_md_escape(log['message'])}")
    if log["duplicates"]:
        ex = sum(1 for d in log["duplicates"] if d["kind"] == "exact")
        pos = len(log["duplicates"]) - ex
        out.append(f"- Duplicates: {ex} exact removed, {pos} possible (kept, flagged for review)")
    out.append("\n## Column transforms")
    out.append("| Column | Converted | Flagged |\n|---|---|---|")
    for t in log["transforms"]:
        out.append(f"| {_md_escape(t['column'])} | {t['converted']} | {t['flagged']} |")
    if log["flagged"]:
        out.append(f"\n## ⚑ Cells flagged for review ({len(log['flagged'])}) — "
                   "hard failures kept raw; warnings converted, please verify")
        out.append("| Row | Column | Value | Reason |\n|---|---|---|---|")
        for f in log["flagged"][:50]:
            out.append(f"| {f['row']} | {_md_escape(f['column'])} | {_md_escape(f['value'])} | {_md_escape(f['reason'])} |")
        if len(log["flagged"]) > 50:
            out.append(f"_…and {len(log['flagged']) - 50} more_")
    if log["validation"]:
        out.append(f"\n## ⚑ Validation failures ({len(log['validation'])})")
        out.append("| Row | Column | Issue |\n|---|---|---|")
        for v in log["validation"][:50]:
            out.append(f"| {v.get('row', '-')} | {_md_escape(v.get('col', '-'))} | {_md_escape(v['issue'])} |")
    return "\n".join(out)


def _md_escape(v):
    return str("" if v is None else v).replace("|", "\\|").replace("\r\n", "<br>").replace("\n", "<br>")


# --------------------------------------------------------------------------- #
# Reuse: emit a runner (.py) + card (.md) for a recurring doc type, so the next
# agent re-runs the same extraction/cleanup with ~no reasoning tokens. The card
# is read FIRST (verify the source still matches) before the .py is executed.
# --------------------------------------------------------------------------- #
_RUNNER_TEMPLATE = '''#!/usr/bin/env python
"""data-@@KIND@@ reusable runner — @@TITLE@@
Doc type: @@DOCTYPE@@.  Generated @@DATE@@ by data-@@KIND@@.

REUSE — do NOT blind-run on new files. FIRST read the companion card "@@CARD@@", inspect the
source, and confirm it still matches the baked spec. This runner self-checks and WARNS on
mismatch (missing fields / new columns), but a human/agent should verify first. If the
layout changed (e.g. a column was added), update SPEC or regenerate via data-@@KIND@@.

Usage:  python @@PYNAME@@ <doc1> [doc2 ...] [-o out.xlsx]
"""
import sys
import pathlib

MODE = "@@KIND@@"
SPEC = @@SPEC_REPR@@
EXPECTED = @@EXPECTED_REPR@@

# The engine (dataclean.py / ingest.py / extract.py) is deployed ALONGSIDE this runner in the
# working folder, so this folder is self-contained — no dependency on the plugin install.
sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))
try:
    import dataclean   # noqa: E402
    import ingest      # noqa: E402
    import extract     # noqa: E402
except ModuleNotFoundError as e:
    raise SystemExit("Engine file missing next to this runner (%s). This folder should also "
                     "contain dataclean.py, ingest.py and extract.py — regenerate the runner "
                     "via data-extract / data-tidy." % e.name)


def _args(argv):
    out, paths, it = "extracted.xlsx", [], iter(argv)
    for a in it:
        if a == "-o":
            out = next(it)
        else:
            paths.append(a)
    if not paths:
        raise SystemExit("Usage: python @@PYNAME@@ <doc...> [-o out.xlsx]")
    return paths, out


def main(argv):
    paths, out = _args(argv)
    if MODE == "extract":
        records, flags_all = [], []
        for p in paths:
            text, _ = ingest.read_text(p)
            low = text.lower()
            missing = [f["name"] for f in SPEC
                       if not any(l.lower() in low for l in f.get("labels", [f["name"]]))]
            if missing:
                print("[WARN] %s: expected fields not found -> %s; source may have changed "
                      "- verify before trusting output." % (p, missing))
            rec, flags = extract.extract_fields(p, SPEC)
            records.append(rec)
            flags_all.append(flags)
        header, rows = extract.fields_to_table(records, [f["name"] for f in SPEC])
        dataclean.write_xlsx(header, rows, out)
        print(extract.render_fields_report(records, flags_all))
    else:
        rows, note = ingest.read_any(paths[0])
        print("ingest:", note)
        hi = SPEC.get("header_row")
        if hi is None:
            hi = dataclean.detect_header(rows)
        srccols = [str(c).strip() for c in rows[hi]]
        exp = EXPECTED.get("columns", [])
        miss = [c for c in exp if c not in srccols]
        new = [c for c in srccols if c and c not in exp]
        if miss:
            print("[WARN] %s: expected columns missing -> %s; verify before trusting output."
                  % (paths[0], miss))
        if new:
            print("[WARN] %s: NEW columns not in the recipe -> %s; a column may have been "
                  "added - verify / regenerate." % (paths[0], new))
        header, clean, log = dataclean.apply_recipe(rows, SPEC)
        dataclean.write_xlsx(header, clean, out)
        print(dataclean.render_report(log))
    print("\\nWrote %s" % out)


if __name__ == "__main__":
    main(sys.argv[1:])
'''


def _build_card(mode, doctype, title, date, pyname, spec):
    L = [f"# data-{mode} runner — {title}", "",
         f"**Doc type:** {doctype}  ", f"**Generated:** {date} (by data-{mode})  ",
         f"**Runner:** `{pyname}`", "",
         "## Before you run it — DO THIS FIRST",
         "1. **Read this card.**",
         "2. **Inspect the current source file(s)** (open / preview them).",
         "3. **Compare to *Expected source* below** — layouts drift (a column/field added, "
         "renamed or moved).",
         f"4. If it **matches**, run the runner (cheap, deterministic). If it **changed**, "
         f"update the baked spec or **regenerate** via data-{mode} — do NOT blind-run a "
         "stale spec.", "", "## Expected source (verify this)"]
    if mode == "extract":
        L += ["Fields pulled (output ← labels searched):", "",
              "| Output field | Type | Source labels |", "|---|---|---|"]
        L += [f"| {f['name']} | {f.get('type', 'text')} | {', '.join(f.get('labels', [f['name']]))} |"
              for f in spec]
        outcols = [f["name"] for f in spec]
    else:
        L += [f"Source columns the recipe maps (header row {spec.get('header_row', 'auto')}):", "",
              "| Source column | → Output | Type |", "|---|---|---|"]
        L += [f"| {c.get('source')} | {c.get('target')} | {c.get('type', 'text')} |"
              for c in spec.get("columns", [])]
        outcols = [c.get("target") for c in spec.get("columns", [])]
    L += ["", "## Output",
          f"A clean `.xlsx` with columns: {', '.join(str(c) for c in outcols)}; plus a flag/change report.",
          "", "## Usage", "```", f"python {pyname} <doc1> [doc2 ...] [-o out.xlsx]", "```",
          "The runner self-checks and warns on mismatch (missing fields / new columns), but verify first.",
          "", "## Self-contained folder",
          "`dataclean.py`, `ingest.py` and `extract.py` are bundled in this folder alongside the "
          "runner — it runs standalone, with nothing else installed. Keep them together (e.g. on "
          "your synced or shared file store). Regenerating refreshes these engine copies.",
          "", "## Data handling",
          "This runner is plain Python and makes no network calls — running it sends your data "
          "nowhere. (An AI agent that reads the data to build or check the recipe does send what it "
          "reads to your AI provider.) See the toolkit's `PRINCIPLES.md` (§ Data handling).",
          "", "## Maintenance", "Regenerate if the source layout changes materially."]
    return "\n".join(L) + "\n"


ENGINE_FILES = ("dataclean.py", "ingest.py", "extract.py")


def emit_runner(out_dir, doctype, mode, spec, title=None, date=None):
    """Write a SELF-CONTAINED reuse bundle for a recurring doc type into out_dir (the user's
    working folder): the runner (.py), the card (.md), AND a copy of the engine
    (dataclean/ingest/extract) so the runner imports from its own folder — no dependency on
    the plugin install. mode 'extract' (spec = field dicts) or 'tidy' (spec = recipe dict)."""
    import shutil
    assert mode in ("extract", "tidy"), mode
    out = Path(out_dir).resolve()
    out.mkdir(parents=True, exist_ok=True)
    title = title or doctype
    date = date or dt.date.today().strftime(DATE_OUT)
    safe = re.sub(r"[^A-Za-z0-9._-]+", "-", doctype).strip("-") or "doc"
    pyname, cardname = f"{mode}_{safe}.py", f"{mode}_{safe}.md"
    if mode == "extract":
        expected = {"fields": [{"name": f["name"], "labels": f.get("labels", [f["name"]])}
                               for f in spec]}
    else:
        expected = {"columns": [c.get("source") for c in spec.get("columns", [])],
                    "header_row": spec.get("header_row")}
    # deploy the engine into the working folder (skip if out_dir IS the engine source dir)
    engine_src = Path(__file__).resolve().parent
    deployed = []
    if out != engine_src:
        for f in ENGINE_FILES:
            src = engine_src / f
            if src.exists():
                shutil.copy2(src, out / f)
                deployed.append(f)
    py = (_RUNNER_TEMPLATE
          .replace("@@KIND@@", mode).replace("@@TITLE@@", title)
          .replace("@@DOCTYPE@@", doctype).replace("@@DATE@@", date)
          .replace("@@PYNAME@@", pyname).replace("@@CARD@@", cardname)
          .replace("@@SPEC_REPR@@", repr(spec)).replace("@@EXPECTED_REPR@@", repr(expected)))
    (out / pyname).write_text(py, encoding="utf-8")
    (out / cardname).write_text(_build_card(mode, doctype, title, date, pyname, spec),
                                encoding="utf-8")
    return {"runner": str(out / pyname), "card": str(out / cardname), "engine": deployed}


if __name__ == "__main__":
    raw = [
        ["ABC Capital — investor commitments", "", ""],
        ["CONFIDENTIAL", "", ""],
        ["Inv. Name", "Commit", "Close"],
        ["Acme Pension  ", "£1,000,000", "12/06/2026"],
        ["Beta Family Office", "S$ 2.5m", "2026-06-13"],
        ["acme pension", "£1,000,000", "5 Jun 2026"],          # fuzzy dup of Acme
        ["Gamma Trust", "not provided", "soon"],               # two bad cells -> flagged
        ["", "", ""],                                          # blank
        ["TOTAL", "£3,500,000", ""],                           # totals row
    ]
    recipe = {
        "columns": [
            {"source": "Inv. Name", "target": "Investor", "type": "text"},
            {"source": "Commit", "target": "Commitment", "type": "currency", "currency": "GBP"},
            {"source": "Close", "target": "Close date", "type": "date"},
        ],
        "dedup_keys": ["Investor"],
        "validate": [{"col": "Investor", "required": True}],
    }
    h, rows, log = apply_recipe(raw, recipe)
    print(render_profile(profile_table(["Inv. Name", "Commit", "Close"], raw[3:])))
    print("\n" + render_report(log))
    out = Path(__file__).resolve().parent / "_selftest_clean.xlsx"
    write_xlsx(h, rows, out)
    print("\n[self-test] clean rows:", rows)
    out.unlink()

    # --- D: semantic type detection (categorical / ordinal) ---
    sizes = ["S", "M", "L", "M", "S", "XL", "L", "M", "S", "L"]
    statuses = ["Open", "Closed", "Open", "Pending", "Open", "Closed",
                "Pending", "Open", "Closed", "Open"]
    print("\n[self-test] infer ordinal   :", _infer_type(sizes))
    print("[self-test] infer categorical:", _infer_type(statuses))
    print("[self-test] ordinal order    :", match_ordinal(set(sizes)))
    assert _infer_type(sizes) == "ordinal", _infer_type(sizes)
    assert _infer_type(statuses) == "categorical", _infer_type(statuses)
    # Large columns take the strided-sample path (>_INFER_SAMPLE): classification must be
    # unchanged. Guards the profiling speed fix (parser-frac on a sample, not the full column).
    big_n = _INFER_SAMPLE * 3
    assert _infer_type([f"{i%28+1:02d}/{i%12+1:02d}/2025" for i in range(big_n)]) == "date"
    assert _infer_type([f"{i}.50" for i in range(big_n)]) == "number"
    assert _infer_type([("Open" if i % 2 else "Closed") for i in range(big_n)]) == "categorical"

    # --- C: string hygiene (encoding / case / specials) ---
    v, note, _ = _convert("Beta Famille FranÃ§aise", {"type": "text", "fix_encoding": True})
    print("[self-test] fix_encoding     :", repr(v), "|", note)
    v2, n2, _ = _convert("  ACME   pension!!  ", {"type": "text", "case": "title",
                                                   "strip_specials": True})
    print("[self-test] case+strip       :", repr(v2), "|", n2)
    assert "ç" in v, v
    assert v2 == "Acme Pension", v2
    print("[self-test] D + C passed.")

    # --- B: categorical value standardisation (propose -> confirm -> apply) ---
    countries = ["USA", "U.S.A.", "USA", "usa", "United States", "Canada", "canada"]
    clusters = propose_value_map(countries)
    print("\n" + render_value_map_proposals(clusters, "Country"))
    assert clusters[0]["canonical"] == "USA", clusters[0]
    vmap = value_map_from_clusters(clusters)            # user confirmed all
    sv, svnote, _ = _convert("u.s.a.", {"type": "categorical", "value_map": vmap})
    print("[self-test] standardise      :", repr(sv), "|", svnote)
    assert sv == "USA", sv
    clm = propose_value_map(["united states", "UNITED STATES", "United States"],
                            master=["United States"])
    assert clm[0]["canonical"] == "United States" and clm[0]["from_master"], clm

    # --- A: quality / health report ---
    qh = ["Amount", "Country", "Notes"]
    qr = [["1,000", "USA", "ok"], ["2,500", "U.S.A.", "FranÃ§aise"], ["3.2m", "usa", "fine"],
          ["soon", "United States", "ok  "], ["4,000", "Canada", "ok"], ["5,000", "canada", "x"],
          ["6,000", "USA", "ok"], ["7,000", "usa", "ok"], ["8,000", "Canada", "ok"], ["", "", ""]]
    q = score_quality(qh, qr)
    print("\n" + render_quality_report(q))
    assert q["overall_grade"] in ("A", "B", "C", "D", "F")
    assert q["columns"][0]["consistency"] < 100         # 'soon' isn't a number
    assert any("standardis" in i["msg"] or "cluster" in i["msg"] for i in q["issues"])
    print("\n[self-test] B + A passed.")

    # --- Decimal amounts (exact, no binary-float drift) ---
    n1, _ = parse_number("0.1")
    n2, _ = parse_number("0.2")
    assert isinstance(n1, Decimal) and (n1 + n2) == Decimal("0.3"), (n1, n2, n1 + n2)
    assert parse_number("1.2m")[0] == Decimal("1200000.0"), parse_number("1.2m")
    assert parse_number("15%")[0] == Decimal("0.15"), parse_number("15%")
    assert parse_number("(500)")[0] == Decimal("-500"), parse_number("(500)")
    cval, _, _ = _convert("£1,234.50", {"type": "currency", "currency": "GBP"})
    assert cval == Decimal("1234.50"), cval

    # --- currency: bare '$' is ambiguous (not auto-USD); disambiguated/expected resolve ---
    (amt, code), note = parse_currency("$1,000")
    assert code is None and "ambiguous" in note, (code, note)
    assert parse_currency("S$ 2.5m")[0] == (Decimal("2500000.0"), "SGD"), parse_currency("S$ 2.5m")
    assert parse_currency("US$ 50")[0][1] == "USD"
    # expected currency resolves the bare-$ ambiguity (note cleared)
    v3, n3, _ = _convert("$1,000", {"type": "currency", "currency": "SGD"})
    assert v3 == Decimal("1000") and n3 == "", (v3, n3)
    assert _infer_type(["$100", "$2,000", "$3,500"]) == "currency"   # bare $ still reads as money

    # --- currency code_target: emit the code into its own column ---
    raw_cur = [["Amount"], ["S$ 1,000"], ["US$ 2,000"], ["$ 3,000"]]
    rec_cur = {"columns": [{"source": "Amount", "target": "Amount", "type": "currency",
                            "currency": "SGD", "code_target": "Currency"}]}
    hc, rc, _ = apply_recipe(raw_cur, rec_cur)
    assert hc == ["Amount", "Currency"], hc
    assert [r[1] for r in rc] == ["SGD", "USD", "SGD"], rc   # bare $ -> expected SGD
    assert rc[0][0] == Decimal("1000"), rc[0]

    # --- currency code_source: a separate Currency column is carried through even when the
    # amount is UNPARSEABLE ('pending') or blank — an unparseable amount costs the row its
    # amount, never its currency ---
    raw_cs = [["Amount", "Currency"], ["£1,234.50", "GBP"], ["pending", "GBP"], ["", "GBP"]]
    rec_cs = {"columns": [
        {"source": "Amount", "target": "Amount", "type": "currency",
         "code_target": "Currency", "code_source": "Currency"},
    ]}
    hcs, rcs, lcs = apply_recipe(raw_cs, rec_cs)
    assert hcs == ["Amount", "Currency"], hcs
    assert [r[1] for r in rcs] == ["GBP", "GBP", "GBP"], rcs   # code preserved on every row
    assert rcs[0][0] == Decimal("1234.50") and rcs[1][0] == "pending", rcs
    assert sum(1 for f in lcs["flagged"] if f["column"] == "Amount") == 1, lcs["flagged"]  # only 'pending'
    # symbol-less amount + code_source resolves the code and quiets the "unknown code" note
    v_cs, n_cs, _ = _convert("1,500.00", {"type": "currency"}, "GBP")
    assert v_cs == Decimal("1500.00") and n_cs == "", (v_cs, n_cs)
    assert currency_code("1,500.00", {"type": "currency"}, "GBP") == "GBP"
    print("[self-test] Decimal + currency passed.")
